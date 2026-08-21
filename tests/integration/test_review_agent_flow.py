"""Semantic-review path: M4 assets -> Agent route -> fail-closed comparison gate."""

from __future__ import annotations

import json
from collections import deque
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from award_audit.agent import loop as loop_mod
from award_audit.agent.harness.models import CaseSeed, HarnessOutcome
from award_audit.agent.harness.persistence import CaseRepository
from award_audit.agent.investigation import InvestigationAgent
from award_audit.agent.review_agent.readers import M4AssetReader
from award_audit.agent.review_agent.runner import SemanticReviewRunner
from award_audit.agent.review_agent.service import ReviewAgent
from award_audit.agent.review_workflow import (
    prepare_review_batch,
    run_audit_stage,
    run_queued_review_cases,
)
from award_audit.agent.toolkit.contracts import (
    EvidenceAssetRecord,
    ToolBudgetLimits,
    ToolObservation,
)
from award_audit.agent.toolkit.registry import ToolRegistry
from award_audit.agent.toolkit.safety import inspect_evidence_file
from award_audit.agent.toolkit.web import Attachment, PageContent
from award_audit.core.pipeline.importer import import_file
from award_audit.core.pipeline.store import Store
from award_audit.core.reference.ledger import LedgerEntry
from award_audit.core.reference.resource_map import ResourceMapEntry
from award_audit.core.reference.template_registry import build_template_spec


class FakeLlm:
    def __init__(self, responses: list[object]) -> None:
        self._responses = deque(responses)

    def json_call(self, _system: str, _user: str, *, max_tokens: int) -> Any:
        assert max_tokens in {1200, 4000}
        return self._responses.popleft()


class CompareReadyInvestigationLlm:
    def json_call(self, _system: str, _user: str, *, max_tokens: int) -> Any:
        assert max_tokens == 1400
        return {
            "kind": "compare",
            "reason": "M4 has a verified local roster asset for semantic routing.",
        }


def test_semantic_runner_starts_supplement_with_fresh_runtime_budget(
    tmp_path, monkeypatch
) -> None:  # noqa: ANN001
    store = Store(tmp_path / "semantic-supplement-budget.db")
    batch_id = store.create_batch("semantic-supplement-budget")
    repository = CaseRepository(store)
    state, _ = repository.create_or_get(CaseSeed(
        batch_id=batch_id,
        resource_code="06020007",
        award_name="中国专利奖",
        year="2025",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="重新核验官方名单",
    ))
    state.status = "waiting_human"
    state.budget.calls = 23
    state.budget.asset_calls = 14
    state.step_count = 68
    state.elapsed_ms = 311_844
    state.tool_trace = [ToolObservation(
        call_id="old-attempt",
        tool_name="extract_pdf_text",
        started_at="2026-08-12T00:00:00Z",
        finished_at="2026-08-12T00:00:01Z",
        duration_ms=1000,
        ok=True,
    )]
    repository.save(state)
    state = repository.request_supplement(
        state.case_id,
        "使用修复后的批处理重新核验",
        expected_version=state.state_version,
    )
    runner = SemanticReviewRunner(
        repository,
        review_llm=CompareReadyInvestigationLlm(),
        investigation_agent=InvestigationAgent(
            CompareReadyInvestigationLlm(),
            ToolRegistry(),
            allowed_roots=[tmp_path],
        ),
        allowed_roots=[tmp_path],
        tool_limits=ToolBudgetLimits(max_calls=24),
    )
    captured: dict[str, object] = {}

    def fake_run_langgraph_case(**kwargs):  # noqa: ANN003, ANN202
        current = kwargs["state"]
        captured.update({
            "calls": current.budget.calls,
            "asset_calls": current.budget.asset_calls,
            "step_count": current.step_count,
            "elapsed_ms": current.elapsed_ms,
            "trace_count": len(current.tool_trace),
            "pending_supplement": current.pending_supplement,
        })
        return HarnessOutcome(state=current, stopped_reason="captured")

    monkeypatch.setattr(runner, "_run_langgraph_case", fake_run_langgraph_case)
    outcome = runner.run(state.case_id)

    assert outcome.stopped_reason == "captured"
    assert captured == {
        "calls": 0,
        "asset_calls": 0,
        "step_count": 0,
        "elapsed_ms": 0,
        "trace_count": 0,
        "pending_supplement": "",
    }
    attempt = store.list_audit_attempts(state.case_id)[-1]
    assert attempt["kind"] == "supplement"
    assert attempt["supplement_request"] == "使用修复后的批处理重新核验"
    store.close()


def test_semantic_runner_persists_llm_route_before_comparison(tmp_path) -> None:  # noqa: ANN001
    store = Store(tmp_path / "review-agent-flow.db")
    batch_id = store.create_batch("review-agent-flow")
    source_url = "https://official.example/result.xlsx"
    path = tmp_path / "official-list.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "获奖名单"
    sheet.append(["作品名称", "单位名称"])
    sheet.append(["项目甲", "甲大学"])
    workbook.save(path)
    inspection = inspect_evidence_file(
        path,
        max_bytes=20 * 1024 * 1024,
        allowed_kinds={"xlsx", "xls"},
    )
    asset = EvidenceAssetRecord(
        url=source_url,
        parent_url="https://official.example/notice",
        label="2026 年获奖名单",
        kind="unknown",
        status="parsed",
        local_path=str(path),
        sha256=inspection.sha256,
        content_type=inspection.content_type,
        metadata={
            "title": "2026 年示例奖获奖名单",
            "sample_rows": [["作品", "单位"], ["项目甲", "甲大学"]],
            "anchors": ["Sheet1!A1:B2"],
        },
    )
    report = {
        "resource_code": "04050014",
        "award_name": "示例奖",
        "year": "2026",
        "verdict": "无法核对",
        "confidence": "low",
        "source_url": "https://official.example/notice",
        "source_urls": ["https://official.example/notice"],
        "submitted_count": 1,
        "evidence_assets": [asset.model_dump(mode="json")],
    }
    result_id = store.add_audit_results(batch_id, [report])[0]
    stage_item = store.claim_stage_item(batch_id, "04050014", "2026", worker="test")
    assert stage_item is not None
    store.finish_stage_item(
        batch_id,
        "04050014",
        "2026",
        status="done",
        current_result_id=result_id,
        worker="test",
        expected_version=int(stage_item["state_version"]),
    )
    stage = store.claim_batch_stage(batch_id, "m4", worker="test")
    assert stage is not None
    store.finish_batch_stage(
        batch_id,
        "m4",
        "done",
        worker="test",
        expected_version=int(stage["state_version"]),
    )
    state, _ = CaseRepository(store).create_or_get(CaseSeed(
        batch_id=batch_id,
        origin_m4_result_id=result_id,
        resource_code="04050014",
        award_name="示例奖",
        year="2026",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="验证示例奖名单来源。",
        submitted_summary={
            "submitted_rows": 1,
            "row_conservation": {
                "total_rows": 1,
                "assigned_rows": 1,
                "ambiguous_rows": 0,
                "unassigned_rows": 0,
            },
            "role_scopes": [{
                "scope_key": "work_or_project:all",
                "role_type": "work_or_project",
                "role_label": "项目/成果",
                "required": True,
                "profile": {"primary_alternatives": [["ZPMC"]]},
                "business_scope": {"year": "2026"},
                "submitted_row_count": 1,
                "submitted_identity_count": 1,
                "submitted_identities": {"project-1": "项目甲"},
            }],
        },
        known_urls=["https://official.example/notice"],
    ))
    assert state.case_id > 0
    # Case snapshots keep submitted scope facts, while scope IDs belong to the
    # durable audit_scope ledger. The runner must bridge that boundary itself.
    assert "scope_id" not in state.submitted_summary["role_scopes"][0]

    asset_id = f"sha256:{inspection.sha256}"
    llm = FakeLlm([
        {
            "requests": [{
                "asset_id": asset_id,
                "subunit_id": "document",
                "content_kind": "spreadsheet_sheet",
                "reason": "确认表头和名单样例。",
            }],
            "reason": "需要查看工作表。",
        },
        {
            "case_recommendation": "compare",
            "assessments": [{
                "asset_id": asset_id,
                "scope_ids": [1],
                "role": "project",
                "material_relation": "primary",
                "version_relation": "same",
                "roster_contribution": "include",
                "confidence": 0.99,
                "reason": "标题、年度和表格样例均对应目标名单。",
            }],
            "selected_assets": [asset_id],
            "version_groups": [{
                "key": "2026-project",
                "asset_ids": [asset_id],
                "merge_allowed": True,
                "reason": "单一主名单。",
            }],
            "reason": "可进入本地逐行比较。",
        },
    ])

    outcome = SemanticReviewRunner(
        CaseRepository(store),
        ReviewAgent(
            llm,
            M4AssetReader({asset_id: asset}, allowed_roots=[tmp_path]),
        ),
        investigation_agent=InvestigationAgent(
            CompareReadyInvestigationLlm(),
            ToolRegistry(),
            allowed_roots=[tmp_path],
        ),
        allowed_roots=[tmp_path],
    ).run(state.case_id)

    assert outcome.stopped_reason == "review_agent_comparison_complete"
    assert outcome.state.latest_verification is not None
    assert outcome.state.latest_verification.coverage_complete == "yes"
    assert outcome.state.latest_verification.source_authority == "official"
    routes = store.list_evidence_asset_routes(state.case_id)
    assert len(routes) == 1
    assert routes[0]["route_source"] == "llm"
    assert routes[0]["route_status"] == "routed"
    attempt = store.list_audit_attempts(state.case_id)[-1]
    assert attempt["conclusion_readiness"] == "ready_for_human"
    comparisons = store.list_scope_comparisons(
        state.case_id, attempt_id=int(attempt["attempt_id"])
    )
    assert len(comparisons) == 1 and comparisons[0]["status"] == "complete"
    assert comparisons[0]["comparison_result"] == "matched"
    snapshot = store.get_audit_case_snapshot(state.case_id)
    assert snapshot is not None
    traces = snapshot["tool_trace"]
    graph_names = [
        trace["tool_name"] for trace in traces
        if trace["tool_name"].startswith("langgraph:")
    ]
    assert graph_names[:3] == [
        "langgraph:prepare_case",
        "langgraph:retrieve_memory",
        "langgraph:semantic_plan",
    ]
    assert graph_names[-6:] == [
        "langgraph:semantic_route_assets",
        "langgraph:build_exact_matches_and_candidates",
        "langgraph:semantic_adjudicate_identities",
        "langgraph:deterministic_verify",
        "langgraph:persist",
        "langgraph:waiting_human",
    ]
    review_trace = next(trace for trace in traces if trace["tool_name"] == "review_asset_relations")
    assert review_trace["input_summary"]["request_count"] == 1
    assert len(review_trace["input_summary"]["case_packet_sha256"]) == 64
    assert review_trace["output_summary"]["validation_status"] == "accepted"
    assert review_trace["output_summary"]["raw_json_summary"]["selected_assets"] == [asset_id]
    assert snapshot["budget"]["calls"] == 3
    assert snapshot["budget"]["asset_calls"] == 1
    assert snapshot["step_count"] == 9
    assert snapshot["elapsed_ms"] > 0
    assert attempt["elapsed_ms"] == snapshot["elapsed_ms"]
    store.close()


def test_import_m4_discovery_then_semantic_review_to_verifier(
    tmp_path: Path,
    monkeypatch,
) -> None:
    submission_dir = tmp_path / "submission"
    submission_dir.mkdir()
    submission_path = submission_dir / "CON_GG_XK_RCPY_GXDJSCGR-示例奖-2026.xlsx"
    submission_book = Workbook()
    submission_sheet = submission_book.active
    submission_sheet.title = "数据"
    submission_sheet.append(["ZYLBM", "ZYLB", "XMMC", "XRYXM", "ND"])
    submission_sheet.append(["资源项码", "资源项", "项目名称", "获奖人", "年度"])
    submission_sheet.append(["04059999", "示例奖", "项目甲", "张三", "2026"])
    submission_book.save(submission_path)
    imported = import_file(submission_path, submission_dir.name)
    spec = build_template_spec(
        imported.claimed_table_code,
        imported.sheet_name,
        imported.header_codes,
        imported.header_names,
    )
    official_path = tmp_path / "official-list.xlsx"
    official_book = Workbook()
    official_sheet = official_book.active
    official_sheet.title = "获奖名单"
    official_sheet.append(["项目名称", "获奖人"])
    official_sheet.append(["项目甲", "张三"])
    official_book.save(official_path)
    official_url = "https://official.example/assets/2026-list.xlsx"
    notice_url = "https://official.example/notices/2026"
    page = PageContent(
        url=notice_url,
        status=200,
        title="2026 年示例奖获奖名单",
        text="现公布 2026 年示例奖获奖名单，附件为完整名单。",
        attachments=[Attachment(
            text="2026 年获奖名单 Excel",
            url=official_url,
            is_excel=True,
        )],
    )
    monkeypatch.setattr(
        loop_mod.tools,
        "fetch_page",
        lambda _url, timeout=15.0: page,
    )
    monkeypatch.setattr(
        loop_mod.tools,
        "download_file",
        lambda _url, _directory, timeout=30.0, **_kwargs: official_path,
    )
    store = Store(tmp_path / "import-m4-agent.db")
    prepared = prepare_review_batch(
        submission_dir,
        store,
        imported_files=[imported],
        registry={imported.claimed_table_code: spec},
        resource_map={"04059999": ResourceMapEntry(
            resource_code="04059999",
            resource_name="示例奖",
            table_code=imported.claimed_table_code,
        )},
        ledger={"04059999": LedgerEntry(
            resource_code="04059999",
            resource_name="示例奖",
            expected_count=1,
            collect_url=notice_url,
        )},
    )
    m4 = run_audit_stage(
        store,
        prepared,
        prober=lambda _url: (200, ""),
        workdir=tmp_path,
    )
    assert m4.bridge.created == 1
    store.close()

    class PacketAwareLlm:
        def __init__(self) -> None:
            self._xlsx_asset_id = ""
            self._scope_id = 0
            self._role = ""
            self._other_asset_ids: list[str] = []

        def json_call(self, _system: str, user: str, *, max_tokens: int) -> Any:
            payload = json.loads(user)
            if max_tokens == 1200:
                assets = payload["assets"]
                spreadsheet = next(asset for asset in assets if asset["kind"] == "xlsx")
                self._xlsx_asset_id = spreadsheet["asset_id"]
                self._other_asset_ids = [
                    asset["asset_id"] for asset in assets
                    if asset["asset_id"] != self._xlsx_asset_id
                ]
                scope = payload["scopes"][0]
                self._scope_id = scope["scope_id"]
                self._role = scope["role"]
                return {
                    "requests": [{
                        "asset_id": self._xlsx_asset_id,
                        "subunit_id": "document",
                        "content_kind": "spreadsheet_sheet",
                        "reason": "确认名单表头和样例。",
                    }],
                    "reason": "需要读取已发现的 Excel 候选。",
                }
            assessments = [{
                "asset_id": self._xlsx_asset_id,
                "scope_ids": [self._scope_id],
                "role": self._role,
                "material_relation": "primary",
                "version_relation": "same",
                "roster_contribution": "include",
                "confidence": 0.99,
                "reason": "页面标题、附件标签和 Sheet 样例均对应目标年度完整名单。",
            }]
            assessments.extend({
                "asset_id": asset_id,
                "scope_ids": [],
                "role": self._role,
                "material_relation": "unrelated",
                "version_relation": "independent",
                "roster_contribution": "exclude",
                "confidence": 0.99,
                "reason": "该材料仅为公告载体，不直接贡献本 scope 名单。",
            } for asset_id in self._other_asset_ids)
            return {
                "case_recommendation": "compare",
                "assessments": assessments,
                "selected_assets": [self._xlsx_asset_id],
                "version_groups": [{
                    "key": "2026-main-list",
                    "asset_ids": [self._xlsx_asset_id],
                    "merge_allowed": True,
                    "reason": "单一目标年度主名单。",
                }],
                "reason": "使用目标年度主附件进行逐 scope 本地比较。",
            }

    llm = PacketAwareLlm()
    results = run_queued_review_cases(
        tmp_path / "import-m4-agent.db",
        prepared.batch_id,
        evidence_roots=[tmp_path],
        semantic_runner_factory=lambda current_store, roots: SemanticReviewRunner(
            CaseRepository(current_store),
            review_llm=llm,
            allowed_roots=roots,
        ),
    )

    assert len(results) == 1
    assert results[0]["conclusion_readiness"] == "ready_for_human"
    final_store = Store(tmp_path / "import-m4-agent.db")
    case_id = int(results[0]["case_id"])
    routes = final_store.list_evidence_asset_routes(case_id)
    assert {route["route_source"] for route in routes} == {"llm"}
    assert {route["route_status"] for route in routes} == {"routed", "excluded"}
    attempt = final_store.list_audit_attempts(case_id)[-1]
    comparisons = final_store.list_scope_comparisons(
        case_id,
        attempt_id=int(attempt["attempt_id"]),
    )
    assert len(comparisons) == 1 and comparisons[0]["comparison_result"] == "matched"
    final_store.close()

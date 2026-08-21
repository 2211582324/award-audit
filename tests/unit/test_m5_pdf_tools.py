"""M5.2 production PDF/OCR/vision tool tests with no real API calls."""

from __future__ import annotations

import json
import time
from pathlib import Path

import openpyxl
import pytest
from PIL import Image
from pypdf import PdfWriter
from reportlab.pdfgen import canvas

from award_audit.agent.toolkit import (
    SafeToolExecutor,
    ToolBudgetLimits,
    ToolExecutionContext,
    build_default_registry,
)
from award_audit.agent.toolkit import pdf as pdf_tools
from award_audit.agent.toolkit.isolation import IsolatedCallTimeout, run_isolated
from award_audit.agent.toolkit.pdf import PdfRuntimeError, resolve_poppler_command

ROOT = Path(__file__).resolve().parents[2]
GOLDEN = ROOT / "tests/data/m5_golden/pdf"


def _context(*roots: Path, **limits: object) -> ToolExecutionContext:
    return ToolExecutionContext.create(roots, ToolBudgetLimits(max_calls=20, **limits))


def _image(path: Path, size: tuple[int, int] = (240, 160)) -> Path:
    Image.new("RGB", size, "white").save(path, format="PNG")
    return path


def test_pdf_inspection_and_bounded_page_extraction() -> None:
    executor = SafeToolExecutor(build_default_registry())
    context = _context(GOLDEN)
    path = GOLDEN / "mixed_roster.pdf"

    inspected = executor.execute("inspect_pdf", {"path": str(path)}, context)
    extracted = executor.execute(
        "extract_pdf_text",
        {"path": str(path), "pages": [1], "max_chars_per_page": 100},
        context,
    )

    assert inspected.ok
    assert inspected.data["page_count"] == 2
    assert inspected.data["digital_pages"] == [1]
    assert inspected.data["scan_candidate_pages"] == [2]
    page = extracted.data["pages"][0]
    assert extracted.ok and page["page"] == 1 and page["is_truncated"] is True
    assert page["tables"] and page["table_rows"] >= 8
    assert context.budget.pdf_pages == 2  # page 1 is not charged twice


def test_pdfplumber_fallback_handles_pypdf_incompatible_pdf(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = GOLDEN / "digital_roster.pdf"
    monkeypatch.setattr(
        pdf_tools,
        "_reader",
        lambda _path: (_ for _ in ()).throw(PdfRuntimeError("pypdf rejected PDF")),
    )

    inspected = pdf_tools.inspect_pdf(path)
    extracted = pdf_tools.extract_pdf_text(path, [1])

    assert inspected.page_count == 2
    assert 1 in inspected.digital_pages
    assert extracted[0].text_chars > 0


def test_pdf_text_extraction_builds_coherent_roster_evidence(tmp_path: Path) -> None:
    submitted = tmp_path / "submitted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["CSDWMC"])
    sheet.append(["Team name"])
    sheet.append(["Team Alpha"])
    sheet.append(["Team Beta"])
    workbook.save(submitted)
    workbook.close()
    pdf_path = tmp_path / "roster.pdf"
    document = canvas.Canvas(str(pdf_path))
    document.drawString(72, 760, "2025 Rural Service Competition winners")
    document.drawString(72, 730, "Team Alpha")
    document.drawString(72, 700, "Team Beta")
    document.save()

    result = SafeToolExecutor(build_default_registry()).execute(
        "extract_pdf_text",
        {
            "path": str(pdf_path),
            "pages": [1],
            "submitted_path": str(submitted),
            "match_fields": ["CSDWMC"],
            "expected_award_name": "National Rural Service Competition",
            "award_aliases": ["Rural Service Competition"],
            "expected_year": "2025",
            "expected_scope_count": 2,
            "source_url": "https://competition.example.edu.cn/roster.pdf",
            "official_secondary_domains": ["competition.example.edu.cn"],
        },
        _context(tmp_path),
    )

    assert result.ok and result.evidence_facts
    fact = result.evidence_facts[0]
    assert fact.status == "complete"
    assert fact.target_match == fact.year_match == "yes"
    assert fact.source_level == "official_secondary"
    assert fact.expected_count == fact.observed_count == 2
    assert fact.coverage_complete is True
    assert fact.artifact_hashes == [result.sha256]


def test_pdf_text_extraction_uses_routed_scope_rows_as_denominator(tmp_path: Path) -> None:
    submitted = tmp_path / "submitted-scopes.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XMBH", "XMLB"])
    sheet.append(["项目编号", "项目类别"])
    sheet.append(["ZL-GOLD-1", "中国专利金奖项目"])
    sheet.append(["ZL-GOLD-2", "中国专利金奖项目"])
    sheet.append(["ZL-SILVER-1", "中国专利银奖项目"])
    workbook.save(submitted)
    workbook.close()
    pdf_path = tmp_path / "gold.pdf"
    document = canvas.Canvas(str(pdf_path))
    document.drawString(72, 760, "2025 中国专利金奖项目名单")
    document.drawString(72, 730, "ZL-GOLD-1")
    document.drawString(72, 700, "ZL-GOLD-2")
    document.save()

    result = SafeToolExecutor(build_default_registry()).execute(
        "extract_pdf_text",
        {
            "path": str(pdf_path), "pages": [1], "submitted_path": str(submitted),
            "match_fields": ["XMBH"],
            "submitted_scope_filter": {"XMLB": "中国专利金奖项目"},
            "scope_id": 42, "role_type": "work_or_project",
            "expected_award_name": "中国专利奖", "expected_year": "2025",
            "expected_scope_count": 2,
        },
        _context(tmp_path),
    )

    assert result.ok
    fact = result.evidence_facts[0]
    assert fact.scope_id == 42
    assert fact.role_type == "work_or_project"
    assert fact.submitted_count == 2
    assert fact.expected_count == fact.observed_count == 2
    assert "ZL-SILVER-1" not in result.data["submitted_identity_items"].values()


def test_pdf_text_extraction_persists_each_unmatched_submitted_identity(
    tmp_path: Path,
) -> None:
    submitted = tmp_path / "submitted-with-missing.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["CSDWMC"])
    sheet.append(["Team name"])
    sheet.append(["Team Alpha"])
    sheet.append(["Team Beta"])
    sheet.append(["Team Gamma"])
    workbook.save(submitted)
    workbook.close()
    pdf_path = tmp_path / "partial-roster.pdf"
    document = canvas.Canvas(str(pdf_path))
    document.drawString(72, 760, "2025 Rural Service Competition winners")
    document.drawString(72, 730, "Team Alpha")
    document.drawString(72, 700, "Team Beta")
    document.save()

    context = _context(tmp_path)
    result = SafeToolExecutor(build_default_registry()).execute(
        "extract_pdf_text",
        {
            "path": str(pdf_path),
            "pages": [1],
            "submitted_path": str(submitted),
            "match_fields": ["CSDWMC"],
            "expected_award_name": "National Rural Service Competition",
            "award_aliases": ["Rural Service Competition"],
            "expected_year": "2025",
            "expected_scope_count": 3,
            "source_url": "https://competition.example.edu.cn/roster.pdf",
            "official_secondary_domains": ["competition.example.edu.cn"],
        },
        context,
    )

    assert result.ok and result.evidence_facts
    fact = result.evidence_facts[0]
    assert fact.observed_count == 2
    assert fact.expected_count == 3
    assert fact.coverage_complete is False
    assert fact.missing_items == ["Team Gamma"]
    assert fact.missing_item_count == 1
    assert result.data["missing_items"] == ["Team Gamma"]
    assert result.data["missing_item_count"] == 1
    assert context.trace[-1].output_summary["verification_facts"]["missing_items"] == [
        "Team Gamma"
    ]


def test_pdf_roster_inherits_identity_only_from_verified_parent_attachment(
    tmp_path: Path,
) -> None:
    submitted = tmp_path / "submitted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["CSDWMC"])
    sheet.append(["Team name"])
    sheet.append(["Team Alpha"])
    sheet.append(["Team Beta"])
    workbook.save(submitted)
    workbook.close()
    pdf_path = tmp_path / "roster-without-heading.pdf"
    document = canvas.Canvas(str(pdf_path))
    document.drawString(72, 730, "Team Alpha")
    document.drawString(72, 700, "Team Beta")
    document.save()

    result = SafeToolExecutor(build_default_registry()).execute(
        "extract_pdf_text",
        {
            "path": str(pdf_path),
            "pages": [1],
            "submitted_path": str(submitted),
            "match_fields": ["CSDWMC"],
            "expected_award_name": "National Rural Service Competition",
            "expected_year": "2025",
            "expected_scope_count": 2,
            "source_url": "https://competition.example.edu.cn/roster.pdf",
            "official_secondary_domains": ["competition.example.edu.cn"],
            "parent_page_url": "https://competition.example.edu.cn/award-2025",
            "parent_attachment_linked": True,
            "parent_award_name": "National Rural Service Competition",
            "parent_year": "2025",
            "parent_source_level": "official_secondary",
        },
        _context(tmp_path),
    )

    assert result.ok and result.evidence_facts
    fact = result.evidence_facts[0]
    assert fact.status == "complete"
    assert fact.target_match == fact.year_match == "yes"
    assert fact.expected_count == fact.observed_count == 2
    assert fact.coverage_complete is True
    assert fact.relationship_confirmed is True
    assert fact.comparison_scope == "verified_parent_page_and_pdf_roster"


def test_pdf_page_limit_and_encrypted_pdf_fail_closed(tmp_path: Path) -> None:
    executor = SafeToolExecutor(build_default_registry())
    limited = _context(GOLDEN, max_pdf_pages=1)
    over_limit = executor.execute(
        "inspect_pdf", {"path": str(GOLDEN / "digital_roster.pdf")}, limited
    )
    assert over_limit.error_code == "PDF_LIMIT_EXCEEDED"

    encrypted_path = tmp_path / "encrypted.pdf"
    writer = PdfWriter()
    writer.add_blank_page(width=200, height=200)
    writer.encrypt("secret")
    with encrypted_path.open("wb") as handle:
        writer.write(handle)
    encrypted = executor.execute(
        "inspect_pdf", {"path": str(encrypted_path)}, _context(tmp_path)
    )
    assert encrypted.error_code == "PDF_ENCRYPTED"


@pytest.mark.skipif(not resolve_poppler_command("pdftoppm"), reason="Poppler unavailable")
def test_render_explicit_page_creates_page_level_artifact(tmp_path: Path) -> None:
    path = GOLDEN / "digital_roster.pdf"
    executor = SafeToolExecutor(build_default_registry())
    context = _context(GOLDEN, tmp_path)
    result = executor.execute(
        "render_pdf_pages",
        {
            "path": str(path),
            "pages": [2],
            "output_dir": str(tmp_path / "render"),
            "dpi": 150,
            "source_url": "https://example.com/roster.pdf",
        },
        context,
    )
    assert result.ok and len(result.artifacts) == 1
    page = result.data["pages"][0]
    artifact = result.artifacts[0]
    assert page["page"] == 2 and Path(page["path"]).is_file()
    assert artifact.metadata["page"] == 2 and artifact.metadata["dpi"] == 150
    assert artifact.metadata["derived_from_sha256"] == result.sha256
    assert context.budget.rendered_pages == 1 and context.budget.image_pixels > 0


def test_fake_rapidocr_returns_boxes_and_confidence(tmp_path: Path) -> None:
    image_path = _image(tmp_path / "page.png")

    class Engine:
        def __call__(self, _path: str):  # noqa: ANN202
            return [
                [[[0, 0], [100, 0], [100, 20], [0, 20]], "1 张伟明 清华大学", 0.98],
                [[[0, 30], [100, 30], [100, 50], [0, 50]], "一等奖", 0.96],
            ], [0.01]

    registry = build_default_registry(ocr_engine_factory=Engine)
    context = _context(tmp_path)
    result = SafeToolExecutor(registry).execute(
        "ocr_image",
        {"images": [{"path": str(image_path), "page": 1, "total_pages": 1}]},
        context,
    )
    page = result.data["pages"][0]
    assert result.ok and page["average_confidence"] == 0.97
    assert len(page["lines"][0]["box"]) == 4 and page["text"].startswith("1 张伟明")
    assert page["detected_numbers"] == [1]
    assert page["sequence_contiguous"] is True and page["needs_vision"] is False
    assert context.budget.ocr_pages == 1


def test_fake_vision_is_schema_validated_and_lazy(tmp_path: Path) -> None:
    image_path = _image(tmp_path / "page.png")
    prompts: list[str] = []

    class Client:
        provider = "openai"
        model = "fake-vision"

        def vision_json_call(self, _system, user, *_args, **_kwargs):  # noqa: ANN002, ANN003, ANN202
            prompts.append(user)
            return {
                "page": 1,
                "total_pages": 1,
                "is_roster_page": True,
                "headers": ["序号", "姓名", "单位", "等级"],
                "entries": [
                    {"no": 1, "name": "张伟明", "org": "清华大学", "level": "一等奖"}
                ],
                "first_no": 1,
                "last_no": 1,
                "truncated": False,
                "unreadable": [],
                "confidence": 0.99,
                "visible_row_count": 1,
                "all_rows_extracted": True,
            }

    calls = 0

    def client_factory() -> Client:
        nonlocal calls
        calls += 1
        return Client()

    registry = build_default_registry(vision_client_factory=client_factory)
    assert calls == 0  # registry construction must not load model configuration
    result = SafeToolExecutor(registry).execute(
        "vision_extract_roster",
        {
            "images": [{"path": str(image_path), "page": 1, "total_pages": 1}],
            "ocr_text_by_page": {1: "OCR-ROW-1"},
        },
        _context(tmp_path),
    )
    assert result.ok and result.data["complete"] is True and calls == 1
    assert result.data["pages"][0]["image_sha256"]
    assert prompts and "OCR-ROW-1" in prompts[0]


def test_vision_visible_row_count_mismatch_is_structured_failure(tmp_path: Path) -> None:
    image_path = _image(tmp_path / "page.png")

    class Client:
        provider = "openai"
        model = "fake-incomplete"

        def vision_json_call(self, *_args, **_kwargs):  # noqa: ANN002, ANN003, ANN202
            return {
                "is_roster_page": True,
                "entries": [{"no": 1, "name": "only one row"}],
                "first_no": 1,
                "last_no": 1,
                "truncated": False,
                "unreadable": [],
                "confidence": 0.99,
                "visible_row_count": 2,
                "all_rows_extracted": True,
            }

    result = SafeToolExecutor(build_default_registry(vision_client_factory=Client)).execute(
        "vision_extract_roster",
        {"images": [{"path": str(image_path), "page": 1, "total_pages": 1}]},
        _context(tmp_path),
    )
    assert result.error_code == "VISION_EXTRACTION_FAILED"
    assert result.data["complete"] is False
    assert result.data["errors"][0]["error_code"] == "VISION_OUTPUT_INVALID"


def test_vision_incomplete_page_is_retried_before_success(tmp_path: Path) -> None:
    image_path = _image(tmp_path / "page.png")
    calls = 0

    class Client:
        provider = "openai"
        model = "fake-retry"

        def vision_json_call(self, *_args, **_kwargs):  # noqa: ANN002, ANN003, ANN202
            nonlocal calls
            calls += 1
            return {
                "is_roster_page": True,
                "entries": [{"no": 1, "name": "one row"}],
                "first_no": 1,
                "last_no": 1,
                "truncated": False,
                "unreadable": [],
                "confidence": 0.7 if calls == 1 else 0.99,
                "visible_row_count": 1,
                "all_rows_extracted": True,
            }

    result = SafeToolExecutor(build_default_registry(vision_client_factory=Client)).execute(
        "vision_extract_roster",
        {"images": [{"path": str(image_path), "page": 1, "total_pages": 1}]},
        _context(tmp_path),
    )
    assert result.ok and result.data["complete"] is True
    assert calls == 2


def test_invalid_vision_page_is_a_structured_failure(tmp_path: Path) -> None:
    image_path = _image(tmp_path / "page.png")

    class Client:
        provider = "openai"
        model = "fake-invalid"

        def vision_json_call(self, *_args, **_kwargs):  # noqa: ANN002, ANN003, ANN202
            return {"page": 99, "total_pages": 1, "entries": "not-a-list"}

    result = SafeToolExecutor(build_default_registry(vision_client_factory=Client)).execute(
        "vision_extract_roster",
        {"images": [{"path": str(image_path), "page": 1, "total_pages": 1}]},
        _context(tmp_path),
    )
    assert result.error_code == "VISION_EXTRACTION_FAILED"
    assert result.data["complete"] is False and result.data["errors"][0]["page"] == 1


def test_compare_roster_requires_complete_sequence_and_readable_pages(tmp_path: Path) -> None:
    entry = {"no": 1, "name": "张伟明", "org": "清华大学", "level": "一等奖"}
    page = {
        "page": 1,
        "total_pages": 1,
        "entries": [entry],
        "first_no": 1,
        "last_no": 1,
        "truncated": False,
        "unreadable": [],
        "confidence": 0.99,
    }
    executor = SafeToolExecutor(build_default_registry())
    context = _context(tmp_path)
    complete = executor.execute(
        "compare_roster",
        {"submitted": [entry], "official_pages": [page], "expected_total": 1},
        context,
    )
    page["truncated"] = True
    truncated = executor.execute(
        "compare_roster",
        {"submitted": [entry], "official_pages": [page], "expected_total": 1},
        context,
    )
    assert complete.data["consistent"] is True
    assert truncated.data["coverage_complete"] is False
    assert truncated.data["manual_review_required"] is True
    assert "page_truncated_or_unreadable" in truncated.warnings
    trace = json.dumps([item.model_dump(mode="json") for item in context.trace], ensure_ascii=False)
    assert "张伟明" not in trace and "清华大学" not in trace
    assert "content_redacted" in trace


def test_image_pixel_and_ocr_page_budgets_fail_closed(tmp_path: Path) -> None:
    first = _image(tmp_path / "first.png", (20, 20))
    second = _image(tmp_path / "second.png", (20, 20))

    class Engine:
        def __call__(self, _path: str):  # noqa: ANN202
            return [], [0.01]

    registry = build_default_registry(ocr_engine_factory=Engine)
    too_many = SafeToolExecutor(registry).execute(
        "ocr_image",
        {"images": [
            {"path": str(first), "page": 1, "total_pages": 2},
            {"path": str(second), "page": 2, "total_pages": 2},
        ]},
        _context(tmp_path, max_ocr_pages=1),
    )
    too_large = SafeToolExecutor(registry).execute(
        "ocr_image",
        {"images": [{"path": str(first), "page": 1, "total_pages": 1}]},
        _context(tmp_path, max_image_pixels=100),
    )
    assert too_many.error_code == "TOOL_BUDGET_EXCEEDED"
    assert too_large.error_code == "IMAGE_LIMIT_EXCEEDED"


def test_pdf_path_dpi_and_vision_budget_boundaries(tmp_path: Path) -> None:
    image_path = _image(tmp_path / "page.png")

    class Client:
        provider = "openai"
        model = "must-not-run"

        def vision_json_call(self, *_args, **_kwargs):  # noqa: ANN002, ANN003, ANN202
            raise AssertionError("vision must not run after a budget rejection")

    executor = SafeToolExecutor(build_default_registry(vision_client_factory=Client))
    outside = executor.execute(
        "inspect_pdf",
        {"path": str(GOLDEN / "digital_roster.pdf")},
        _context(tmp_path),
    )
    invalid_dpi = executor.execute(
        "render_pdf_pages",
        {
            "path": str(GOLDEN / "digital_roster.pdf"),
            "pages": [1],
            "output_dir": str(tmp_path),
            "dpi": 300,
        },
        _context(GOLDEN, tmp_path),
    )
    vision_budget = executor.execute(
        "vision_extract_roster",
        {"images": [{"path": str(image_path), "page": 1, "total_pages": 1}]},
        _context(tmp_path, max_vision_pages=0),
    )
    assert outside.error_code == "UNSAFE_PATH"
    assert invalid_dpi.error_code == "TOOL_INPUT_INVALID"
    assert vision_budget.error_code == "TOOL_BUDGET_EXCEEDED"


def test_isolated_parser_timeout_terminates_worker() -> None:
    started = time.monotonic()
    with pytest.raises(IsolatedCallTimeout):
        run_isolated(time.sleep, args=(5,), timeout_seconds=0.1)
    assert time.monotonic() - started < 3

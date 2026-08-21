"""L5-agent 工具层测试：HTML 正文/附件解析（离线夹具）+ 官网 Excel 解析。"""

from __future__ import annotations

import openpyxl

from award_audit.agent import tools

# 仿真实公示页（提交-13 研创赛形态）：正文 + 附件1盖章版PDF + 附件2Excel + 名单截图 + logo装饰图
_HTML = """
<html><head><style>.x{color:red}</style><script>var a=1;</script></head>
<body>
<img src="/static/logo.png">
<h1>"华为杯"第七届中国研究生人工智能创新大赛获奖名单公布</h1>
<p>发布时间：2025-11-06</p>
<p>经大赛组委会秘书处核实，现正式公布获奖名单（名单附后）。</p>
<img src="/upload/mingdan_1.jpg">
<a href="/upload/hjmd_seal.pdf">附件1： 获奖名单（盖章版）</a>
<a href="/upload/hjmd.xlsx">附件2： 获奖名单（Excel）</a>
<a href="/other/page">返回</a>
</body></html>
"""


# 功能：验证 HTML 解析产出干净正文、识别附件（仅 xlsx 标记 Excel）、收集内容图并滤掉 logo
# 设计：夹具含 script/logo/普通链接干扰，断言附件 2 个、内容图恰 1 张（mingdan_1.jpg）、logo 被滤
def test_parse_html_extracts_text_and_attachments() -> None:
    text, atts, images = tools.parse_html(_HTML, "https://cpipc.acge.org.cn/cw/detail/x/y")
    assert "获奖名单公布" in text and "var a=1" not in text
    assert len(atts) == 2
    excel = [a for a in atts if a.is_excel]
    assert len(excel) == 1
    assert excel[0].url == "https://cpipc.acge.org.cn/upload/hjmd.xlsx"
    assert "附件2" in excel[0].text
    assert images == ["https://cpipc.acge.org.cn/upload/mingdan_1.jpg"]


# 功能：验证无附件的普通页面返回空附件与空图片清单
# 设计：只有普通导航链接的 HTML，断言均为空——识别不误报
def test_parse_html_no_attachments() -> None:
    text, atts, images = tools.parse_html('<a href="/home">首页</a><p>正文</p>', "https://x.cn/")
    assert atts == [] and images == [] and "正文" in text


def test_fetch_page_marks_visible_text_truncation(monkeypatch) -> None:  # noqa: ANN001
    class FakeResponse:
        status_code = 200
        headers = {"content-type": "text/html; charset=utf-8"}
        url = "https://8.8.8.8/long"
        content = (
            "<html><body>"
            + ("名单内容" * (tools.MAX_TEXT_CHARS // 4 + 100))
            + "</body></html>"
        ).encode()

    class FakeClient:
        def __init__(self, **kwargs):  # noqa: ANN003
            pass

        def __enter__(self):  # noqa: ANN204
            return self

        def __exit__(self, *args):  # noqa: ANN002, ANN204
            return False

        def get(self, url):  # noqa: ANN001, ANN201
            return FakeResponse()

    import httpx

    monkeypatch.setattr(httpx, "Client", FakeClient)

    page = tools.fetch_page("https://8.8.8.8/long")

    assert len(page.text) == tools.MAX_TEXT_CHARS
    assert page.text_truncated is True
    assert page.original_text_chars > len(page.text)


# 功能：验证 <base href> 被尊重——相对链接以 base 标签为基准解析，而非页面 URL（研创赛平台实测坑）
# 设计：页面 URL 深在 /cw/detail/xxx，<base> 指向站根；断言图片/附件解析到站根而非页面路径
def test_parse_html_honors_base_href() -> None:
    html = ("<head><base href='https://cpipc.acge.org.cn/'></head><body>"
            "<img src='source/website/images/46.png'>"
            "<a href='source/hjmd.xlsx'>附件2：获奖名单（Excel）</a></body>")
    page_url = "https://cpipc.acge.org.cn/cw/detail/aaa/bbb"
    text, atts, images = tools.parse_html(html, page_url)
    assert images == ["https://cpipc.acge.org.cn/source/website/images/46.png"]
    assert len(atts) == 1 and atts[0].url == "https://cpipc.acge.org.cn/source/hjmd.xlsx"
    assert atts[0].is_excel


def test_parse_html_recognizes_excel_from_download_link_text() -> None:
    html = """
    <a href="/system/_content/download.jsp?urltype=news.DownloadAttachUrl&amp;wbfileid=1">
      1.2023年“最美高校辅导员”候选人.xlsx
    </a>
    <a href="/system/_content/download.jsp?urltype=news.DownloadAttachUrl&amp;wbfileid=2">
      2.第十五届“高校辅导员年度人物”候选人.xlsx
    </a>
    <img src="/images/ico1-dl.png">
    <img src="/images/wx-ewm-99.jpg">
    """

    _text, attachments, images = tools.parse_html(
        html,
        "https://www.chinazy.org/info/1014/15997.htm",
    )

    assert [item.text for item in attachments] == [
        "1.2023年“最美高校辅导员”候选人.xlsx",
        "2.第十五届“高校辅导员年度人物”候选人.xlsx",
    ]
    assert all(item.is_excel for item in attachments)
    assert images == []


# 功能：验证官网 Excel 名单解析：取首个有数据的 sheet、返回原始网格
# 设计：造一个表头与我们模板完全不同的官网风格 xlsx（奖项等级/作品/学校/成员），
#       断言网格行数与首行表头——列语义留给 LLM 映射，解析层不做模板假设
def test_parse_award_excel(tmp_path) -> None:
    p = tmp_path / "官网名单.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "获奖名单"
    ws.append(["序号", "奖项等级", "作品名称", "学校", "团队成员"])
    ws.append([1, "一等奖", "智能体研究", "某大学", "张三、李四"])
    ws.append([2, "二等奖", "视觉导航", "另一大学", "王五"])
    wb.save(p)
    grid = tools.parse_award_excel(p)
    assert grid["sheet"] == "获奖名单" and grid["n_rows"] == 3
    assert grid["rows"][0][1] == "奖项等级"


# 功能：多 sheet（按等级分页）的名单须全部合并，且每段前注入「【等级：页名】」标记
# 设计：造一等奖/二等奖两个 sheet，断言两页数据都在、页名被登记、标记行落在各段之首——
#       否则只读第一页会把二等奖整体误判为"官网查无/疑似多采"
def test_parse_award_excel_multi_sheet(tmp_path) -> None:
    p = tmp_path / "分级名单.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "一等奖"
    ws1.append(["作品名称", "参赛单位"])
    ws1.append(["作品甲", "甲大学"])
    ws2 = wb.create_sheet("二等奖")
    ws2.append(["作品名称", "参赛单位"])
    ws2.append(["作品乙", "乙大学"])
    wb.save(p)
    grid = tools.parse_award_excel(p)
    assert grid["sheets"] == ["一等奖", "二等奖"]
    assert grid["sheet"] == "一等奖 / 二等奖"
    assert grid["rows"][0] == ["【等级：一等奖】"]
    flat = ["|".join(r) for r in grid["rows"]]
    assert "作品甲|甲大学" in flat and "作品乙|乙大学" in flat  # 两页都在
    assert ["【等级：二等奖】"] in grid["rows"]                  # 第二段标记也注入
    assert grid["sheet_grids"] == [
        {
            "sheet": "一等奖",
            "n_rows": 2,
            "rows": [["作品名称", "参赛单位"], ["作品甲", "甲大学"]],
            "truncated": False,
        },
        {
            "sheet": "二等奖",
            "n_rows": 2,
            "rows": [["作品名称", "参赛单位"], ["作品乙", "乙大学"]],
            "truncated": False,
        },
    ]


def test_parse_award_excel_default_reads_more_than_2000_rows(tmp_path) -> None:
    path = tmp_path / "large-roster.xlsx"
    workbook = openpyxl.Workbook(write_only=True)
    sheet = workbook.create_sheet("名单")
    sheet.append(["姓名"])
    for index in range(1, 5_109):
        sheet.append([f"候选人{index}"])
    workbook.save(path)

    grid = tools.parse_award_excel(path)

    assert grid["n_rows"] == 5_109
    assert grid["truncated"] is False
    assert grid["sheet_row_counts"] == {"名单": 5_109}


def test_parse_award_excel_reports_explicit_row_limit(tmp_path) -> None:
    path = tmp_path / "bounded-roster.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "名单"
    for index in range(12):
        sheet.append([f"第{index}行"])
    workbook.save(path)
    workbook.close()

    grid = tools.parse_award_excel(path, max_rows=10)

    assert grid["n_rows"] == 10
    assert grid["truncated"] is True
    assert grid["truncated_sheets"] == ["名单"]
    assert grid["row_limit"] == 10


def test_semantic_roster_extraction_does_not_treat_team_school_as_organization() -> None:
    grid = {
        "sheet_grids": [{
            "sheet": "Sheet1", "n_rows": 4, "truncated": False,
            "rows": [
                ["第一届志愿服务技能大赛一等奖获奖团队名单"],
                ["序号", "学校", "团队名称", "组别", "指导导师姓名"],
                ["1", "甲大学", "甲团队", "技术类", "张老师"],
                ["2", "乙大学", "乙团队", "教育类", "李老师"],
            ],
        }],
    }

    records = tools.extract_semantic_roster_records(grid)

    assert [(record["role_type"], record["identity"]) for record in records] == [
        ("team", "甲团队"), ("team", "乙团队"),
    ]
    assert records[0]["category_values"] == ["技术类"]


def test_semantic_roster_extraction_detects_organization_award_table() -> None:
    grid = {
        "sheet_grids": [{
            "sheet": "组织奖", "n_rows": 4, "truncated": False,
            "rows": [
                ["第一届乡村振兴志愿服务技能大赛优秀组织单位"],
                ["序号", "参赛单位"],
                ["1", "甲大学"],
                ["2", "乙大学"],
            ],
        }],
    }

    records = tools.extract_semantic_roster_records(grid)

    assert [(record["role_type"], record["identity"]) for record in records] == [
        ("organization", "甲大学"), ("organization", "乙大学"),
    ]


def test_semantic_roster_extraction_keeps_project_category() -> None:
    grid = {
        "sheet_grids": [{
            "sheet": "PDF page 1", "n_rows": 2, "truncated": False,
            "rows": [
                ["序号", "项目名称", "项目类别", "申请人", "学校名称"],
                ["1", "项目甲", "青年基金项目", "张三", "甲大学"],
            ],
        }],
    }

    records = tools.extract_semantic_roster_records(grid)

    assert records[0]["identity"] == "项目甲"
    assert records[0]["category_values"] == ["青年基金项目"]


# 功能：acquire_excel_grid 从页面附件取到第一份可解析网格 + 溯源（collect/回写共用的纯抓取函数）
# 设计：mock fetch_page 返回带 Excel 附件的页、download/parse 返回网格，断言拿到 grid 与命中附件 URL；
#       download mock 用 **kw 吸收 excel_only/referer，贴合真实调用签名
def test_acquire_excel_grid(tmp_path, monkeypatch) -> None:  # noqa: ANN001
    page = tools.PageContent(url="https://x/a", status=200, text="",
                             attachments=[tools.Attachment(text="名单", url="https://x/m.xlsx", is_excel=True)],
                             images=["https://x/pic.png"])
    monkeypatch.setattr(tools, "fetch_page", lambda url, timeout=15.0: page)
    monkeypatch.setattr(tools, "download_file", lambda url, d, timeout=30.0, **kw: d / "m.xlsx")
    monkeypatch.setattr(tools, "parse_award_excel",
                        lambda p, max_rows=2000: {"sheet": "s", "n_rows": 1, "rows": [["题目"], ["甲"]]})
    got = tools.acquire_excel_grid(["https://x/a"], tmp_path)
    assert got is not None
    assert got.source_url == "https://x/m.xlsx" and got.page_url == "https://x/a"
    assert got.grid["rows"] == [["题目"], ["甲"]]
    assert "https://x/pic.png" in got.found_assets  # 页面资产也登记，供未命中转人工


# 功能：acquire_excel_grid 全程取不到可解析 Excel 时返回 None（页面死链/无附件）
# 设计：fetch_page 返回 404，断言 None——collect 遇此跳过、审查时留待人工
def test_acquire_excel_grid_uses_resolved_page_for_attachment_referer(
    tmp_path, monkeypatch
) -> None:  # noqa: ANN001
    requested = "https://x/notice"
    resolved = "https://x/notice;jsessionid=abc"
    attachment_url = "https://x/download?id=1"
    page = tools.PageContent(
        url=resolved,
        status=200,
        text="",
        attachments=[tools.Attachment(text="list", url=attachment_url, is_excel=True)],
    )
    downloads: list[tuple[str, str]] = []

    def download(url, destination, timeout=30.0, **kwargs):  # noqa: ANN001, ARG001
        downloads.append((url, str(kwargs["referer"])))
        return destination / "list.xlsx"

    monkeypatch.setattr(tools, "fetch_page", lambda _url, timeout=15.0: page)
    monkeypatch.setattr(tools, "download_file", download)
    monkeypatch.setattr(
        tools,
        "parse_award_excel",
        lambda _path, max_rows=2000: {"sheet": "s", "n_rows": 1, "rows": [["x"]]},
    )

    got = tools.acquire_excel_grid([requested], tmp_path)

    assert got is not None
    assert got.page_url == resolved
    assert got.attachment_parent_urls[attachment_url] == resolved
    assert downloads == [(attachment_url, resolved)]


def test_acquire_excel_grid_none(tmp_path, monkeypatch) -> None:  # noqa: ANN001
    monkeypatch.setattr(tools, "fetch_page",
                        lambda url, timeout=15.0: tools.PageContent(url=url, status=404, text=""))
    assert tools.acquire_excel_grid(["https://x/dead"], tmp_path) is None


def test_acquire_excel_grid_reuses_direct_attachment_when_parent_page_fails(
    tmp_path,
    monkeypatch,
) -> None:  # noqa: ANN001
    parent_url = "https://x/notice"
    attachment_url = "https://x/files/list.xlsx"
    monkeypatch.setattr(
        tools,
        "fetch_page",
        lambda url, timeout=15.0: tools.PageContent(url=url, status=503, text=""),
    )
    downloaded: list[tuple[str, str]] = []

    def download(url, destination, timeout=30.0, **kwargs):  # noqa: ANN001, ARG001
        downloaded.append((url, str(kwargs.get("referer", ""))))
        return destination / "list.xlsx"

    monkeypatch.setattr(tools, "download_file", download)
    monkeypatch.setattr(
        tools,
        "parse_award_excel",
        lambda path, max_rows=2000: {
            "sheet": "名单",
            "n_rows": 2,
            "rows": [["姓名"], ["张三"]],
        },
    )

    got = tools.acquire_excel_grid(
        [parent_url],
        tmp_path,
        direct_attachment_urls=[attachment_url],
        direct_referer=parent_url,
    )

    assert got is not None
    assert got.source_urls == [attachment_url]
    assert got.page_url == parent_url
    assert got.found_assets == [attachment_url]
    assert downloaded == [(attachment_url, parent_url)]


# 功能：多个赛道附件被合并成一份网格，各段前注入「【名单：<赛道>】」标记（研创赛一奖多赛道）
# 设计：页面两份 Excel 附件，parse 各返回一行数据，断言 sheets/source_urls 收齐两份、两段标记与数据都在
def test_acquire_excel_grid_merges_multiple(tmp_path, monkeypatch) -> None:  # noqa: ANN001
    page = tools.PageContent(url="https://x/a", status=200, text="",
                             attachments=[
                                 tools.Attachment(text="环境空间设计赛道", url="https://x/s1.xlsx", is_excel=True),
                                 tools.Attachment(text="服装与服饰设计赛道", url="https://x/s2.xlsx", is_excel=True)])
    monkeypatch.setattr(tools, "fetch_page", lambda url, timeout=15.0: page)
    monkeypatch.setattr(tools, "download_file", lambda url, d, timeout=30.0, **kw: d / "s.xlsx")
    monkeypatch.setattr(tools, "parse_award_excel",
                        lambda p, max_rows=2000: {"sheet": "s", "n_rows": 2, "rows": [["作品"], ["甲作品"]]})
    got = tools.acquire_excel_grid(["https://x/a"], tmp_path)
    assert got is not None
    assert got.source_urls == ["https://x/s1.xlsx", "https://x/s2.xlsx"]
    assert got.grid["sheets"] == ["环境空间设计赛道", "服装与服饰设计赛道"]
    assert ["【名单：环境空间设计赛道】"] in got.grid["rows"]
    assert ["【名单：服装与服饰设计赛道】"] in got.grid["rows"]
    flat = ["|".join(r) for r in got.grid["rows"]]
    assert flat.count("甲作品") == 2  # 两个赛道段各一份


def test_acquire_excel_grid_reports_unprocessed_attachment_group(
    tmp_path,
    monkeypatch,
) -> None:  # noqa: ANN001
    urls = [f"https://x/group-{index}.xlsx" for index in range(1, 4)]
    page = tools.PageContent(
        url="https://x/a",
        status=200,
        text="",
        attachments=[
            tools.Attachment(text=f"第{index}组", url=url, is_excel=True)
            for index, url in enumerate(urls, start=1)
        ],
    )
    monkeypatch.setattr(tools, "fetch_page", lambda url, timeout=15.0: page)
    monkeypatch.setattr(
        tools,
        "download_file",
        lambda url, destination, timeout=30.0, **kwargs: destination / "group.xlsx",
    )
    monkeypatch.setattr(
        tools,
        "parse_award_excel",
        lambda path, max_rows=100_000: {
            "sheet": "名单",
            "n_rows": 2,
            "rows": [["项目名称"], ["示例项目"]],
        },
    )

    acquired = tools.acquire_excel_grid([page.url], tmp_path, max_total=2)

    assert acquired is not None
    assert acquired.discovered_attachment_urls == urls
    assert acquired.attempted_attachment_urls == urls[:2]
    assert acquired.failed_attachment_urls == []
    assert acquired.unprocessed_attachment_urls == [urls[2]]
    assert acquired.all_attachments_processed is False


def test_acquire_excel_grid_retries_a_transient_attachment_failure(
    tmp_path, monkeypatch
) -> None:  # noqa: ANN001
    attachment = tools.Attachment(text="list", url="https://x/list.xlsx", is_excel=True)
    page = tools.PageContent(url="https://x/notice", status=200, text="", attachments=[attachment])
    calls = 0

    def download(url, destination, timeout=30.0, **kwargs):  # noqa: ANN001, ARG001
        nonlocal calls
        calls += 1
        if calls == 1:
            raise RuntimeError("temporary timeout")
        return destination / "list.xlsx"

    monkeypatch.setattr(tools, "fetch_page", lambda _url, timeout=15.0: page)
    monkeypatch.setattr(tools, "download_file", download)
    monkeypatch.setattr(tools, "parse_award_excel", lambda _path, max_rows=2000: {"rows": [["x"]]})

    acquired = tools.acquire_excel_grid([page.url], tmp_path)

    assert acquired is not None
    assert calls == 2
    assert acquired.failed_attachment_urls == []


# 功能：年度抽取 extract_years——识别 4 位合规年份（2015–2035）、跨年区间返回多值、无年度返回空集
# 设计：断言"YYYY年第X届"取到年、纯届次"第七届"取空、"2024-2025学年"取两值、区间外(1999)被滤
def test_extract_years() -> None:
    assert tools.extract_years("2025年第三届全国大学生大赛获奖名单") == {"2025"}
    assert tools.extract_years("第七届获奖名单") == set()
    assert tools.extract_years("2024-2025学年名单") == {"2024", "2025"}
    assert tools.extract_years("咨询电话1999-2050") == set()  # 区间外伪年份被滤


# 功能：附件名无年度时，采集侧把详情页标题的唯一年度折进「【名单：X（YYYY）】」标签（① 年度分片采集侧）
# 设计：页面 title 含 2025、两个赛道附件名都不含年度；断言两段标记都被折入（2025），供审查侧按年度选源
def test_acquire_folds_page_title_year(tmp_path, monkeypatch) -> None:  # noqa: ANN001
    page = tools.PageContent(url="https://x/a", status=200, text="",
                             title="2025年第三届华为杯研究生大赛获奖名单公布",
                             attachments=[
                                 tools.Attachment(text="环境空间设计赛道", url="https://x/s1.xlsx", is_excel=True),
                                 tools.Attachment(text="服装与服饰设计赛道", url="https://x/s2.xlsx", is_excel=True)])
    monkeypatch.setattr(tools, "fetch_page", lambda url, timeout=15.0: page)
    monkeypatch.setattr(tools, "download_file", lambda url, d, timeout=30.0, **kw: d / "s.xlsx")
    monkeypatch.setattr(tools, "parse_award_excel",
                        lambda p, max_rows=2000: {"sheet": "s", "n_rows": 2, "rows": [["作品"], ["甲作品"]]})
    got = tools.acquire_excel_grid(["https://x/a"], tmp_path)
    assert got is not None
    assert ["【名单：环境空间设计赛道（2025）】"] in got.grid["rows"]
    assert ["【名单：服装与服饰设计赛道（2025）】"] in got.grid["rows"]
    assert tools.extract_years("环境空间设计赛道（2025）") == {"2025"}  # 审查侧能从来源名读回年度

"""统一产出 report 的单元测试：write_reports 传 audit_reports 追加 L5 段/sheet，不传则与旧产出兼容。"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from award_audit.core.pipeline.engine import BatchResult, FileResult
from award_audit.core.pipeline.report import write_reports


# 造一个最小批次结果（一个全绿文件、无 issue）
def _result() -> BatchResult:
    return BatchResult(batch="提交-T", files=[
        FileResult(file="f.xlsx", claimed_table_code="CON_X", n_rows=2, issues=[]),
    ])


# 造一条联网核对结论 dict（EvidenceReport.model_dump() 形态）
def _arep(code: str, verdict: str, **kw: object) -> dict:
    base: dict = {
        "resource_code": code, "award_name": "测试奖", "year": "2024", "verdict": verdict,
        "confidence": "high", "source_kind": "excel", "source_url": "https://gov/x",
        "page_year": "2024", "extracted_count": 2, "submitted_count": 2,
        "missing": [], "extra": [], "source_urls": ["https://gov/p"], "found_assets": [],
        "evidence": [], "notes": "",
    }
    base.update(kw)
    return base


# 功能：验证 write_reports 传 audit_reports 时 md 追加「联网核对(L5)」段、xlsx 加「联网核对」sheet
# 设计：最小 BatchResult + 两条 audit dict（一致 / 无法核对带人工入口），断言 md 含 verdict 与人工入口 URL、
#       openpyxl 断言第二 sheet 存在且含表头+两条结论行——统一产出把 L5 并进同一份反馈
def test_write_reports_appends_l5_section(tmp_path: Path) -> None:
    reps = [
        _arep("04050014", "一致"),
        _arep("04050099", "无法核对", source_kind="none",
              found_assets=["https://gov/list.png"], source_urls=["https://gov/p2"]),
    ]
    xlsx, md = write_reports(_result(), out_dir=tmp_path, audit_reports=reps)
    text = md.read_text(encoding="utf-8")
    assert "## 联网核对(L5)" in text
    assert "无法核对" in text and "https://gov/list.png" in text  # 结论 + 人工入口 URL 都进 md
    wb = openpyxl.load_workbook(xlsx)
    assert "联网核对" in wb.sheetnames
    assert wb["联网核对"].max_row == 3  # 表头 + 两条结论


# 功能：验证 L5 反馈把分诊落到 md（分诊统计行 + 转人工/快速放行标签 + 降级原因）与 xlsx（分诊/降级原因列表头）
# 设计：一条一致（放行）+ 一条无法核对带降级码（转人工），断言 md 含分诊标签与原因中文、xlsx 表头有「分诊」「降级原因」
def test_l5_report_surfaces_triage(tmp_path: Path) -> None:
    reps = [
        _arep("04050014", "一致"),
        _arep("04050099", "无法核对", confidence="low", source_kind="none",
              reason_codes=["no_list"], found_assets=["https://gov/list.png"]),
    ]
    xlsx, md = write_reports(_result(), out_dir=tmp_path, audit_reports=reps)
    text = md.read_text(encoding="utf-8")
    assert "分诊" in text and "转人工" in text and "快速放行" in text  # 分诊统计行 + 表格分诊列
    assert "未获取到任何官网名单" in text  # 降级原因中文标签进 md
    header = [c.value for c in openpyxl.load_workbook(xlsx)["联网核对"][1]]
    assert "分诊" in header and "降级原因" in header


# 功能：验证不传 audit_reports 时产出与旧行为一致——md 无 L5 段、xlsx 仅「反馈意见」一个 sheet（向后兼容）
# 设计：同一 BatchResult 不传 audit_reports 调 write_reports，断言 md 不含「联网核对(L5)」、workbook 只一个 sheet
def test_write_reports_without_audit_is_backward_compatible(tmp_path: Path) -> None:
    xlsx, md = write_reports(_result(), out_dir=tmp_path)
    assert "## 联网核对(L5)" not in md.read_text(encoding="utf-8")
    assert openpyxl.load_workbook(xlsx).sheetnames == ["反馈意见"]


def test_write_reports_appends_bounded_m5_case_status(tmp_path: Path) -> None:
    cases = [{
        "case_id": 7,
        "resource_code": "04050014",
        "award_name": "测试奖",
        "year": "2026",
        "status": "waiting_human",
        "trigger_codes": ["PDF_ONLY"],
        "confidence": "medium",
        "recommendation": "证据待人工复核",
        "evidence_sources": ["https://official.example/list.pdf"],
        "evidence_hashes": ["a" * 64],
        "evidence_times": ["2026-07-25T00:00:00Z"],
        "verification_action": "supplement",
        "human_decision": "",
        "human_decision_summary": "",
        "reviewed_by": "",
        "state_version": 3,
    }]
    xlsx, md = write_reports(_result(), out_dir=tmp_path, audit_cases=cases)
    text = md.read_text(encoding="utf-8")
    assert "## M5 疑难案件" in text
    assert "waiting_human" in text and "待复核台终审" in text
    assert "https://official.example/list.pdf" in text and "local_path" not in text
    workbook = openpyxl.load_workbook(xlsx)
    assert workbook.sheetnames == ["反馈意见", "疑难案件"]
    header = [cell.value for cell in workbook["疑难案件"][1]]
    assert "证据哈希" in header and "人工决定" in header

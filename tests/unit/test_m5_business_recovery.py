"""Offline regressions for the six submission-14 business failures."""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

from award_audit.agent.harness.client import FakeAgentClient
from award_audit.agent.harness.models import (
    AuditCaseState,
    CaseSeed,
    EvidenceCandidate,
    NextAction,
)
from award_audit.agent.harness.persistence import CaseRepository
from award_audit.agent.harness.runner import (
    EvidenceHarness,
    _asset_followup_allowed,
    _attempted_urls,
    _collect_arguments,
    _consume_pending_media,
    _deterministic_action,
    _extract_arguments,
    _fetch_arguments,
    _image_roster_arguments,
    _next_unattempted_known_url,
    _pdf_extract_arguments,
    _route_web_result_to_scopes,
    _search_arguments,
)
from award_audit.agent.toolkit import pdf as pdf_tools
from award_audit.agent.toolkit import registry as registry_module
from award_audit.agent.toolkit.contracts import (
    EvidenceArtifact,
    EvidenceFact,
    ToolBudgetLimits,
    ToolBudgetState,
    ToolObservation,
    ToolResult,
)
from award_audit.agent.toolkit.registry import (
    CollectSpreadsheetAttachmentsInput,
    ExtractSearchDocumentInput,
    SafeToolExecutor,
    ToolExecutionContext,
    ToolRegistry,
    VerifyPageImageRosterInput,
    _match_award_title,
    _submitted_match_items_from_paths,
    build_default_registry,
)
from award_audit.agent.toolkit.search import ExtractResponse, FakeSearchProvider
from award_audit.agent.toolkit.testing import register_fake_tool
from award_audit.agent.toolkit.web import Attachment, PageContent, parse_html
from award_audit.agent.verification import EvidenceVerifier
from award_audit.agent.verification.models import AutoApprovalPolicy, VerificationReport
from award_audit.agent.verification.service import (
    FakeVerifierClient,
    build_evidence_snapshot,
    decide_review_route,
    deterministic_verify,
)
from award_audit.core.pipeline.store import Store


def test_submitted_identity_reader_does_not_truncate_large_roster(tmp_path: Path) -> None:
    path = tmp_path / "large-submission.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XMBH"])
    sheet.append(["项目编号"])
    for index in range(2505):
        sheet.append([f"P-{index:04d}"])
    workbook.save(path)

    identities = _submitted_match_items_from_paths(
        [path],
        ["XMBH"],
        ToolExecutionContext.create([tmp_path]),
    )

    assert len(identities) == 2505


def test_cnipa_downfile_jsp_is_discovered_as_pdf_attachment() -> None:
    url = (
        "/module/download/downfile.jsp?classid=0&"
        "showname=award-list.pdf&filename=evidence.pdf"
    )
    _text, attachments, _images = parse_html(
        f'<meta name="Image" content="{url}">',
        "https://www.cnipa.gov.cn/art/2025/6/5/example.html",
    )

    assert len(attachments) == 1
    assert attachments[0].url.startswith(
        "https://www.cnipa.gov.cn/module/download/downfile.jsp"
    )


def test_submitted_identity_reader_disambiguates_same_title_rows(tmp_path: Path) -> None:
    path = tmp_path / "same-title.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XMMC", "XFZRXM", "XDWMC"])
    sheet.append(["项目名称", "负责人", "单位"])
    for index in range(4):
        sheet.append(["同名研究", f"负责人{index}", f"单位{index}"])
    workbook.save(path)

    identities = _submitted_match_items_from_paths(
        [path],
        ["XMMC", "XFZRXM", "XDWMC"],
        ToolExecutionContext.create([tmp_path]),
    )

    assert len(identities) == 4
    assert all("+" in field_code and ";" in display for field_code, _, display in identities)


def test_submitted_identity_reader_disambiguates_same_title_across_sheets(
    tmp_path: Path,
) -> None:
    path = tmp_path / "same-title-multi-sheet.xlsx"
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "第一批"
    first.append(["XMMC", "XFZRXM", "XDWMC"])
    first.append(["项目名称", "负责人", "单位"])
    first.append(["同名研究", "负责人甲", "单位甲"])
    second = workbook.create_sheet("第二批")
    second.append(["XMMC", "XFZRXM", "XDWMC"])
    second.append(["项目名称", "负责人", "单位"])
    second.append(["同名研究", "负责人乙", "单位乙"])
    workbook.save(path)
    workbook.close()

    identities = _submitted_match_items_from_paths(
        [path],
        ["XMMC", "XFZRXM", "XDWMC"],
        ToolExecutionContext.create([tmp_path]),
    )

    assert len(identities) == 2
    assert {display for _field, _key, display in identities} == {
        "同名研究;负责人甲;单位甲",
        "同名研究;负责人乙;单位乙",
    }


def test_verifier_aggregates_distinct_complete_pdf_siblings() -> None:
    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="05040003",
        award_name="教育部人文社会科学研究一般项目",
        year="2025",
        trigger_codes=["PDF_ONLY"],
        objective="核验同一公示页的多个 PDF 附件",
    ), ToolBudgetState())
    parent = "https://example.gov.cn/notice"
    results = [
        ToolResult(
            ok=True,
            data={
                "evidence_group": parent,
                "document_complete": True,
                "matched_identity_hashes": hashes,
            },
            evidence_facts=[_fact(
                url=f"{parent}/{index}.pdf",
                expected=4,
                observed=2,
                complete=False,
            )],
        )
        for index, hashes in enumerate((["a", "b"], ["c", "d"]), start=1)
    ]

    snapshot = build_evidence_snapshot(state, results)

    assert snapshot.expected_count == 4
    assert snapshot.observed_count == 4
    assert snapshot.explicit_coverage_complete is True


def _fact(
    *,
    url: str = "https://example.gov.cn/award",
    level: str = "official_primary",
    expected: int = 25,
    observed: int = 25,
    complete: bool = True,
) -> EvidenceFact:
    return EvidenceFact(
        status="complete" if complete else "partial",
        award_name="示例奖",
        year="2025",
        target_match="yes",
        year_match="yes",
        source_url=url,
        source_level=level,
        expected_count=expected,
        observed_count=observed,
        coverage_complete=complete,
        extraction_method="offline_fixture",
    )


def test_generic_case_scope_uses_submission_without_resource_profile() -> None:
    matched, mode = _match_award_title(
        "全国高校辅导员年度人物",
        "教育部办公厅关于开展2023年“最美高校辅导员”暨第十五届"
        "“高校辅导员年度人物”推选展示活动的通知 - 中华人民共和国教育部政府门户网站",
    )
    assert matched is True and mode == "scope_variant"
    adjacent_award, adjacent_mode = _match_award_title(
        "全国高校辅导员年度人物",
        "2023年“最美大学生”“最美高校辅导员”候选人公示_央广网",
    )
    assert adjacent_award is True and adjacent_mode == "semantic_core"
    different_role, different_role_mode = _match_award_title(
        "全国高校辅导员年度人物",
        "2023年全国高校教师年度人物候选人公示",
    )
    assert different_role is False and different_role_mode == "none"
    non_result, non_result_mode = _match_award_title(
        "全国高校辅导员年度人物",
        "2023年高校辅导员能力提升培训班开班通知",
    )
    assert non_result is False and non_result_mode == "none"

    state = CaseSeed(
        batch_id=1,
        resource_code="NEW-RESOURCE",
        award_name="全国高校辅导员年度人物",
        year="2023",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="核验完整业务口径",
        submitted_summary={
            "submitted_rows": 20,
            "reference_rows": 50,
            "submission_file": "submitted.xlsx",
            "match_fields": ["XRYXM"],
        },
    )
    from award_audit.agent.harness.models import AuditCaseState
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    case_state = AuditCaseState.from_seed(state, ToolBudgetState())
    arguments = _fetch_arguments(case_state, {"url": "https://edu.cnr.cn/example"})
    assert arguments["expected_scope_count"] == 20
    assert arguments["award_aliases"] == []
    assert arguments["section_keywords"] == []
    assert arguments["section_exclude_keywords"] == []
    assert arguments["relationship_terms"] == []
    collect_arguments = _collect_arguments(case_state, {})
    assert collect_arguments["include_attachment_keywords"] == []
    assert collect_arguments["exclude_attachment_keywords"] == []
    assert collect_arguments["expected_scope_count"] == 20

    snapshot = build_evidence_snapshot(case_state, [
        ToolResult(ok=True, evidence_facts=[_fact(expected=20, observed=20, complete=True)])
    ])
    report = deterministic_verify(snapshot)
    assert snapshot.expected_count == snapshot.observed_count == 20
    assert report.coverage_complete == "yes"


def test_fetch_page_compares_all_submission_files_in_one_case(
    tmp_path: Path,
    monkeypatch,
) -> None:  # noqa: ANN001
    submitted_paths = [tmp_path / "part-a.xlsx", tmp_path / "part-b.xlsx"]
    for path, name in zip(submitted_paths, ["张三", "李四"], strict=True):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["XRYXM"])
        sheet.append(["获奖人"])
        sheet.append([name])
        workbook.save(path)
        workbook.close()
    url = "https://official.example/award"
    monkeypatch.setattr(
        registry_module.web,
        "fetch_page",
        lambda *_args, **_kwargs: PageContent(
            url=url,
            status=200,
            title="2026年新奖项公示",
            text="2026年新奖项公示名单：张三、李四",
        ),
    )

    result = SafeToolExecutor(build_default_registry()).execute(
        "fetch_web_page",
        {
            "url": url,
            "expected_award_name": "新奖项",
            "expected_year": "2026",
            "submitted_paths": [str(path) for path in submitted_paths],
            "match_fields": ["XRYXM"],
            "expected_scope_count": 2,
        },
        ToolExecutionContext.create([tmp_path]),
    )

    assert result.ok
    assert result.data["submitted_count"] == 2
    assert result.data["matched_items"] == ["张三", "李四"]
    assert result.data["coverage_complete"] is True
    assert result.data["identity_version"] == "identity-v2"


def test_fetch_page_truncated_body_cannot_claim_complete_coverage(
    tmp_path: Path,
    monkeypatch,
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XRYXM"])
    sheet.append(["张三"])
    workbook.save(submitted)
    workbook.close()
    page = PageContent(
        url="https://example.gov.cn/award",
        status=200,
        title="2026年新奖项名单",
        text="新奖项 2026 张三",
        text_truncated=True,
        original_text_chars=50_000,
    )
    monkeypatch.setattr(registry_module.web, "fetch_page", lambda url: page)

    result = SafeToolExecutor(build_default_registry()).execute(
        "fetch_web_page",
        {
            "url": page.url,
            "expected_award_name": "新奖项",
            "expected_year": "2026",
            "submitted_paths": [str(submitted)],
            "match_fields": ["XRYXM"],
        },
        ToolExecutionContext.create([tmp_path]),
    )

    assert result.ok
    assert result.is_truncated is True
    assert result.data["text_truncated"] is True
    assert result.data["coverage_complete"] is False
    assert result.evidence_facts[0].coverage_complete is False
    assert "网页正文已截断，无法证明名单完整" in result.evidence_facts[0].missing_evidence


def test_fetch_page_matching_body_still_queues_complete_attachment_group(
    tmp_path: Path,
    monkeypatch,
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XRYXM"])
    sheet.append(["张三"])
    workbook.save(submitted)
    workbook.close()
    attachment_urls = [
        f"https://example.gov.cn/group-{index}.pdf" for index in range(1, 26)
    ]
    page = PageContent(
        url="https://example.gov.cn/award",
        status=200,
        title="2026年新奖项名单",
        text="新奖项 2026 张三",
        attachments=[
            Attachment(text=f"第{index}组名单.pdf", url=url, is_excel=False)
            for index, url in enumerate(attachment_urls, start=1)
        ],
    )
    monkeypatch.setattr(registry_module.web, "fetch_page", lambda url: page)

    result = SafeToolExecutor(build_default_registry()).execute(
        "fetch_web_page",
        {
            "url": page.url,
            "expected_award_name": "新奖项",
            "expected_year": "2026",
            "submitted_paths": [str(submitted)],
            "match_fields": ["XRYXM"],
        },
        ToolExecutionContext.create([tmp_path]),
    )

    assert result.ok
    assert result.data["coverage_complete"] is False
    assert result.data["next_evidence_stage"] == "spreadsheet_processing"
    assert result.data["candidate_attachment_urls"] == attachment_urls
    assert result.evidence_facts[0].coverage_complete is False


def test_fetch_page_matching_body_still_queues_all_discovered_images(
    tmp_path: Path,
    monkeypatch,
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XRYXM"])
    sheet.append(["张三"])
    workbook.save(submitted)
    workbook.close()
    image_urls = [
        f"https://example.gov.cn/roster-{index}.jpg" for index in range(1, 26)
    ]
    page = PageContent(
        url="https://example.gov.cn/award",
        status=200,
        title="2026年新奖项名单",
        text="新奖项 2026 张三",
        images=image_urls,
    )
    monkeypatch.setattr(registry_module.web, "fetch_page", lambda url: page)

    result = SafeToolExecutor(build_default_registry()).execute(
        "fetch_web_page",
        {
            "url": page.url,
            "expected_award_name": "新奖项",
            "expected_year": "2026",
            "submitted_paths": [str(submitted)],
            "match_fields": ["XRYXM"],
        },
        ToolExecutionContext.create([tmp_path]),
    )

    assert result.ok
    assert result.data["coverage_complete"] is False
    assert result.data["next_evidence_stage"] == "image_processing"
    assert result.data["candidate_image_urls"] == image_urls


def test_fetch_arguments_rejects_model_invented_relationship_terms() -> None:
    from award_audit.agent.harness.models import AuditCaseState
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="02050015",
        award_name="最美教师",
        year="2025",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="核验公开名单",
    ), ToolBudgetState())

    arguments = _fetch_arguments(state, {
        "url": "https://news.example.cn/list",
        "relationship_terms": ["名单", "公布", "发布"],
    })

    assert arguments["relationship_terms"] == []


def test_s01_image_arguments_do_not_leak_relationship_only_fields() -> None:
    from award_audit.agent.harness.models import AuditCaseState
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="06090003",
        award_name="全国高校辅导员年度人物",
        year="2023",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="核验页面图片名单",
        submitted_summary={
            "submission_file": "submitted.xlsx",
            "match_fields": ["XRYXM"],
            "submitted_rows": 20,
        },
    ), ToolBudgetState())
    state.evidence_progress.pending_media_source_url = "https://example.cn/page"
    state.evidence_progress.pending_media_page_title = "候选人公示"
    state.evidence_progress.pending_media_urls = ["https://example.cn/roster.jpg"]

    arguments = _image_roster_arguments(state, {
        "relationship_terms": ["不应传递到图片工具"],
    })

    assert "relationship_terms" not in arguments
    assert arguments["expected_scope_count"] == 20
    assert arguments["image_urls"] == ["https://example.cn/roster.jpg"]


def test_image_roster_processes_all_images_from_a_bounded_page() -> None:
    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="05060001",
        award_name="示例图片名单",
        year="2025",
        trigger_codes=["IMAGE_ONLY"],
        objective="核验多张名单图片",
        submitted_summary={
            "submission_file": "submitted.xlsx",
            "match_fields": ["XMMC"],
            "submitted_rows": 9,
        },
    ), ToolBudgetState())
    urls = [f"https://example.cn/roster-{index}.png" for index in range(9)]
    state.evidence_progress.pending_media_source_url = "https://example.cn/page"
    state.evidence_progress.pending_media_page_title = "名单"
    state.evidence_progress.pending_media_urls = urls

    arguments = _image_roster_arguments(state, {})
    processed = arguments["image_urls"]
    _consume_pending_media(state, processed)

    assert processed == urls[:6]
    assert state.evidence_progress.pending_media_urls == urls[6:]
    assert state.evidence_progress.pending_media_source_url == "https://example.cn/page"


def test_known_html_is_processed_before_discovered_page_images() -> None:
    from award_audit.agent.toolkit.testing import register_fake_tool

    page_url = "https://example.gov.cn/results"
    image_url = "https://example.gov.cn/results-1.png"
    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="COMPETITION",
        award_name="Example Competition",
        year="2025",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="verify all business roles",
        known_urls=[page_url, image_url],
    ), ToolBudgetState())
    state.evidence_progress.pending_media_source_url = page_url
    state.evidence_progress.pending_media_urls = [image_url]
    registry = ToolRegistry()
    register_fake_tool(registry, "fetch_web_page", [ToolResult(ok=True)])
    register_fake_tool(registry, "verify_page_image_roster", [ToolResult(ok=True)])

    decision = _deterministic_action(state, [], registry)

    assert decision is not None
    action, reason = decision
    assert action.tool_name == "fetch_web_page"
    assert action.arguments["url"] == page_url
    assert reason == "known_html_processed_before_linked_assets"


def test_pending_images_precede_complete_snapshot_corroboration() -> None:
    from award_audit.agent.toolkit.testing import register_fake_tool

    page_url = "https://official.example/notice"
    image_url = "https://official.example/page-2.png"
    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="MULTI-IMAGE",
        award_name="Example Award",
        year="2025",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="finish known images before search",
        known_urls=[page_url],
    ), ToolBudgetState())
    state.evidence_progress.pending_media_source_url = page_url
    state.evidence_progress.pending_media_urls = [image_url]
    state.tool_trace.append(ToolObservation(
        call_id="page-call", tool_name="fetch_web_page",
        started_at="2025-01-01T00:00:00Z", finished_at="2025-01-01T00:00:01Z",
        duration_ms=1000, input_summary={"url": page_url}, ok=True,
    ))
    complete = ToolResult(ok=True, evidence_facts=[EvidenceFact(
        status="complete", award_name="Example Award", year="2025",
        target_match="yes", year_match="yes", source_url=page_url,
        source_level="publisher_secondary", coverage_complete=True,
        document_complete=True,
    )])
    registry = ToolRegistry()
    register_fake_tool(registry, "verify_page_image_roster", [ToolResult(ok=True)])
    register_fake_tool(registry, "search_official_award", [ToolResult(ok=True)])

    decision = _deterministic_action(state, [complete], registry)

    assert decision is not None
    action, reason = decision
    assert action.tool_name == "verify_page_image_roster"
    assert reason == "pending_page_images_processed_without_agent_turn"


def test_generic_award_match_uses_shared_distinctive_core_without_profile() -> None:
    from award_audit.agent.harness.models import AuditCaseState
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="NEW-COMPETITION",
        award_name="中国研究生创新实践系列大赛：全国农科研究生志愿服务技能大赛",
        year="2023",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="核验四个赛事附件",
        submitted_summary={
            "submission_file": "submitted.xlsx",
            "match_fields": ["CSDWMC"],
            "attachment_match_fields": ["CSDWMC", "XCSDW"],
            "submitted_rows": 307,
            "reference_rows": 307,
        },
    ), ToolBudgetState())

    fetch = _fetch_arguments(state, {"url": "https://cpipc.acge.org.cn/page"})
    collect = _collect_arguments(state, {})

    assert fetch["award_aliases"] == []
    assert fetch["official_secondary_domains"] == []
    assert collect["award_aliases"] == []
    assert collect["official_secondary_domains"] == []
    matched, mode = _match_award_title(
        state.award_name,
        "第三届乡村振兴志愿服务技能大赛获奖名单",
    )
    assert matched is True and mode == "shared_core"


def test_s04_pdf_extraction_receives_trusted_case_scope_and_provenance() -> None:
    from award_audit.agent.harness.models import AuditCaseState
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    pdf_path = "tmp/evidence/award-roster.pdf"
    pdf_url = "https://cpipc.acge.org.cn/sysFile/downFile.do?fileId=pdf-1"
    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="04030060",
        award_name="中国研究生创新实践系列大赛：全国农科研究生志愿服务技能大赛",
        year="2023",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="核验 PDF 获奖名单",
        submitted_summary={
            "submission_file": "submitted.xlsx",
            "match_fields": ["CSDWMC", "XMMC"],
            "reference_rows": 325,
        },
    ), ToolBudgetState())
    state.artifacts.append(EvidenceArtifact(
        kind="pdf",
        source_url=pdf_url,
        local_path=pdf_path,
        content_type="application/pdf",
        sha256="a" * 64,
        size_bytes=1024,
        fetched_at="2026-07-29T00:00:00Z",
        metadata={
            "page_url": "https://cpipc.acge.org.cn/detail/award-1",
            "attachment_linked": True,
            "page_observed_award_name": state.award_name,
            "page_observed_year": "2023",
            "page_source_level": "official_secondary",
        },
    ))

    arguments = _pdf_extract_arguments(state, {"path": pdf_path, "pages": [1]})

    assert arguments["submitted_path"] == "submitted.xlsx"
    assert arguments["match_fields"] == ["CSDWMC", "XMMC"]
    assert arguments["expected_award_name"] == state.award_name
    assert arguments["award_aliases"] == []
    assert arguments["expected_year"] == "2023"
    assert arguments["expected_scope_count"] == 325
    assert arguments["source_url"] == pdf_url
    assert arguments["official_secondary_domains"] == []
    assert arguments["parent_attachment_linked"] is True
    assert arguments["parent_award_name"] == state.award_name
    assert arguments["parent_year"] == "2023"
    assert arguments["parent_source_level"] == "official_secondary"


def test_generic_first_search_rejects_model_supplied_trust_metadata() -> None:
    from award_audit.agent.harness.models import AuditCaseState
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="NEW-RESOURCE",
        award_name="全国高校辅导员年度人物",
        year="2023",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="查找带附件的公开名单",
    ), ToolBudgetState())

    arguments = _search_arguments(state, {
        "strategy": "site",
        "official_domains": ["moe.gov.cn", "univs.cn"],
        "organizer": "教育部",
    })

    assert arguments["strategy"] == "broad"
    assert arguments["official_domains"] == []
    assert arguments["official_secondary_domains"] == []
    assert arguments["require_award_name_match"] is True
    assert "organizer" not in arguments


def test_generic_non_source_failure_does_not_force_same_site_search() -> None:
    from award_audit.agent.harness.models import AuditCaseState
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="NEW-RESOURCE",
        award_name="全国青年公益创新竞赛",
        year="2025",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="在给定来源不足时查找其他同奖项来源",
        known_urls=["https://publisher.example.org/adjacent-award"],
    ), ToolBudgetState())
    state.evidence_progress.source_failures = 1

    arguments = _search_arguments(state, {})

    assert arguments["strategy"] == "broad"
    assert "site_domains" not in arguments


def test_second_generic_search_uses_attachment_strategy() -> None:
    from award_audit.agent.harness.models import AuditCaseState
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="NEW-RESOURCE",
        award_name="全国高校辅导员年度人物",
        year="2023",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="宽泛搜索后继续查找公开附件",
        known_urls=["https://publisher.example.org/award"],
    ), ToolBudgetState())
    state.evidence_progress.search_round = 1

    arguments = _search_arguments(state, {})

    assert arguments["strategy"] == "attachment"
    assert arguments["max_results"] == 8
    assert "site_domains" not in arguments


def test_generic_failed_source_search_uses_known_host_without_trusting_it() -> None:
    from award_audit.agent.harness.models import AuditCaseState
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="NEW-RESOURCE",
        award_name="全国教材建设奖",
        year="2025",
        trigger_codes=["SOURCE_UNREACHABLE"],
        objective="恢复失效的教育部获奖名单来源",
        known_urls=["http://www.moe.gov.cn/old-award.pdf"],
    ), ToolBudgetState())
    state.evidence_progress.source_failures = 1
    state.tool_trace.append(ToolObservation(
        call_id="failed-source",
        tool_name="fetch_web_page",
        started_at="2025-01-01T00:00:00Z",
        finished_at="2025-01-01T00:00:01Z",
        duration_ms=1000,
        input_summary={"url": "http://www.moe.gov.cn/old-award.pdf"},
        ok=False,
        error_code="HTTP_ERROR",
    ))

    arguments = _search_arguments(state, {"strategy": "attachment"})

    assert arguments["strategy"] == "site"
    assert arguments["site_domains"] == ["moe.gov.cn"]
    assert arguments["official_domains"] == []
    assert arguments["require_award_name_match"] is True
    assert arguments["recovery_terms"] == ["old-award.pdf"]


def test_second_failed_source_search_broadens_beyond_original_domain() -> None:
    from award_audit.agent.harness.models import AuditCaseState
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    failed_url = "https://www.example.gov.cn/archive/t20250905_1411955.html"
    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="NEW-RESOURCE",
        award_name="示例教师团队",
        year="2025",
        trigger_codes=["SOURCE_UNREACHABLE"],
        objective="原域名恢复失败后扩大公开来源",
        known_urls=[failed_url],
    ), ToolBudgetState())
    state.evidence_progress.search_round = 1
    state.evidence_progress.source_failures = 2
    failed_trace = ToolObservation(
        call_id="failed-source",
        tool_name="fetch_web_page",
        started_at="2025-01-01T00:00:00Z",
        finished_at="2025-01-01T00:00:01Z",
        duration_ms=1000,
        input_summary={"url": failed_url},
        ok=False,
        error_code="HTTP_ERROR",
    )
    state.tool_trace.append(failed_trace)

    arguments = _search_arguments(state, {})

    assert arguments["strategy"] == "broad"
    assert "site_domains" not in arguments
    assert arguments["recovery_terms"] == ["t20250905_1411955.html"]


def test_later_search_drops_document_id_after_exact_recovery_was_attempted() -> None:
    from award_audit.agent.harness.models import AuditCaseState
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    original = "https://www.example.gov.cn/archive/t20250905_1411955.html"
    recovered = "https://mirror.example.gov.cn/t20250905_1411955.html"
    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="NEW-RESOURCE",
        award_name="未登记教师团队",
        year="2025",
        trigger_codes=["SOURCE_UNREACHABLE"],
        objective="精确恢复耗尽后跨来源寻找可解析正文",
        known_urls=[original],
    ), ToolBudgetState())
    state.evidence_progress.search_round = 1
    state.evidence_progress.source_failures = 2
    state.evidence_progress.candidates = [EvidenceCandidate(
        url=recovered,
        status="succeeded",
        attempts=2,
    )]
    state.tool_trace.append(ToolObservation(
        call_id="failed-source",
        tool_name="fetch_web_page",
        started_at="2025-01-01T00:00:00Z",
        finished_at="2025-01-01T00:00:01Z",
        duration_ms=1000,
        input_summary={"url": original},
        ok=False,
        error_code="HTTP_ERROR",
    ))

    arguments = _search_arguments(state, {})

    assert arguments["strategy"] == "broad"
    assert "site_domains" not in arguments
    assert "recovery_terms" not in arguments


def test_exhausted_first_recovery_queue_stops_after_bounded_search() -> None:
    from award_audit.agent.harness.models import AuditCaseState
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="NEW-RESOURCE",
        award_name="示例教师团队",
        year="2025",
        trigger_codes=["SOURCE_UNREACHABLE"],
        objective="首轮恢复失败后继续跨域搜索",
        known_urls=["https://www.example.gov.cn/old-page"],
    ), ToolBudgetState())
    state.evidence_progress.search_round = 1
    state.evidence_progress.source_failures = 2
    state.evidence_progress.candidates = [EvidenceCandidate(
        url="https://files.example.gov.cn/recovered-page",
        status="succeeded",
        attempts=1,
    )]
    registry = ToolRegistry()
    register_fake_tool(
        registry,
        "search_official_award",
        [ToolResult(ok=True)],
    )

    decision = _deterministic_action(state, [], registry)

    assert decision is not None
    action, reason = decision
    assert action.action == "manual"
    assert reason == "bounded_search_candidates_exhausted"


def test_forced_next_recovery_round_reaches_executor(tmp_path: Path) -> None:
    store = Store(tmp_path / "forced-next-search.db")
    batch_id = store.create_batch("forced-next-search")
    repository = CaseRepository(store)
    recovered_url = "https://publisher.example.cn/current-result"
    state, _created = repository.create_or_get(CaseSeed(
        batch_id=batch_id,
        resource_code="NEW-RESOURCE",
        award_name="示例教师团队",
        year="2025",
        trigger_codes=["SOURCE_UNREACHABLE"],
        objective="首轮候选失败后必须执行下一轮搜索",
        known_urls=["https://www.example.gov.cn/old-page"],
    ))
    state.reason_codes = ["official_search_candidates_ready"]
    state.evidence_progress.search_round = 1
    state.evidence_progress.source_failures = 2
    state.evidence_progress.candidates = [EvidenceCandidate(
        url="https://files.example.gov.cn/failed-page",
        status="succeeded",
        attempts=1,
    )]
    failed_trace = ToolObservation(
        call_id="failed-source",
        tool_name="fetch_web_page",
        started_at="2025-01-01T00:00:00Z",
        finished_at="2025-01-01T00:00:01Z",
        duration_ms=1000,
        input_summary={"url": "https://www.example.gov.cn/old-page"},
        ok=False,
        error_code="HTTP_ERROR",
    )
    state.tool_trace.append(failed_trace)
    repository.save(state, traces=[failed_trace])
    registry = ToolRegistry()
    searched = register_fake_tool(registry, "search_official_award", [ToolResult(
        ok=True,
        data={
            "provider": "offline",
            "strategy": "broad",
            "candidate_count": 1,
            "official_candidate_count": 0,
            "candidates": [{
                "url": recovered_url,
                "source_level": "publisher_secondary",
                "provider": "offline",
                "rank": 1,
                "title": "示例教师团队 2025 结果名单",
            }],
        },
    )])
    fetched = register_fake_tool(registry, "fetch_web_page", [ToolResult(
        ok=True,
        source_url=recovered_url,
        evidence_facts=[_fact(url=recovered_url, level="publisher_secondary")],
    )])

    outcome = EvidenceHarness(
        repository=repository,
        registry=registry,
        agent_client=FakeAgentClient([]),
        allowed_roots=[tmp_path],
        verifier=EvidenceVerifier(),
    ).run(state.case_id)

    assert len(searched.calls) == 0
    assert len(fetched.calls) == 0
    assert outcome.state.status == "waiting_human"
    store.close()


def test_collect_arguments_filters_scope_metadata_to_tool_contract(tmp_path: Path) -> None:
    submission = tmp_path / "submission.xlsx"
    submission.touch()
    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="SCOPE-TOOL",
        award_name="Example Award",
        year="2025",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="Process a routed PDF",
        submitted_summary={
            "submission_files": [str(submission)],
            "match_fields": ["ZPMC"],
            "role_scopes": [{
                "scope_id": 7,
                "role_type": "team",
                "required": True,
                "business_scope": {"category": "Final"},
                "submitted_identity_count": 2,
                "profile": {
                    "primary_alternatives": [["ZPMC"]],
                    "section_include_terms": ["Winning Teams"],
                    "section_exclude_terms": ["Organization Award"],
                },
            }],
        },
    ), ToolBudgetState())
    url = "https://example.gov.cn/results.pdf"
    state.evidence_progress.pending_attachment_page_urls = [
        "https://example.gov.cn/results"
    ]
    state.evidence_progress.pending_attachment_urls = [url]
    state.evidence_progress.pending_attachment_parent_urls = {
        url: "https://example.gov.cn/results"
    }

    arguments = _collect_arguments(state, {})

    assert "section_keywords" not in arguments
    assert "section_exclude_keywords" not in arguments
    assert CollectSpreadsheetAttachmentsInput.model_validate(arguments)


def test_image_attempt_does_not_hide_unfetched_parent_page() -> None:
    parent = "https://example.gov.cn/results"
    image = "https://example.gov.cn/results-1.png"
    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="KNOWN-PAGE",
        award_name="Example Award",
        year="2025",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="Fetch the parent page",
        known_urls=[parent, image],
    ), ToolBudgetState())
    state.tool_trace.append(ToolObservation(
        call_id="image-call",
        tool_name="verify_page_image_roster",
        started_at="2025-01-01T00:00:00Z",
        finished_at="2025-01-01T00:00:01Z",
        duration_ms=1000,
        input_summary={"page_url": parent, "image_urls": [image]},
        ok=True,
    ))

    assert _next_unattempted_known_url(state) == parent


def test_html_body_routes_distinct_role_sections_without_refetch() -> None:
    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="MULTI-ROLE-WEB",
        award_name="Example Award",
        year="2025",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="Route one HTML page to multiple roles",
        submitted_summary={
            "role_scopes": [
                {
                    "scope_id": 11,
                    "scope_key": "team:final",
                    "role_type": "team",
                    "required": True,
                    "submitted_identities": {"team-a": "Team A"},
                    "profile": {"section_include_terms": ["Winning Teams"]},
                },
                {
                    "scope_id": 12,
                    "scope_key": "organization:final",
                    "role_type": "organization",
                    "required": True,
                    "submitted_identities": {"org-a": "University A"},
                    "profile": {"section_include_terms": ["Organization Units"]},
                },
            ],
        },
    ), ToolBudgetState())
    result = ToolResult(
        ok=True,
        source_url="https://example.gov.cn/results",
        data={
            "text": "Winning Teams\nTeam A\nOrganization Units\nUniversity A",
        },
        evidence_facts=[EvidenceFact(
            status="partial",
            award_name="Example Award",
            year="2025",
            target_match="yes",
            year_match="yes",
            source_url="https://example.gov.cn/results",
            source_level="official_primary",
            matched_items=["Team A"],
        )],
    )

    _route_web_result_to_scopes(state, result)

    assert [(fact.scope_id, fact.matched_items) for fact in result.evidence_facts] == [
        (11, ["Team A"]),
        (12, ["University A"]),
    ]
    assert all(fact.document_complete for fact in result.evidence_facts)


def test_extract_search_arguments_match_the_strict_tool_contract() -> None:
    from award_audit.agent.harness.models import AuditCaseState
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="03020004",
        award_name="全国教材建设奖",
        year="2025",
        trigger_codes=["SOURCE_UNREACHABLE"],
        objective="提取搜索候选页面",
        submitted_summary={
            "submission_file": "submitted.xlsx",
            "match_fields": ["ISBN", "JCMC"],
            "reference_rows": 989,
        },
    ), ToolBudgetState())

    arguments = _extract_arguments(state, {
        "url": "https://www.moe.gov.cn/recovered.pdf",
        "official_domains": ["moe.gov.cn"],
        "official_secondary_domains": ["example.edu.cn"],
        "section_exclude_keywords": ["拟推荐"],
    })

    validated = ExtractSearchDocumentInput.model_validate(arguments)
    assert validated.expected_award_name == "全国教材建设奖"
    assert validated.expected_year == "2025"
    assert validated.expected_scope_count == 989
    assert "official_domains" not in arguments
    assert "official_secondary_domains" not in arguments
    assert "section_exclude_keywords" not in arguments


def test_extract_search_query_uses_candidate_title_and_case_count() -> None:
    from award_audit.agent.harness.models import AuditCaseState
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="NEW-RESOURCE",
        award_name="全国高校黄大年式教师团队",
        year="2025",
        trigger_codes=["SOURCE_UNREACHABLE"],
        objective="恢复教育部名单正文",
        submitted_summary={
            "submission_file": "submitted.xlsx",
            "match_fields": ["TDMC"],
            "submitted_rows": 190,
            "reference_rows": 190,
        },
    ), ToolBudgetState())
    candidate_url = (
        "http://www.moe.gov.cn/jyb_xxgk/s5743/s5744/A10/202509/"
        "t20250905_1411955.html"
    )
    state.evidence_progress.candidates.append(EvidenceCandidate(
        url=candidate_url,
        source_level="official_secondary",
        provider="fake",
        rank=1,
        title="第三批认定名单和第四批入围名单",
    ))

    arguments = _extract_arguments(state, {
        "url": candidate_url,
    })

    assert arguments["search_query"] == (
        "site:moe.gov.cn 全国高校黄大年式教师团队 2025 "
        "第三批认定名单和第四批入围名单 190项 名单公示"
    )


def test_s01_image_roster_selects_target_section_and_excludes_nomination_group(
    tmp_path: Path,
    monkeypatch,
) -> None:
    submitted = tmp_path / "submitted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XRYXM"])
    sheet.append(["入选人姓名"])
    sheet.append(["王奇"])
    sheet.append(["杨帆"])
    workbook.save(submitted)
    workbook.close()

    image_urls = [f"https://edu.cnr.cn/group-{index}.png" for index in range(1, 5)]
    fixture = Path("tests/data/m5_golden/vision/clean_roster.png")

    def fake_download(url, destination, **_kwargs):  # noqa: ANN001
        target = Path(destination) / Path(url).name
        shutil.copyfile(fixture, target)
        return target

    monkeypatch.setattr(registry_module.web, "download_file", fake_download)

    responses = iter([
        {
            "section_title": "2023年最美大学生候选人",
            "entries": [{"no": 1, "name": "学生甲"}],
        },
        {
            "section_title": "2023年最美高校辅导员候选人",
            "entries": [{"no": 1, "name": "辅导员甲"}],
        },
        {
            "section_title": "第十五届高校辅导员年度人物候选人",
            "entries": [
                {"no": 1, "name": "王奇"},
                {"no": 2, "name": "杨帆"},
            ],
        },
        {
            "section_title": "第十五届高校辅导员年度人物提名候选人",
            "entries": [{"no": 1, "name": "提名甲"}],
        },
    ])

    class Client:
        provider = "fake"
        model = "fake-vision"

        def __init__(self) -> None:
            self.page = 0

        def vision_json_call(self, *_args, **_kwargs):  # noqa: ANN002, ANN003, ANN202
            self.page += 1
            payload = next(responses)
            entries = payload["entries"]
            return {
                "page": self.page,
                "total_pages": 4,
                "is_roster_page": True,
                "headers": ["序号", "姓名"],
                **payload,
                "first_no": 1,
                "last_no": len(entries),
                "truncated": False,
                "unreadable": [],
                "confidence": 0.99,
            }

    executor = SafeToolExecutor(build_default_registry(
        vision_client_factory=Client,
    ))
    result = executor.execute(
        "verify_page_image_roster",
        {
            "page_url": "https://edu.cnr.cn/native/gd/20230511/t20230511_526247128.shtml",
            "page_title": "2023年最美大学生、最美高校辅导员候选人公示",
            "image_urls": image_urls,
            "submitted_path": str(submitted),
            "match_fields": ["XRYXM"],
            "expected_award_name": "全国高校辅导员年度人物",
            "expected_year": "2023",
            "expected_scope_count": 2,
            "destination_dir": str(tmp_path),
        },
        ToolExecutionContext.create([tmp_path]),
    )

    assert result.ok
    assert result.data["selected_sections"] == ["第十五届高校辅导员年度人物候选人"]
    assert result.data["matched_items"] == ["王奇", "杨帆"]
    assert result.data["missing_items"] == []
    assert result.data["extra_items"] == []
    assert result.data["coverage_complete"] is True
    assert result.data["processed_image_urls"] == image_urls
    assert result.data["failed_image_urls"] == []
    assert result.data["unprocessed_image_urls"] == []
    assert result.data["all_images_processed"] is True
    assert result.evidence_facts[0].comparison_scope == "全国高校辅导员年度人物"


def test_image_roster_continuation_marker_does_not_fail_readable_asset(
    tmp_path: Path,
    monkeypatch,
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted-continuation.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XMMC"])
    sheet.append(["Project name"])
    sheet.append(["Project 12"])
    workbook.save(submitted)
    workbook.close()

    image_url = "https://example.gov.cn/roster-part-2.png"
    fixture = Path("tests/data/m5_golden/vision/clean_roster.png")

    def fake_download(url, destination, **_kwargs):  # noqa: ANN001
        target = Path(destination) / Path(url).name
        shutil.copyfile(fixture, target)
        return target

    monkeypatch.setattr(registry_module.web, "download_file", fake_download)

    class Client:
        provider = "fake"
        model = "fake-vision"

        def vision_json_call(self, *_args, **_kwargs):  # noqa: ANN002, ANN003, ANN202
            return {
                "page": 1,
                "total_pages": 1,
                "is_roster_page": True,
                "section_title": "2025 Example Award roster",
                "headers": ["No.", "Project"],
                "entries": [{"no": 12, "name": "Project 12"}],
                "first_no": 12,
                "last_no": 12,
                # Real providers use this to indicate that the list continues.
                "truncated": True,
                "unreadable": [],
                "confidence": 0.99,
            }

    result = SafeToolExecutor(build_default_registry(
        vision_client_factory=Client,
    )).execute(
        "verify_page_image_roster",
        {
            "page_url": "https://example.gov.cn/notice",
            "page_title": "2025 Example Award roster",
            "image_urls": [image_url],
            "submitted_path": submitted,
            "match_fields": ["XMMC"],
            "expected_award_name": "Example Award",
            "expected_year": "2025",
            "expected_scope_count": 1,
            "destination_dir": tmp_path,
        },
        ToolExecutionContext.create([tmp_path]),
    )

    assert result.ok
    assert result.data["processed_image_urls"] == [image_url]
    assert result.data["failed_image_urls"] == []
    assert result.data["all_images_processed"] is True
    assert result.data["matched_items"] == ["Project 12"]


def test_image_roster_retries_only_failed_vision_page(tmp_path: Path, monkeypatch) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted-retry.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XMMC"])
    sheet.append(["Project name"])
    sheet.append(["Project Retry"])
    workbook.save(submitted)
    workbook.close()
    fixture = Path("tests/data/m5_golden/vision/clean_roster.png")
    image_url = "https://example.gov.cn/retry.png"

    def fake_download(_url, destination, **_kwargs):  # noqa: ANN001, ANN202
        target = Path(destination) / "retry.png"
        shutil.copyfile(fixture, target)
        return target

    monkeypatch.setattr(registry_module.web, "download_file", fake_download)
    calls = {"count": 0}

    class Client:
        provider = "fake"
        model = "fake-vision"

        def vision_json_call(self, *_args, **_kwargs):  # noqa: ANN002, ANN003, ANN202
            calls["count"] += 1
            if calls["count"] <= 3:
                return {"entries": "invalid"}
            return {
                "page": 1,
                "total_pages": 1,
                "is_roster_page": True,
                "section_title": "2025 Example Award roster",
                "entries": [{"no": 1, "name": "Project Retry"}],
                "first_no": 1,
                "last_no": 1,
                "truncated": False,
                "unreadable": [],
                "confidence": 0.99,
            }

    context = ToolExecutionContext.create(
        [tmp_path], ToolBudgetLimits(max_vision_pages=2)
    )
    result = SafeToolExecutor(build_default_registry(
        vision_client_factory=Client,
    )).execute(
        "verify_page_image_roster",
        {
            "page_url": "https://example.gov.cn/notice",
            "page_title": "2025 Example Award roster",
            "image_urls": [image_url],
            "submitted_path": submitted,
            "match_fields": ["XMMC"],
            "expected_award_name": "Example Award",
            "expected_year": "2025",
            "expected_scope_count": 1,
            "destination_dir": tmp_path,
        },
        context,
    )

    assert result.ok
    assert calls["count"] == 4
    assert context.budget.vision_pages == 2
    assert result.data["processed_image_urls"] == [image_url]
    assert result.data["failed_image_urls"] == []
    assert result.data["vision_error_count"] == 0


def test_empty_roster_page_is_failed_not_processed(tmp_path: Path, monkeypatch) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted-empty.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XMMC"])
    sheet.append(["Project name"])
    sheet.append(["Expected Project"])
    workbook.save(submitted)
    workbook.close()
    fixture = Path("tests/data/m5_golden/vision/clean_roster.png")
    image_url = "https://example.gov.cn/empty-roster.png"

    def fake_download(_url, destination, **_kwargs):  # noqa: ANN001, ANN202
        target = Path(destination) / "empty-roster.png"
        shutil.copyfile(fixture, target)
        return target

    monkeypatch.setattr(registry_module.web, "download_file", fake_download)

    class Client:
        provider = "fake"
        model = "fake-vision"

        def vision_json_call(self, *_args, **_kwargs):  # noqa: ANN002, ANN003, ANN202
            return {
                "page": 1,
                "total_pages": 1,
                "is_roster_page": True,
                "section_title": "2025 Example Award roster",
                "entries": [],
                "truncated": False,
                "unreadable": [],
                "confidence": 0.9,
            }

    result = SafeToolExecutor(build_default_registry(
        vision_client_factory=Client,
    )).execute(
        "verify_page_image_roster",
        {
            "page_url": "https://example.gov.cn/notice",
            "page_title": "2025 Example Award roster",
            "image_urls": [image_url],
            "submitted_path": submitted,
            "match_fields": ["XMMC"],
            "expected_award_name": "Example Award",
            "expected_year": "2025",
            "expected_scope_count": 1,
            "destination_dir": tmp_path,
        },
        ToolExecutionContext.create([tmp_path]),
    )

    assert result.ok
    assert result.data["processed_image_urls"] == []
    assert result.data["failed_image_urls"] == [image_url]
    assert result.data["all_images_processed"] is False


def test_roster_page_with_sequence_gap_is_failed(tmp_path: Path, monkeypatch) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted-gap.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XMMC"])
    sheet.append(["Project name"])
    sheet.append(["Project 1"])
    workbook.save(submitted)
    workbook.close()
    fixture = Path("tests/data/m5_golden/vision/clean_roster.png")
    image_url = "https://example.gov.cn/gapped-roster.png"

    def fake_download(_url, destination, **_kwargs):  # noqa: ANN001, ANN202
        target = Path(destination) / "gapped-roster.png"
        shutil.copyfile(fixture, target)
        return target

    monkeypatch.setattr(registry_module.web, "download_file", fake_download)

    class Client:
        provider = "fake"
        model = "fake-vision"

        def vision_json_call(self, *_args, **_kwargs):  # noqa: ANN002, ANN003, ANN202
            return {
                "page": 1,
                "total_pages": 1,
                "is_roster_page": True,
                "section_title": "2025 Example Award roster",
                "entries": [
                    {"no": 1, "name": "Project 1"},
                    {"no": 3, "name": "Project 3"},
                ],
                "first_no": 1,
                "last_no": 3,
                "truncated": False,
                "unreadable": [],
                "confidence": 0.9,
            }

    result = SafeToolExecutor(build_default_registry(
        vision_client_factory=Client,
    )).execute(
        "verify_page_image_roster",
        {
            "page_url": "https://example.gov.cn/notice",
            "page_title": "2025 Example Award roster",
            "image_urls": [image_url],
            "submitted_path": submitted,
            "match_fields": ["XMMC"],
            "expected_award_name": "Example Award",
            "expected_year": "2025",
            "expected_scope_count": 1,
            "destination_dir": tmp_path,
        },
        ToolExecutionContext.create([tmp_path]),
    )

    assert result.ok
    assert result.data["processed_image_urls"] == []
    assert result.data["failed_image_urls"] == [image_url]
    assert result.data["image_page_summaries"][0]["row_count_complete"] is False


def test_image_roster_filters_decorations_and_batches_more_than_twenty_pages(
    tmp_path: Path,
    monkeypatch,
) -> None:  # noqa: ANN001
    from PIL import Image

    submitted = tmp_path / "submitted-many.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XMMC"])
    sheet.append(["项目名称"])
    for index in range(1, 24):
        sheet.append([f"项目{index}"])
    workbook.save(submitted)
    workbook.close()

    roster_fixture = Path("tests/data/m5_golden/vision/clean_roster.png")
    image_urls = [f"https://example.gov.cn/page-{index}.png" for index in range(1, 24)]
    image_urls.extend(
        f"https://example.gov.cn/decor-{index}.png" for index in range(1, 4)
    )

    def fake_download(url, destination, **_kwargs):  # noqa: ANN001, ANN202
        target = Path(destination) / Path(url).name
        if "decor" in url:
            Image.new("RGB", (75, 30), "white").save(target)
        else:
            shutil.copyfile(roster_fixture, target)
        return target

    monkeypatch.setattr(registry_module.web, "download_file", fake_download)
    counter = {"page": 0}

    class Client:
        provider = "fake"
        model = "fake-vision"

        def vision_json_call(self, *_args, **_kwargs):  # noqa: ANN002, ANN003, ANN202
            counter["page"] += 1
            page = counter["page"]
            return {
                "page": page,
                "total_pages": 23,
                "is_roster_page": True,
                "section_title": "2025年示例项目名单",
                "entries": [{"no": page, "name": f"项目{page}"}],
                "first_no": page,
                "last_no": page,
                "truncated": False,
                "unreadable": [],
                "confidence": 0.99,
            }

    result = SafeToolExecutor(build_default_registry(
        vision_client_factory=Client,
    )).execute(
        "verify_page_image_roster",
        {
            "page_url": "https://example.gov.cn/notice",
            "page_title": "2025年示例项目名单",
            "image_urls": image_urls,
            "submitted_path": submitted,
            "match_fields": ["XMMC"],
            "expected_award_name": "示例项目",
            "expected_year": "2025",
            "expected_scope_count": 23,
            "destination_dir": tmp_path,
        },
        ToolExecutionContext.create(
            [tmp_path],
            ToolBudgetLimits(max_downloads=30, max_vision_pages=30),
        ),
    )

    assert result.ok
    assert counter["page"] == 23
    assert len(result.artifacts) == 26
    assert result.evidence_facts[0].document_count == 23
    assert result.evidence_facts[0].coverage_complete is True


def test_vision_project_and_unit_fields_map_to_unified_identity(
    tmp_path: Path,
    monkeypatch,
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted-project.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XMMC"])
    sheet.append(["Project name"])
    sheet.append(["Project Alpha"])
    workbook.save(submitted)
    workbook.close()
    fixture = Path("tests/data/m5_golden/vision/clean_roster.png")

    def fake_download(_url, destination, **_kwargs):  # noqa: ANN001, ANN202
        target = Path(destination) / "project.png"
        shutil.copyfile(fixture, target)
        return target

    monkeypatch.setattr(registry_module.web, "download_file", fake_download)

    class Client:
        provider = "fake"
        model = "fake-vision"

        def vision_json_call(self, *_args, **_kwargs):  # noqa: ANN002, ANN003, ANN202
            return {
                "page": 1,
                "total_pages": 1,
                "is_roster_page": True,
                "section_title": "General projects",
                "entries": [{
                    "no": 1,
                    "project_name": "Project Alpha",
                    "unit": "University A",
                    "award_level": "General",
                }],
                "first_no": 1,
                "last_no": 1,
                "truncated": False,
                "unreadable": [],
                "confidence": 0.99,
            }

    result = SafeToolExecutor(build_default_registry(
        vision_client_factory=Client,
    )).execute(
        "verify_page_image_roster",
        {
            "page_url": "https://example.gov.cn/results",
            "page_title": "2025 Example Award",
            "image_urls": ["https://example.gov.cn/project.png"],
            "submitted_path": submitted,
            "match_fields": ["XMMC"],
            "expected_award_name": "Example Award",
            "expected_year": "2025",
            "expected_scope_count": 1,
            "destination_dir": tmp_path,
        },
        ToolExecutionContext.create([tmp_path]),
    )

    assert result.ok
    assert result.data["matched_items"] == ["Project Alpha"]
    assert result.data["identity_records"][0]["org"] == "University A"
    assert result.data["failed_image_urls"] == []


def test_generic_image_roster_arguments_follow_tool_contract() -> None:
    from award_audit.agent.harness.models import AuditCaseState
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="NEW-RESOURCE",
        award_name="全国青年公益创新竞赛",
        year="2025",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="核验网页图片名单",
    ), ToolBudgetState())
    state.submitted_summary = {
        "submission_file": "submitted.xlsx",
        "match_fields": ["XM"],
        "submitted_rows": 4,
        "official_domains": ["example.gov.cn"],
    }
    state.evidence_progress.pending_media_source_url = "https://example.org/page"
    state.evidence_progress.pending_media_page_title = "2025年获奖名单"
    state.evidence_progress.pending_media_urls = ["https://example.org/roster.jpg"]

    arguments = _image_roster_arguments(state, {})

    validated = VerifyPageImageRosterInput.model_validate(arguments)
    assert validated.expected_scope_count == 4
    assert arguments["official_domains"] == ["example.gov.cn"]
    assert arguments["official_secondary_domains"] == []


def test_generic_scope_excludes_named_sibling_categories() -> None:
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    generic = {
        "scope_id": 18,
        "role_type": "work_or_project",
        "required": True,
        "profile": {"primary_alternatives": [["XMBH"], ["XMMC"]]},
        "business_scope": {"ZYLBM": "05040003"},
        "submitted_identity_count": 248,
    }
    planning = {
        "scope_id": 20,
        "role_type": "work_or_project",
        "required": True,
        "profile": {"primary_alternatives": [["XMBH"], ["XMMC"]]},
        "business_scope": {"ZYLBM": "05040003", "XMLB": "Planning Fund"},
        "submitted_identity_count": 1259,
    }
    youth = {
        "scope_id": 21,
        "role_type": "work_or_project",
        "required": True,
        "profile": {"primary_alternatives": [["XMBH"], ["XMMC"]]},
        "business_scope": {"ZYLBM": "05040003", "XMLB": "Youth Fund"},
        "submitted_identity_count": 1747,
    }
    state = AuditCaseState.from_seed(CaseSeed(
        batch_id=1,
        resource_code="05040003",
        award_name="Example Research Award",
        year="2025",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="Compare scoped research projects",
        submitted_summary={
            "submission_file": "submitted.xlsx",
            "match_fields": ["XMBH", "XMMC"],
            "role_scopes": [generic, planning, youth],
        },
    ), ToolBudgetState())
    state.artifacts = [EvidenceArtifact(
        kind="pdf",
        source_url="https://example.gov.cn/west.pdf",
        local_path="west.pdf",
        content_type="application/pdf",
        sha256="a" * 64,
        size_bytes=1,
        fetched_at="2025-01-01T00:00:00Z",
        metadata={"label": "Western projects"},
    )]

    arguments = _pdf_extract_arguments(state, {
        "path": "west.pdf",
        "pages": [1],
        "scope_id": 18,
    })

    assert arguments["expected_scope_count"] == 248
    assert arguments["submitted_scope_filter"] == {}
    assert arguments["submitted_scope_exclude"] == {
        "XMLB": ["Planning Fund", "Youth Fund"]
    }


def test_media_queue_retains_failed_and_unprocessed_images() -> None:
    from award_audit.agent.harness.models import AuditCaseState
    from award_audit.agent.harness.runner import _update_media_queue_after_verification
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    state = AuditCaseState.from_seed(
        CaseSeed(
            batch_id=1,
            resource_code="NEW-RESOURCE",
            award_name="示例奖",
            year="2025",
            trigger_codes=["IMAGE_ONLY"],
            objective="核验图片名单",
        ),
        ToolBudgetState(),
    )
    urls = [f"https://example.gov.cn/roster-{index}.jpg" for index in range(1, 4)]
    state.evidence_progress.pending_media_source_url = "https://example.gov.cn/page"
    state.evidence_progress.pending_media_urls = urls

    _update_media_queue_after_verification(
        state,
        ToolResult(
            ok=True,
            data={
                "processed_image_urls": [urls[0]],
                "failed_image_urls": [urls[1]],
                "unprocessed_image_urls": [urls[2]],
                "all_images_processed": False,
            },
        ),
        urls,
    )

    assert state.evidence_progress.pending_media_urls == [urls[2]]
    assert state.evidence_progress.media_failed_urls == [urls[1]]
    assert state.evidence_progress.pending_media_source_url


def test_s02_auto_approval_requires_strict_official_complete_policy() -> None:
    secondary = deterministic_verify(build_evidence_snapshot_from_fact(
        _fact(level="publisher_secondary")
    ))
    official = deterministic_verify(build_evidence_snapshot_from_fact(_fact()))
    policy = AutoApprovalPolicy(enabled=True)

    assert decide_review_route(secondary, policy) == "waiting_human"
    assert decide_review_route(official, policy) == "auto_approve"
    assert decide_review_route(official, AutoApprovalPolicy(enabled=False)) == "waiting_human"


def test_s02_strict_policy_is_explicitly_wired_to_completed_state(tmp_path: Path) -> None:
    store = Store(tmp_path / "strict-auto.db")
    batch_id = store.create_batch("s02-policy")
    repository = CaseRepository(store)
    state, _created = repository.create_or_get(CaseSeed(
        batch_id=batch_id,
        resource_code="02050015",
        award_name="示例奖",
        year="2025",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="验证严格自动批准边界",
    ))
    registry = ToolRegistry()
    register_fake_tool(registry, "evidence_tool", [ToolResult(
        ok=True,
        evidence_facts=[_fact()],
    )])
    outcome = EvidenceHarness(
        repository=repository,
        registry=registry,
        agent_client=FakeAgentClient([
            NextAction(action="call_tool", tool_name="evidence_tool"),
            NextAction(action="finish", reason_summary="官方完整证据"),
        ]),
        allowed_roots=[tmp_path],
        verifier=EvidenceVerifier(),
        auto_approval_policy=AutoApprovalPolicy(enabled=True),
    ).run(state.case_id)

    assert outcome.stopped_reason == "auto_approved"
    assert outcome.state.status == "completed"
    assert outcome.state.evidence_progress.phase == "auto_approved"
    store.close()


def test_verifier_keeps_best_complete_fact_without_inferior_candidate_gaps() -> None:
    complete = _fact(level="publisher_secondary")
    partial = _fact(
        url="https://example.gov.cn/partial",
        level="official_secondary",
        observed=3,
        complete=False,
    ).model_copy(update={
        "missing_items": ["示例人员"],
        "missing_item_count": 1,
        "missing_evidence": ["候选转载页只包含局部名单"],
    })
    snapshot = build_evidence_snapshot_from_fact_list([complete, partial])

    assert snapshot.explicit_coverage_complete is True
    assert snapshot.observed_count == snapshot.expected_count == 25
    assert snapshot.missing_evidence == []


def test_verifier_model_cannot_turn_partial_overlap_into_zero_overlap() -> None:
    partial = _fact(expected=93, observed=91, complete=False)
    snapshot = build_evidence_snapshot_from_fact(partial)
    model = VerificationReport(
        target_match="no",
        year_match="no",
        source_authority="unknown",
        coverage_complete="no",
        recommended_action="manual",
        reason_codes=["target_mismatch", "year_mismatch", "zero_overlap"],
        deterministic_action="manual",
    )

    report = EvidenceVerifier(FakeVerifierClient([model])).verify(snapshot)

    assert report.target_match == "yes"
    assert report.year_match == "yes"
    assert report.coverage_complete == "no"
    assert "zero_overlap" not in report.reason_codes
    assert "target_mismatch" not in report.reason_codes


def test_attachment_urls_count_as_attempted_and_are_not_refetched_as_pages() -> None:
    pdf_url = "https://example.gov.cn/download?id=roster"
    image_url = "https://example.gov.cn/roster.png"
    state = AuditCaseState.from_seed(
        CaseSeed(
            batch_id=1,
            resource_code="04030052",
            award_name="示例竞赛",
            year="2025",
            trigger_codes=["PDF_ONLY"],
            objective="核验附件名单",
            known_urls=["https://example.gov.cn/page", pdf_url, image_url],
        ),
        ToolBudgetState(),
    )
    state.tool_trace = [ToolObservation(
        call_id="call-1",
        tool_name="extract_pdf_text",
        started_at="2026-01-01T00:00:00Z",
        finished_at="2026-01-01T00:00:01Z",
        duration_ms=1000,
        input_summary={
            "source_url": pdf_url,
            "image_urls": [image_url],
        },
        ok=True,
    )]

    assert {pdf_url, image_url}.issubset(_attempted_urls(state))


def test_asset_followup_continues_when_page_body_lacks_award_but_year_matches() -> None:
    result = ToolResult(
        ok=True,
        evidence_facts=[EvidenceFact(
            status="conflict",
            award_name="",
            year="2025",
            target_match="no",
            year_match="yes",
            source_url="https://example.gov.cn/page",
            source_level="official_secondary",
            coverage_complete=False,
        )],
    )

    assert _asset_followup_allowed(result) is True


def test_default_m5_budget_can_finish_large_multi_pdf_roster() -> None:
    limits = ToolBudgetLimits()

    assert limits.max_calls >= 16
    assert limits.max_pdf_pages >= 250
    assert pdf_tools.MAX_PDF_PAGES >= 250


def test_verifier_retains_real_cross_source_identity_conflict() -> None:
    complete = _fact(level="publisher_secondary")
    conflict = _fact(
        url="https://example.gov.cn/conflict",
        level="official_secondary",
        observed=24,
        complete=False,
    ).model_copy(update={
        "contradictions": ["官方来源使用群体名额，提交材料使用个人姓名"],
    })
    snapshot = build_evidence_snapshot_from_fact_list([complete, conflict])

    assert snapshot.explicit_coverage_complete is True
    assert snapshot.contradictions == ["官方来源使用群体名额，提交材料使用个人姓名"]


def test_verifier_resolves_identity_conflict_with_structured_corroboration() -> None:
    complete = _fact(level="publisher_secondary")
    conflict = _fact(
        url="https://example.gov.cn/conflict",
        level="official_secondary",
        observed=24,
        complete=False,
    ).model_copy(update={
        "contradictions": ["来源使用群体名额，提交材料使用个人姓名，需人工确认对应关系"],
        "missing_items": ["李桂枝", "王伟江"],
        "extra_items": ["保定学院毕业生赴疆任教群体代表"],
    })
    corroboration = _fact(
        url="https://example.gov.cn/group-mapping",
        level="official_secondary",
        observed=1,
        complete=False,
    ).model_copy(update={
        "target_match": "no",
        "year_match": "no",
        "relationship_terms": ["李桂枝", "王伟江", "保定学院毕业生赴疆任教群体代表"],
        "relationship_confirmed": True,
        "relationship_summary": "权威补证来源同时出现差异姓名与群体名称",
    })

    snapshot = build_evidence_snapshot_from_fact_list([complete, conflict, corroboration])
    report = deterministic_verify(snapshot)

    assert snapshot.contradictions == []
    assert "evidence_conflict" not in report.reason_codes


def test_s02_identity_conflict_triggers_targeted_public_corroboration_search(
    tmp_path: Path,
) -> None:
    store = Store(tmp_path / "secondary-cross-check.db")
    batch_id = store.create_batch("s02-cross-check")
    repository = CaseRepository(store)
    secondary_url = "https://news.example.cn/secondary"
    state, _created = repository.create_or_get(CaseSeed(
        batch_id=batch_id,
        resource_code="02050015",
        award_name="示例奖",
        year="2025",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="完整次级来源还需官方交叉核验",
        known_urls=[secondary_url],
    ))
    registry = ToolRegistry()
    official_url = "https://example.gov.cn/official"
    mapping_url = "https://people.example.cn/group-mapping"
    searched = register_fake_tool(registry, "search_official_award", [
        ToolResult(
            ok=True,
            data={
                "strategy": "broad",
                "query": "示例奖 2025 获奖名单 公示",
                "candidates": [{
                    "title": "示例奖正式名单",
                    "url": official_url,
                    "source_level": "official_primary",
                    "provider": "fake",
                    "rank": 1,
                }],
                "candidate_count": 1,
                "official_candidate_count": 1,
            },
            warnings=["search_results_are_leads_not_evidence"],
        ),
        ToolResult(
            ok=True,
            data={
                "strategy": "discrepancy",
                "query": "示例奖 2025 李桂枝 王伟江 保定学院毕业生赴疆任教群体代表 对应关系",
                "candidates": [{
                    "title": "群体代表李桂枝、王伟江",
                    "url": mapping_url,
                    "source_level": "publisher_secondary",
                    "provider": "fake",
                    "rank": 1,
                }],
                "candidate_count": 1,
                "official_candidate_count": 0,
            },
            warnings=["search_results_are_leads_not_evidence"],
        ),
    ])
    conflict = _fact(
        url=official_url,
        level="official_primary",
        observed=24,
        complete=False,
    ).model_copy(update={
        "contradictions": ["来源使用群体名额，提交材料使用个人姓名"],
        "missing_items": ["李桂枝", "王伟江"],
        "extra_items": ["保定学院毕业生赴疆任教群体代表"],
    })
    fetched = register_fake_tool(registry, "fetch_web_page", [
        ToolResult(
            ok=True,
            source_url=secondary_url,
            data={
                "matched_items": ["李桂枝;王伟江"],
                "split_matched_items": ["李桂枝;王伟江"],
                "coverage_complete": True,
            },
            evidence_facts=[_fact(
                url=secondary_url,
                level="publisher_secondary",
            ).model_copy(update={
                "matched_items": ["李桂枝;王伟江"],
                "split_matched_items": ["李桂枝;王伟江"],
            })],
        ),
        ToolResult(
            ok=True,
            source_url=official_url,
            data={
                "comparison_note": "来源使用群体名额，提交材料使用个人姓名",
                "missing_items": ["李桂枝", "王伟江"],
                "extra_items": ["保定学院毕业生赴疆任教群体代表"],
                "coverage_complete": False,
            },
            evidence_facts=[conflict],
        ),
        ToolResult(
            ok=True,
            source_url=mapping_url,
            data={
                "relationship_terms": [
                    "李桂枝",
                    "王伟江",
                    "保定学院毕业生赴疆任教群体代表",
                ],
                "relationship_confirmed": True,
                "relationship_summary": (
                    "补证来源同时出现李桂枝、王伟江和保定学院毕业生赴疆任教群体代表"
                ),
            },
            evidence_facts=[_fact(
                url=mapping_url,
                level="official_secondary",
                observed=1,
                complete=False,
            ).model_copy(update={
                "target_match": "no",
                "year_match": "no",
                "relationship_terms": [
                    "李桂枝",
                    "王伟江",
                    "保定学院毕业生赴疆任教群体代表",
                ],
                "relationship_confirmed": True,
                "relationship_summary": "权威补证来源同时出现差异姓名与群体名称",
            })],
        ),
    ])
    client = FakeAgentClient([
        NextAction(action="manual", reason_summary="官网存在群体名额差异"),
    ])

    outcome = EvidenceHarness(
        repository=repository,
        registry=registry,
        agent_client=client,
        allowed_roots=[tmp_path],
        verifier=EvidenceVerifier(),
    ).run(state.case_id)

    assert len(searched.calls) == 1
    assert [call["url"] for call in fetched.calls] == [
        secondary_url,
        official_url,
    ]
    assert outcome.state.status == "waiting_human"
    assert outcome.state.latest_verification is not None
    assert outcome.state.latest_verification.contradictions
    assert "secondary_evidence_requires_official_corroboration" in (
        outcome.state.reason_codes
    )
    assert outcome.state.evidence_progress.search_round == 1
    assert "identity_discrepancy_recovery_started" not in outcome.state.reason_codes
    store.close()


def test_repeated_failed_known_url_redirects_to_search_recovery(tmp_path: Path) -> None:
    store = Store(tmp_path / "failed-known-url.db")
    batch_id = store.create_batch("failed-known-url")
    repository = CaseRepository(store)
    known_url = "https://news.example.cn/unavailable"
    recovered_url = "https://official.example.gov.cn/recovered"
    state, _created = repository.create_or_get(CaseSeed(
        batch_id=batch_id,
        resource_code="02050015",
        award_name="最美教师",
        year="2025",
        trigger_codes=["SOURCE_UNREACHABLE"],
        objective="已给网址失败后搜索恢复",
        known_urls=[known_url],
    ))
    registry = ToolRegistry()
    fetched = register_fake_tool(registry, "fetch_web_page", [
        ToolResult.failure("HTTP_ERROR", "HTTP 503"),
        ToolResult(ok=True, evidence_facts=[_fact(url=recovered_url)]),
    ])
    searched = register_fake_tool(registry, "search_official_award", [ToolResult(
        ok=True,
        data={
            "strategy": "broad",
                "candidates": [{
                    "url": recovered_url,
                    "source_level": "official_primary",
                    "rank": 1,
                    "title": "最美教师 2025 获奖名单",
                }],
            "candidate_count": 1,
            "official_candidate_count": 1,
        },
        warnings=["search_results_are_leads_not_evidence"],
    )])
    client = FakeAgentClient([
        NextAction(action="call_tool", tool_name="fetch_web_page", arguments={"url": known_url}),
        NextAction(action="call_tool", tool_name="fetch_web_page", arguments={"url": known_url}),
        NextAction(
            action="call_tool",
            tool_name="fetch_web_page",
            arguments={"url": recovered_url},
        ),
        NextAction(action="finish", reason_summary="恢复来源已核验"),
    ])

    outcome = EvidenceHarness(
        repository=repository,
        registry=registry,
        agent_client=client,
        allowed_roots=[tmp_path],
        verifier=EvidenceVerifier(),
    ).run(state.case_id)

    assert len(searched.calls) == 1
    assert len(fetched.calls) == 2
    assert outcome.state.status == "waiting_human"
    assert "repeated_failed_url_redirected_to_search" in outcome.state.reason_codes
    store.close()


def test_extract_search_document_preserves_relationship_corroboration(
    tmp_path: Path,
) -> None:
    url = "https://example.gov.cn/group-mapping"
    provider = FakeSearchProvider([ExtractResponse(
        provider="fake",
        url=url,
        text=(
            "2025年最美教师。保定学院毕业生赴疆任教群体代表李桂枝、王伟江，"
            "继续扎根新疆任教。"
        ),
    )])
    registry = build_default_registry(search_provider_factory=lambda: provider)
    context = ToolExecutionContext.create([tmp_path])

    result = SafeToolExecutor(registry).execute("extract_search_document", {
        "url": url,
        "expected_award_name": "最美教师",
        "expected_year": "2025",
        "relationship_terms": [
            "李桂枝",
            "王伟江",
            "保定学院毕业生赴疆任教群体代表",
        ],
    }, context)

    assert result.ok
    assert result.data["relationship_confirmed"] is True
    assert result.evidence_facts[0].relationship_confirmed is True
    assert result.evidence_facts[0].relationship_terms == [
        "李桂枝",
        "王伟江",
        "保定学院毕业生赴疆任教群体代表",
    ]


def test_extract_search_document_bounds_large_missing_summary(tmp_path: Path) -> None:
    submitted = tmp_path / "submitted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["ISBN", "JCMC"])
    sheet.append(["标准书号", "教材名称"])
    for index in range(30):
        sheet.append([f"978700000{index:04d}", f"未命中教材名称{index:02d}" * 8])
    workbook.save(submitted)
    workbook.close()
    url = "https://example.gov.cn/large-roster"
    provider = FakeSearchProvider([ExtractResponse(
        provider="fake",
        url=url,
        text=("2025年全国教材建设奖公示，名单正文未包含提交记录。" * 80),
    )])
    result = SafeToolExecutor(build_default_registry(
        search_provider_factory=lambda: provider
    )).execute(
        "extract_search_document",
        {
            "url": url,
            "expected_award_name": "全国教材建设奖",
            "expected_year": "2025",
            "submitted_path": submitted,
            "submitted_paths": [submitted],
            "match_fields": ["ISBN", "JCMC"],
            "match_combine": "all",
            "expected_scope_count": 30,
        },
        ToolExecutionContext.create([tmp_path]),
    )

    assert result.ok and result.evidence_facts
    fact = result.evidence_facts[0]
    assert fact.missing_item_count == 30
    assert len(fact.missing_evidence) == 1
    assert len(fact.missing_evidence[0]) <= 500


def test_short_search_excerpt_is_missing_evidence_not_identity_conflict(
    tmp_path: Path,
) -> None:
    submitted = tmp_path / "submitted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["ISBN", "JCMC"])
    sheet.append(["标准书号", "教材名称"])
    sheet.append(["9787000000001", "示例教材"])
    workbook.save(submitted)
    workbook.close()
    url = "https://example.gov.cn/unavailable.pdf"
    provider = FakeSearchProvider([ExtractResponse(
        provider="fake",
        url=url,
        text="Document extraction is currently unavailable.",
    )])
    registry = build_default_registry(search_provider_factory=lambda: provider)
    entry = registry.get("extract_search_document")
    assert entry is not None and entry[0].kind == "general"

    result = SafeToolExecutor(registry).execute(
        "extract_search_document",
        {
            "url": url,
            "expected_award_name": "示例教材奖",
            "expected_year": "2025",
            "submitted_path": submitted,
            "submitted_paths": [submitted],
            "match_fields": ["ISBN", "JCMC"],
            "match_combine": "all",
            "expected_scope_count": 1,
        },
        ToolExecutionContext.create([
            tmp_path,
        ], ToolBudgetLimits(max_searches=0)),
    )

    assert result.ok and result.evidence_facts
    fact = result.evidence_facts[0]
    assert fact.status == "unverified"
    assert fact.target_match == fact.year_match == "uncertain"
    assert fact.coverage_complete is None
    assert fact.observed_count is None
    assert fact.missing_item_count == 0
    assert fact.missing_evidence == ["搜索服务仅返回摘要，缺少可核验名单正文"]


def test_s01_discovered_spreadsheet_is_processed_before_finish(
    tmp_path: Path,
) -> None:
    store = Store(tmp_path / "attachment-transition.db")
    batch_id = store.create_batch("s01-attachment-transition")
    repository = CaseRepository(store)
    submitted = tmp_path / "submitted.xlsx"
    submitted.touch()
    page_url = "https://www.chinazy.org/info/1014/15997.htm"
    state, _created = repository.create_or_get(CaseSeed(
        batch_id=batch_id,
        resource_code="06090003",
        award_name="全国高校辅导员年度人物",
        year="2023",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="网页发现 Excel 后必须先核验附件",
        submitted_summary={
            "submission_file": str(submitted),
            "match_fields": ["XRYXM"],
            "submitted_rows": 20,
        },
        known_urls=[page_url],
    ))
    registry = ToolRegistry()
    fetched = register_fake_tool(registry, "fetch_web_page", [ToolResult(
        ok=True,
        source_url=page_url,
        data={
            "title": "2023年最美高校辅导员候选人公示",
            "next_evidence_stage": "spreadsheet_processing",
            "candidate_attachment_urls": [
                "https://www.chinazy.org/system/_content/download.jsp?id=1"
            ],
        },
        evidence_facts=[_fact(
            url=page_url,
            level="publisher_secondary",
            expected=20,
            observed=0,
            complete=False,
        )],
    )])
    collected = register_fake_tool(
        registry,
        "collect_spreadsheet_attachments",
        [ToolResult(
            ok=True,
            source_url=page_url,
            evidence_facts=[_fact(
                url=page_url,
                level="publisher_secondary",
                expected=20,
                observed=20,
                complete=True,
            )],
        )],
    )
    client = FakeAgentClient([
        NextAction(
            action="call_tool",
            tool_name="fetch_web_page",
            arguments={"url": page_url},
        ),
        NextAction(action="finish", reason_summary="网页正文没有名单"),
        NextAction(action="finish", reason_summary="附件名单核验完成"),
    ])

    outcome = EvidenceHarness(
        repository=repository,
        registry=registry,
        agent_client=client,
        allowed_roots=[tmp_path],
        verifier=EvidenceVerifier(),
    ).run(state.case_id)

    assert fetched.calls
    assert len(collected.calls) == 1
    assert collected.calls[0]["page_urls"] == [page_url]
    assert collected.calls[0]["submitted_path"] == str(submitted)
    assert outcome.state.evidence_progress.pending_attachment_page_urls == []
    assert outcome.state.latest_verification is not None
    assert outcome.state.latest_verification.coverage_complete == "yes"
    store.close()


def test_s01_attachment_download_retries_one_transient_timeout(
    tmp_path: Path,
    monkeypatch,
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted.xlsx"
    remote = tmp_path / "remote.xlsx"
    for path in (submitted, remote):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["XRYXM"])
        sheet.append(["入选人姓名"])
        sheet.append(["王奇"])
        workbook.save(path)
        workbook.close()
    page_url = "https://www.chinazy.org/info/1014/15997.htm"
    attachment_url = "https://www.chinazy.org/system/download?id=3821158"
    monkeypatch.setattr(
        registry_module.web,
        "fetch_page",
        lambda *_args, **_kwargs: PageContent(
            url=page_url,
            status=200,
            text="2023年高校辅导员年度人物候选人",
            title="2023年高校辅导员年度人物候选人公示",
            attachments=[Attachment(
                text="第十五届高校辅导员年度人物候选人.xlsx",
                url=attachment_url,
                is_excel=True,
            )],
        ),
    )
    calls: list[tuple[float, str]] = []

    class ReadTimeout(Exception):
        pass

    def flaky_download(_url, destination, *, timeout, referer, **_kwargs):  # noqa: ANN001
        calls.append((timeout, referer))
        if len(calls) == 1:
            raise ReadTimeout("transient read timeout")
        target = Path(destination) / "downloaded.xlsx"
        shutil.copyfile(remote, target)
        return target

    monkeypatch.setattr(registry_module.web, "download_file", flaky_download)
    executor = SafeToolExecutor(build_default_registry())
    result = executor.execute("collect_spreadsheet_attachments", {
        "page_urls": [page_url],
        "submitted_path": str(submitted),
        "match_fields": ["XRYXM"],
        "include_attachment_keywords": ["高校辅导员年度人物"],
        "exclude_attachment_keywords": ["提名", "最美"],
        "expected_award_name": "全国高校辅导员年度人物",
        "expected_year": "2023",
        "expected_scope_count": 1,
    }, ToolExecutionContext.create([tmp_path]))

    assert result.ok
    assert calls == [(30.0, page_url), (60.0, page_url)]
    assert result.evidence_facts[0].coverage_complete is True


def test_s03_joint_attachment_fact_keeps_award_year_source_and_hashes(
    tmp_path: Path,
    monkeypatch,
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted.xlsx"
    remote = tmp_path / "remote.xlsx"
    for path in (submitted, remote):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["CSDWMC", "XCSDW"])
        sheet.append(["参赛队伍名称", "参赛单位"])
        sheet.append(["测试志愿队", "测试大学"])
        workbook.save(path)
        workbook.close()
    page_url = "https://cpipc.acge.org.cn/cw/detail/example"
    attachment_url = "https://cpipc.acge.org.cn/sysFile/downFile.do?id=1"
    monkeypatch.setattr(
        registry_module.web,
        "fetch_page",
        lambda *_args, **_kwargs: PageContent(
            url=page_url,
            status=200,
            title="第一届乡村振兴志愿服务技能大赛获奖结果",
            text="2023年比赛结果公布，名单见附件。",
            attachments=[Attachment(
                text="获奖团队.xlsx",
                url=attachment_url,
                is_excel=True,
            )],
        ),
    )

    def fake_download(_url, destination, **_kwargs):  # noqa: ANN001
        target = Path(destination) / "downloaded.xlsx"
        shutil.copyfile(remote, target)
        return target

    monkeypatch.setattr(registry_module.web, "download_file", fake_download)
    result = SafeToolExecutor(build_default_registry()).execute(
        "collect_spreadsheet_attachments",
        {
            "page_urls": [page_url],
            "submitted_path": str(submitted),
            "match_fields": ["CSDWMC", "XCSDW"],
            "include_attachment_keywords": ["获奖团队"],
            "expected_award_name": (
                "中国研究生创新实践系列大赛：全国农科研究生志愿服务技能大赛"
            ),
            "award_aliases": ["乡村振兴志愿服务技能大赛"],
            "expected_year": "2023",
            "expected_scope_count": 1,
            "official_secondary_domains": ["cpipc.acge.org.cn"],
        },
        ToolExecutionContext.create([tmp_path]),
    )

    assert result.ok and len(result.artifacts) == 1
    fact = result.evidence_facts[0]
    assert fact.status == "complete"
    assert fact.target_match == fact.year_match == "yes"
    assert fact.source_level == "official_secondary"
    assert fact.coverage_complete is True
    assert fact.document_count == len(fact.artifact_hashes) == 1


def test_truncated_spreadsheet_attachment_cannot_be_complete(
    tmp_path: Path,
    monkeypatch,
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted.xlsx"
    remote = tmp_path / "remote.xlsx"
    submitted_book = openpyxl.Workbook()
    submitted_sheet = submitted_book.active
    submitted_sheet.append(["XRYXM"])
    submitted_sheet.append(["姓名"])
    submitted_sheet.append(["张三"])
    submitted_book.save(submitted)
    submitted_book.close()
    remote_book = openpyxl.Workbook()
    remote_sheet = remote_book.active
    remote_sheet.append(["姓名"])
    remote_sheet.append(["张三"])
    remote_sheet.append(["李四"])
    remote_sheet.append(["王五"])
    remote_book.save(remote)
    remote_book.close()
    page_url = "https://example.gov.cn/notice"
    attachment_url = "https://example.gov.cn/list.xlsx"
    monkeypatch.setattr(
        registry_module.web,
        "fetch_page",
        lambda *_args, **_kwargs: PageContent(
            url=page_url,
            status=200,
            title="2025年示例奖获奖名单",
            text="2025年示例奖名单见附件",
            attachments=[Attachment(
                text="2025年示例奖获奖名单.xlsx",
                url=attachment_url,
                is_excel=True,
            )],
        ),
    )

    def fake_download(_url, destination, **_kwargs):  # noqa: ANN001
        target = Path(destination) / "downloaded.xlsx"
        shutil.copyfile(remote, target)
        return target

    monkeypatch.setattr(registry_module.web, "download_file", fake_download)
    result = SafeToolExecutor(build_default_registry()).execute(
        "collect_spreadsheet_attachments",
        {
            "page_urls": [page_url],
            "submitted_path": str(submitted),
            "match_fields": ["XRYXM"],
            "expected_award_name": "示例奖",
            "expected_year": "2025",
            "expected_scope_count": 1,
            "max_rows_per_file": 3,
        },
        ToolExecutionContext.create([tmp_path]),
    )

    assert result.ok
    assert result.data["spreadsheet_truncated"] is True
    assert result.data["coverage_complete"] is False
    assert result.evidence_facts[0].coverage_complete is False
    assert any(
        "表格读取达到行数上限" in item
        for item in result.evidence_facts[0].missing_evidence
    )


def test_generic_joint_attachments_do_not_hide_unresolved_group_file(
    tmp_path: Path,
    monkeypatch,
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted.xlsx"
    attachment_paths: dict[str, Path] = {}

    def write_names(path: Path, names: list[str]) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["XM"])
        sheet.append(["姓名"])
        for name in names:
            sheet.append([name])
        workbook.save(path)
        workbook.close()

    write_names(submitted, ["甲", "乙", "丙", "丁"])
    for index, names in enumerate((["甲", "乙"], ["丙", "丁"], ["无关人员"]), 1):
        source_url = f"https://example.org/files/{index}.xlsx"
        source_path = tmp_path / f"source-{index}.xlsx"
        write_names(source_path, names)
        attachment_paths[source_url] = source_path

    page_url = "https://example.org/awards/2025/results"
    monkeypatch.setattr(
        registry_module.web,
        "fetch_page",
        lambda *_args, **_kwargs: PageContent(
            url=page_url,
            status=200,
            title="2025年青年公益创新竞赛获奖结果公示",
            text="获奖名单分组列于三个附件。",
            attachments=[
                Attachment(text="A组.xlsx", url=url, is_excel=True)
                for url in attachment_paths
            ],
        ),
    )

    def fake_download(url, destination, **_kwargs):  # noqa: ANN001
        target = Path(destination) / f"downloaded-{Path(url).name}"
        shutil.copyfile(attachment_paths[url], target)
        return target

    monkeypatch.setattr(registry_module.web, "download_file", fake_download)
    result = SafeToolExecutor(build_default_registry()).execute(
        "collect_spreadsheet_attachments",
        {
            "page_urls": [page_url],
            "submitted_path": str(submitted),
            "match_fields": ["XM"],
            "expected_award_name": "全国青年公益创新竞赛",
            "expected_year": "2025",
            "expected_scope_count": 4,
        },
        ToolExecutionContext.create([tmp_path]),
    )

    assert result.ok
    fact = result.evidence_facts[0]
    assert fact.target_match == fact.year_match == "yes"
    assert fact.coverage_complete is False
    assert any(
        "同组仍有未纳入名单范围的表格附件" in item
        for item in fact.missing_evidence
    )
    assert fact.observed_count == fact.expected_count == 4
    assert fact.document_count == len(fact.artifact_hashes) == 2
    assert len(result.artifacts) == 3
    assert result.data["downloaded_attachment_count"] == 3
    assert result.data["attachment_count"] == 2
    assert result.data["unresolved_attachment_urls"] == [
        "https://example.org/files/3.xlsx"
    ]
    assert result.warnings == ["attachments_selected_by_roster_overlap"]


def test_joint_attachment_group_processes_more_than_eight_files_without_silent_drop(
    tmp_path: Path,
    monkeypatch,
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted-nine.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XM"])
    sheet.append(["姓名"])
    for index in range(1, 10):
        sheet.append([f"人员{index}"])
    workbook.save(submitted)
    workbook.close()

    attachment_paths: dict[str, Path] = {}
    attachments: list[Attachment] = []
    for index in range(1, 10):
        url = f"https://example.gov.cn/files/group-{index}.xlsx"
        path = tmp_path / f"group-{index}.xlsx"
        book = openpyxl.Workbook()
        part = book.active
        part.append(["姓名"])
        part.append([f"人员{index}"])
        book.save(path)
        book.close()
        attachment_paths[url] = path
        attachments.append(Attachment(text=f"第{index}组名单.xlsx", url=url, is_excel=True))

    page_url = "https://example.gov.cn/2025-award"
    monkeypatch.setattr(
        registry_module.web,
        "fetch_page",
        lambda *_args, **_kwargs: PageContent(
            url=page_url,
            status=200,
            title="2025年示例奖获奖名单",
            text="2025年示例奖获奖名单分为九组附件。",
            attachments=attachments,
        ),
    )

    def fake_download(url, destination, **_kwargs):  # noqa: ANN001
        target = Path(destination) / f"downloaded-{Path(url).name}"
        shutil.copyfile(attachment_paths[url], target)
        return target

    monkeypatch.setattr(registry_module.web, "download_file", fake_download)
    result = SafeToolExecutor(build_default_registry()).execute(
        "collect_spreadsheet_attachments",
        {
            "page_urls": [page_url],
            "submitted_path": str(submitted),
            "match_fields": ["XM"],
            "expected_award_name": "示例奖",
            "expected_year": "2025",
            "expected_scope_count": 9,
        },
        ToolExecutionContext.create([tmp_path]),
    )

    assert result.ok
    assert result.data["discovered_attachment_count"] == 9
    assert result.data["attempted_attachment_count"] == 9
    assert result.data["downloaded_attachment_count"] == 9
    assert result.data["unprocessed_attachment_urls"] == []
    assert result.data["failed_attachment_urls"] == []
    assert result.data["all_attachments_processed"] is True
    assert result.data["coverage_complete"] is True
    assert len(result.data["attachment_manifest"]) == 9
    assert result.data["attachment_manifest"][0] == {
        "manifest_version": 1,
        "url": "https://example.gov.cn/files/group-1.xlsx",
        "parent_url": page_url,
        "label": "第1组名单.xlsx",
        "kind": "xlsx",
        "status": "parsed",
        "selected": True,
        "matched_identity_count": 1,
        "truncated": False,
        "error_code": "",
        "sheets": [{"sheet": "Sheet", "row_count": 2, "truncated": False}],
    }

    limited = SafeToolExecutor(build_default_registry()).execute(
        "collect_spreadsheet_attachments",
        {
            "page_urls": [page_url],
            "submitted_path": str(submitted),
            "match_fields": ["XM"],
            "expected_award_name": "示例奖",
            "expected_year": "2025",
            "expected_scope_count": 9,
            "max_attachments": 8,
        },
        ToolExecutionContext.create([tmp_path]),
    )
    assert limited.ok
    assert limited.data["unprocessed_attachment_urls"] == [
        "https://example.gov.cn/files/group-9.xlsx"
    ]
    assert limited.data["all_attachments_processed"] is False
    assert limited.data["coverage_complete"] is False


def test_submitted_multi_sheet_identity_is_built_per_sheet(
    tmp_path: Path,
    monkeypatch,
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted-multi-sheet.xlsx"
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "第一组"
    first.append(["XM", "DW"])
    first.append(["姓名", "单位"])
    first.append(["张甲", "甲校"])
    second = workbook.create_sheet("第二组")
    second.append(["DW", "XM"])
    second.append(["单位", "姓名"])
    second.append(["乙校", "李乙"])
    workbook.save(submitted)
    workbook.close()

    remote = tmp_path / "remote-multi-sheet.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["姓名"])
    sheet.append(["张甲"])
    sheet.append(["李乙"])
    workbook.save(remote)
    workbook.close()
    page_url = "https://example.gov.cn/2025-award"
    attachment_url = "https://example.gov.cn/files/list.xlsx"
    monkeypatch.setattr(
        registry_module.web,
        "fetch_page",
        lambda *_args, **_kwargs: PageContent(
            url=page_url,
            status=200,
            title="2025年示例奖获奖名单",
            text="2025年示例奖名单见附件",
            attachments=[Attachment(text="获奖名单.xlsx", url=attachment_url, is_excel=True)],
        ),
    )

    def fake_download(_url, destination, **_kwargs):  # noqa: ANN001
        target = Path(destination) / "downloaded-list.xlsx"
        shutil.copyfile(remote, target)
        return target

    monkeypatch.setattr(registry_module.web, "download_file", fake_download)
    result = SafeToolExecutor(build_default_registry()).execute(
        "collect_spreadsheet_attachments",
        {
            "page_urls": [page_url],
            "submitted_path": str(submitted),
            "match_fields": ["XM"],
            "expected_award_name": "示例奖",
            "expected_year": "2025",
            "expected_scope_count": 2,
        },
        ToolExecutionContext.create([tmp_path]),
    )

    assert result.ok
    assert result.data["submitted_count"] == 2
    assert result.data["submitted_match_count"] == 2
    assert result.data["coverage_complete"] is True


def build_evidence_snapshot_from_fact_list(facts: list[EvidenceFact]):  # noqa: ANN201
    from award_audit.agent.harness.models import AuditCaseState
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    first = facts[0]
    state = AuditCaseState.from_seed(
        CaseSeed(
            batch_id=1,
            resource_code="02050015",
            award_name=first.award_name,
            year=first.year,
            trigger_codes=["COVERAGE_UNKNOWN"],
            objective="核验多来源证据",
        ),
        ToolBudgetState(),
    )
    return build_evidence_snapshot(
        state,
        [ToolResult(ok=True, evidence_facts=[fact]) for fact in facts],
    )


def build_evidence_snapshot_from_fact(fact: EvidenceFact):  # noqa: ANN201
    from award_audit.agent.harness.models import AuditCaseState
    from award_audit.agent.toolkit.contracts import ToolBudgetState

    state = AuditCaseState.from_seed(
        CaseSeed(
            batch_id=1,
            resource_code="02050015",
            award_name=fact.award_name,
            year=fact.year,
            trigger_codes=["COVERAGE_UNKNOWN"],
            objective="核验证据",
        ),
        ToolBudgetState(),
    )
    return build_evidence_snapshot(state, [ToolResult(ok=True, evidence_facts=[fact])])


def test_s03_joint_fact_keeps_award_year_source_and_307_coverage_together() -> None:
    fact = _fact(expected=307, observed=307)
    snapshot = build_evidence_snapshot_from_fact(fact)
    report = deterministic_verify(snapshot)

    assert snapshot.observed_award_names == ["示例奖"]
    assert snapshot.observed_years == ["2025"]
    assert snapshot.source_levels == ["official_primary"]
    assert snapshot.expected_count == snapshot.observed_count == 307
    assert report.recommended_action == "accept_evidence"


def test_s04_spreadsheet_guess_preserves_magic_detected_pdf_for_pdf_stage(
    tmp_path: Path,
    monkeypatch,
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XMMC"])
    sheet.append(["项目名称"])
    sheet.append(["项目一"])
    workbook.save(submitted)
    workbook.close()
    page_url = "https://example.gov.cn/2025-award"
    attachment_url = "https://example.gov.cn/download?id=42"
    monkeypatch.setattr(registry_module.web, "fetch_page", lambda *_args, **_kwargs: PageContent(
        url=page_url,
        status=200,
        text="2025年示例奖附件",
        title="2025年示例奖名单",
        attachments=[Attachment(text="附件一", url=attachment_url, is_excel=False)],
    ))

    def fake_download(_url, destination, **_kwargs):  # noqa: ANN001
        target = Path(destination) / "actual.pdf"
        shutil.copyfile(
            Path("tests/data/m5_golden/pdf/digital_roster.pdf"), target
        )
        return target

    monkeypatch.setattr(registry_module.web, "download_file", fake_download)
    executor = SafeToolExecutor(build_default_registry())
    result = executor.execute("collect_spreadsheet_attachments", {
        "page_urls": [page_url],
        "submitted_path": str(submitted),
        "match_fields": ["XMMC"],
        "expected_award_name": "示例奖",
        "expected_year": "2025",
    }, ToolExecutionContext.create([tmp_path]))

    assert result.ok and [item.kind for item in result.artifacts] == ["pdf"]
    assert result.data["next_evidence_stage"] == "pdf_processing"
    assert "pdf_processing_required" in result.warnings
    metadata = result.artifacts[0].metadata
    assert metadata["attachment_linked"] is True
    assert metadata["page_target_match"] is True
    assert metadata["page_year_match"] is True
    assert metadata["page_observed_award_name"] == "示例奖"
    assert metadata["page_observed_year"] == "2025"


def test_s04_discovered_detail_page_reaches_pdf_before_new_search(
    tmp_path: Path,
) -> None:
    store = Store(tmp_path / "s04-related-page.db")
    batch_id = store.create_batch("s04-related-page")
    repository = CaseRepository(store)
    listing = "https://cpipc.acge.org.cn/cw/contestNews/list/contest/1"
    detail = "https://cpipc.acge.org.cn/cw/contestNews/detail/award-1"
    pdf_url = "https://cpipc.acge.org.cn/sysFile/downFile.do?fileId=pdf-1"
    submitted = tmp_path / "submitted.xlsx"
    submitted.touch()
    state, _created = repository.create_or_get(CaseSeed(
        batch_id=batch_id,
        resource_code="04030060",
        award_name="中国研究生创新实践系列大赛：全国农科研究生志愿服务技能大赛",
        year="2025",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="从通知列表进入详情页并识别 PDF",
        known_urls=[listing],
        submitted_summary={
            "submission_file": str(submitted),
            "match_fields": ["CSDWMC"],
            "reference_rows": 325,
        },
    ))
    registry = ToolRegistry()
    fetched = register_fake_tool(registry, "fetch_web_page", [
        ToolResult(ok=True, source_url=listing, data={
            "source_level": "official_secondary",
            "next_evidence_stage": "page_recovery",
            "candidate_page_urls": [detail],
            "candidate_page_titles": ["2025年获奖名单通知"],
        }),
        ToolResult(ok=True, source_url=detail, data={
            "source_level": "official_secondary",
            "next_evidence_stage": "spreadsheet_processing",
            "candidate_attachment_urls": [pdf_url],
        }),
    ])
    searched = register_fake_tool(
        registry,
        "search_official_award",
        [ToolResult(ok=True)],
    )
    pdf_path = tmp_path / "actual.pdf"
    pdf_path.write_bytes(b"%PDF-1.5\nfixture")
    collected = register_fake_tool(registry, "collect_spreadsheet_attachments", [
        ToolResult(
            ok=True,
            source_url=detail,
            data={
                "next_evidence_stage": "pdf_processing",
                "detected_attachment_kinds": ["pdf"],
                "coverage_complete": False,
            },
            artifacts=[EvidenceArtifact(
                kind="pdf",
                source_url=pdf_url,
                local_path=str(pdf_path),
                content_type="application/pdf",
                sha256="a" * 64,
                size_bytes=pdf_path.stat().st_size,
                fetched_at="2026-07-29T00:00:00Z",
            )],
        )
    ])
    client = FakeAgentClient([
        NextAction(action="call_tool", tool_name="fetch_web_page", arguments={"url": listing}),
        NextAction(action="manual", reason_summary="PDF 内容仍需核验"),
    ])

    outcome = EvidenceHarness(
        repository=repository,
        registry=registry,
        agent_client=client,
        allowed_roots=[tmp_path],
        verifier=EvidenceVerifier(),
    ).run(state.case_id)

    assert [call["url"] for call in fetched.calls] == [listing, detail]
    assert searched.calls == []
    assert len(collected.calls) == 1
    assert collected.calls[0]["page_urls"] == [detail]
    assert [artifact.kind for artifact in outcome.state.artifacts] == ["pdf"]
    assert len(client.calls) == 1
    assert outcome.stopped_reason == "repeated_tool_call_blocked"
    assert store.list_audit_attempts(state.case_id)[-1]["verifier_status"] == "persisted"
    store.close()


def test_s05_candidate_queue_survives_two_failures_and_uses_third_candidate(
    tmp_path: Path,
) -> None:
    store = Store(tmp_path / "recovery.db")
    batch_id = store.create_batch("s05-offline")
    repository = CaseRepository(store)
    state, _created = repository.create_or_get(
        CaseSeed(
            batch_id=batch_id,
            resource_code="03020004",
            award_name="示例奖",
            year="2025",
            trigger_codes=["SOURCE_UNREACHABLE"],
            objective="恢复失效 PDF 来源",
        ),
        tool_limits=ToolBudgetLimits(max_calls=10),
    )
    urls = [f"https://example.gov.cn/candidate-{index}" for index in range(1, 4)]
    registry = ToolRegistry()
    register_fake_tool(registry, "search_official_award", [ToolResult(ok=True, data={
            "candidates": [
                {
                    "url": url, "source_level": "official_primary", "rank": index,
                    "title": f"示例奖 2025 获奖名单 {index}",
                }
                for index, url in enumerate(urls, start=1)
        ],
        "candidate_count": 3,
        "official_candidate_count": 3,
    }, warnings=["search_results_are_leads_not_evidence"])])
    fetched = register_fake_tool(registry, "fetch_web_page", [
        ToolResult.failure("HTTP_ERROR", "HTTP 500"),
        ToolResult.failure("HTTP_ERROR", "HTTP 502"),
        ToolResult(ok=True, evidence_facts=[_fact(url=urls[2])]),
    ])
    client = FakeAgentClient([
        NextAction(action="call_tool", tool_name="search_official_award",
                   arguments={"award_name": "示例奖"}),
        *[
            NextAction(action="call_tool", tool_name="fetch_web_page", arguments={"url": url})
            for url in urls
        ],
        NextAction(action="finish", reason_summary="第三候选证据完整"),
    ])
    outcome = EvidenceHarness(
        repository=repository,
        registry=registry,
        agent_client=client,
        allowed_roots=[tmp_path],
        verifier=EvidenceVerifier(),
    ).run(state.case_id)

    assert outcome.stopped_reason == "recommendation_ready"
    assert len(fetched.calls) == 3
    assert outcome.state.latest_verification is not None
    assert outcome.state.latest_verification.recommended_action == "accept_evidence"
    reloaded = repository.load(state.case_id)
    assert reloaded.evidence_progress.phase == "waiting_human"
    assert [item.status for item in reloaded.evidence_progress.candidates] == [
        "failed", "failed", "succeeded"
    ]
    store.close()


def test_candidate_queue_redirects_repeated_action_to_next_pending_url(
    tmp_path: Path,
) -> None:
    store = Store(tmp_path / "candidate-redirect.db")
    batch_id = store.create_batch("candidate-redirect")
    repository = CaseRepository(store)
    known_url = "https://example.edu.cn/known"
    first_candidate = "https://example.gov.cn/candidate-1"
    second_candidate = "https://example.gov.cn/candidate-2"
    state, _created = repository.create_or_get(CaseSeed(
        batch_id=batch_id,
        resource_code="06090003",
        award_name="全国高校辅导员年度人物",
        year="2023",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="候选队列必须由 Harness 持续处理",
        known_urls=[known_url],
    ))
    registry = ToolRegistry()
    register_fake_tool(registry, "parse_spreadsheet", [ToolResult(ok=True)])
    fetched = register_fake_tool(registry, "fetch_web_page", [
        ToolResult(ok=True, evidence_facts=[_fact(
            url=known_url, expected=50, observed=0, complete=False,
        )]),
        ToolResult(ok=True, evidence_facts=[_fact(
            url=first_candidate, expected=50, observed=20, complete=False,
        )]),
        ToolResult(ok=True, evidence_facts=[_fact(
            url=second_candidate, expected=50, observed=0, complete=False,
        )]),
    ])
    register_fake_tool(registry, "search_official_award", [ToolResult(ok=True, data={
            "candidates": [
                {"url": first_candidate, "source_level": "official_primary", "rank": 1, "title": "2023年获奖名单"},
                {"url": second_candidate, "source_level": "official_primary", "rank": 2, "title": "2023年结果名单"},
        ],
        "candidate_count": 2,
        "official_candidate_count": 2,
    })])
    client = FakeAgentClient([
        NextAction(action="call_tool", tool_name="parse_spreadsheet"),
        NextAction(action="call_tool", tool_name="search_official_award"),
        NextAction(action="call_tool", tool_name="fetch_web_page",
                   arguments={"url": second_candidate}),
        NextAction(action="call_tool", tool_name="parse_spreadsheet"),
        NextAction(action="finish", reason_summary="候选队列已耗尽"),
    ])
    outcome = EvidenceHarness(
        repository=repository,
        registry=registry,
        agent_client=client,
        allowed_roots=[tmp_path],
        verifier=EvidenceVerifier(),
    ).run(state.case_id)

    assert len(fetched.calls) == 3
    assert [call["url"] for call in fetched.calls] == [
        known_url,
        first_candidate,
        second_candidate,
    ]
    assert "pending_candidate_processed_without_agent_turn" in (
        outcome.state.reason_codes
    )
    assert outcome.stopped_reason != "repeated_tool_call_blocked"
    store.close()


def test_cross_year_page_does_not_queue_attachments_and_advances_known_url(
    tmp_path: Path,
) -> None:
    store = Store(tmp_path / "cross-year-known-source.db")
    batch_id = store.create_batch("cross-year-known-source")
    repository = CaseRepository(store)
    wrong_year = "https://competition.example.cn/2025"
    correct_year = "https://competition.example.cn/2023"
    attachment = "https://competition.example.cn/2025-roster.pdf"
    state, _created = repository.create_or_get(CaseSeed(
        batch_id=batch_id,
        resource_code="NEW-RESOURCE",
        award_name="示例竞赛",
        year="2023",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="按案件年份选择已知来源",
        known_urls=[wrong_year, correct_year],
    ))
    registry = ToolRegistry()
    fetched = register_fake_tool(registry, "fetch_web_page", [
        ToolResult(
            ok=True,
            source_url=wrong_year,
            data={
                "next_evidence_stage": "spreadsheet_processing",
                "candidate_attachment_urls": [attachment],
            },
            evidence_facts=[_fact(
                url=wrong_year,
                expected=307,
                observed=0,
                complete=False,
            ).model_copy(update={
                "award_name": "示例竞赛",
                "year": "2025",
                "year_match": "no",
            })],
        ),
        ToolResult(
            ok=True,
            source_url=correct_year,
            evidence_facts=[_fact(
                url=correct_year,
                expected=307,
                observed=307,
                complete=True,
            ).model_copy(update={
                "award_name": "示例竞赛",
                "year": "2023",
            })],
        ),
    ])
    collected = register_fake_tool(
        registry,
        "collect_spreadsheet_attachments",
        [ToolResult(ok=True)],
    )

    outcome = EvidenceHarness(
        repository=repository,
        registry=registry,
        agent_client=FakeAgentClient([
            NextAction(
                action="call_tool",
                tool_name="fetch_web_page",
                arguments={"url": wrong_year},
            ),
            NextAction(action="finish", reason_summary="已找到同年完整来源"),
        ]),
        allowed_roots=[tmp_path],
        verifier=EvidenceVerifier(),
    ).run(state.case_id)

    assert [call["url"] for call in fetched.calls] == [wrong_year, correct_year]
    assert collected.calls == []
    assert outcome.state.latest_verification is not None
    assert outcome.state.latest_verification.year_match == "yes"
    assert outcome.state.latest_verification.coverage_complete == "yes"
    store.close()


def test_candidate_queue_turns_premature_finish_into_next_fetch(tmp_path: Path) -> None:
    store = Store(tmp_path / "candidate-finish-redirect.db")
    batch_id = store.create_batch("candidate-finish-redirect")
    repository = CaseRepository(store)
    first = "https://news.example.cn/candidate-1"
    second = "https://www.chinazy.org/info/1014/15997.htm"
    state, _created = repository.create_or_get(CaseSeed(
        batch_id=batch_id,
        resource_code="06090003",
        award_name="全国高校辅导员年度人物",
        year="2023",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="Agent 不得在候选队列未处理时空转",
    ))
    registry = ToolRegistry()
    register_fake_tool(registry, "search_official_award", [ToolResult(ok=True, data={
            "candidates": [
                {"url": first, "source_level": "publisher_secondary", "rank": 1, "title": "2023年候选人公示名单"},
                {"url": second, "source_level": "unknown", "rank": 2, "title": "2023年获奖结果"},
        ],
        "candidate_count": 2,
        "official_candidate_count": 0,
    })])
    fetched = register_fake_tool(registry, "fetch_web_page", [
        ToolResult(ok=True, source_url=first, evidence_facts=[_fact(
            url=first, expected=20, observed=0, complete=False,
        )]),
        ToolResult(ok=True, source_url=second, evidence_facts=[_fact(
            url=second, expected=20, observed=20, complete=True,
        )]),
    ])
    client = FakeAgentClient([
        NextAction(action="call_tool", tool_name="search_official_award"),
        NextAction(action="finish", reason_summary="过早结束一"),
        NextAction(action="finish", reason_summary="过早结束二"),
        NextAction(action="finish", reason_summary="候选核验完成"),
    ])

    outcome = EvidenceHarness(
        repository=repository,
        registry=registry,
        agent_client=client,
        allowed_roots=[tmp_path],
        verifier=EvidenceVerifier(),
    ).run(state.case_id)

    assert [call["url"] for call in fetched.calls] == [first, second]
    assert "pending_candidate_processed_without_agent_turn" in (
        outcome.state.reason_codes
    )
    assert outcome.state.latest_verification is not None
    assert outcome.state.latest_verification.coverage_complete == "yes"
    store.close()


def test_repeated_known_source_redirects_to_attachment_recovery(
    tmp_path: Path,
) -> None:
    store = Store(tmp_path / "repeat-to-attachment.db")
    batch_id = store.create_batch("repeat-to-attachment")
    repository = CaseRepository(store)
    known = "https://publisher.example.cn/award"
    first = "https://school.example.edu.cn/candidate.pdf"
    second = "https://education.example.gov.cn/candidate.doc"
    state, _created = repository.create_or_get(CaseSeed(
        batch_id=batch_id,
        resource_code="NEW-RESOURCE",
        award_name="示例奖",
        year="2025",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="候选耗尽后继续附件恢复",
        known_urls=[known],
    ))
    registry = ToolRegistry()
    fetched = register_fake_tool(registry, "fetch_web_page", [
        ToolResult(ok=True, source_url=known, evidence_facts=[_fact(
            url=known,
            level="publisher_secondary",
            observed=20,
            complete=False,
        )]),
        ToolResult(ok=True, source_url=first, evidence_facts=[_fact(
            url=first,
            expected=20,
            observed=0,
            complete=False,
        )]),
        ToolResult(ok=True, source_url=second, evidence_facts=[_fact(
            url=second,
            expected=20,
            observed=0,
            complete=False,
        )]),
    ])
    searched = register_fake_tool(registry, "search_official_award", [
        ToolResult(ok=True, data={
            "strategy": "broad",
                "candidates": [
                    {"url": first, "source_level": "institutional_secondary", "rank": 1, "title": "示例奖2025名单"},
                    {"url": second, "source_level": "official_secondary", "rank": 2, "title": "示例奖2025结果公示"},
            ],
            "candidate_count": 2,
            "official_candidate_count": 0,
        }),
        ToolResult(ok=True, data={
            "strategy": "attachment",
            "candidates": [],
            "candidate_count": 0,
            "official_candidate_count": 0,
        }),
    ])
    client = FakeAgentClient([
        NextAction(action="call_tool", tool_name="search_official_award"),
        NextAction(action="finish", reason_summary="候选一尚未检查"),
        NextAction(action="finish", reason_summary="候选二尚未检查"),
        NextAction(
            action="call_tool",
            tool_name="fetch_web_page",
            arguments={"url": known},
        ),
    ])

    outcome = EvidenceHarness(
        repository=repository,
        registry=registry,
        agent_client=client,
        allowed_roots=[tmp_path],
        verifier=EvidenceVerifier(),
    ).run(state.case_id)

    assert len(fetched.calls) == 3
    assert [call["strategy"] for call in searched.calls] == ["broad"]
    assert "attachment_search_started_without_agent_turn" not in (
        outcome.state.reason_codes
    )
    assert len(client.calls) == 0
    assert len(outcome.state.llm_usage) == 0
    assert outcome.stopped_reason == "bounded_search_candidates_exhausted"
    assert outcome.state.latest_verification is not None
    assert outcome.state.latest_verification.coverage_complete == "no"
    store.close()


def test_generic_failed_search_candidate_is_extracted_before_next_candidate(
    tmp_path: Path,
) -> None:
    store = Store(tmp_path / "candidate-extract-fallback.db")
    batch_id = store.create_batch("candidate-extract-fallback")
    repository = CaseRepository(store)
    candidate = "https://example.gov.cn/award/result"
    state, _created = repository.create_or_get(CaseSeed(
        batch_id=batch_id,
        resource_code="NEW-RESOURCE",
        award_name="示例奖",
        year="2025",
        trigger_codes=["SOURCE_UNREACHABLE"],
        objective="候选直连失败后使用搜索提取兜底",
    ))
    registry = ToolRegistry()
    register_fake_tool(registry, "search_official_award", [ToolResult(
        ok=True,
        data={
            "strategy": "broad",
            "candidates": [{
                "url": candidate,
                "source_level": "official_secondary",
                "rank": 1,
                "title": "2025年示例奖认定名单",
            }],
            "candidate_count": 1,
            "official_candidate_count": 1,
        },
    )])
    fetched = register_fake_tool(registry, "fetch_web_page", [
        ToolResult.failure("HTTP_ERROR", "HTTP 404", source_url=candidate),
    ])
    extracted = register_fake_tool(registry, "extract_search_document", [ToolResult(
        ok=True,
        source_url=candidate,
        evidence_facts=[_fact(url=candidate)],
    )])
    client = FakeAgentClient([
        NextAction(action="call_tool", tool_name="search_official_award"),
        NextAction(action="call_tool", tool_name="fetch_web_page",
                   arguments={"url": candidate}),
        NextAction(action="finish", reason_summary="直连失败后过早结束"),
        NextAction(action="finish", reason_summary="提取兜底完成"),
    ])

    outcome = EvidenceHarness(
        repository=repository,
        registry=registry,
        agent_client=client,
        allowed_roots=[tmp_path],
        verifier=EvidenceVerifier(),
    ).run(state.case_id)

    assert len(fetched.calls) == len(extracted.calls) == 1
    assert extracted.calls[0]["url"] == candidate
    assert "failed_candidate_extracted_without_agent_turn" in (
        outcome.state.reason_codes
    )
    assert outcome.state.latest_verification is not None
    assert outcome.state.latest_verification.coverage_complete == "yes"
    store.close()


def test_s06_search_extract_tool_is_an_offline_verifier_ready_fallback(tmp_path: Path) -> None:
    url = "https://www.moe.gov.cn/jyb_xxgk/award.html"
    provider = FakeSearchProvider([ExtractResponse(
        provider="fake",
        url=url,
        text="教育部公布2025年示例奖名单，共25项。第一章 获奖团队名单。",
    )])
    registry = build_default_registry(search_provider_factory=lambda: provider)
    executor = SafeToolExecutor(registry)
    context = ToolExecutionContext.create([tmp_path])

    result = executor.execute("extract_search_document", {
        "url": url,
        "expected_award_name": "示例奖",
        "expected_year": "2025",
        "section_keywords": ["获奖团队名单"],
    }, context)

    assert result.ok and result.evidence_facts
    fact = result.evidence_facts[0]
    assert fact.target_match == fact.year_match == "yes"
    assert fact.source_level in {"official_primary", "official_secondary"}
    assert provider.calls == [{"operation": "extract", "url": url, "query_hint": ""}]


def test_s06_failed_known_source_is_extracted_before_premature_manual(
    tmp_path: Path,
) -> None:
    store = Store(tmp_path / "s06-recovery.db")
    batch_id = store.create_batch("s06-offline")
    repository = CaseRepository(store)
    known = [
        "https://www.moe.gov.cn/old-page-1",
        "https://www.moe.gov.cn/old-page-2",
    ]
    recovered = "https://www.moe.gov.cn/recovered-section"
    state, _created = repository.create_or_get(CaseSeed(
        batch_id=batch_id,
        resource_code="02040005",
        award_name="示例奖",
        year="2025",
        trigger_codes=["SOURCE_UNREACHABLE"],
        objective="恢复教育部名单章节",
        known_urls=known,
    ))
    registry = ToolRegistry()
    fetched = register_fake_tool(registry, "fetch_web_page", [
        ToolResult.failure("HTTP_ERROR", "HTTP 500"),
    ])
    searched = register_fake_tool(registry, "search_official_award", [ToolResult(
        ok=True,
        data={
            "candidates": [{
                "url": recovered,
                "source_level": "official_primary",
                "rank": 1,
            }],
            "candidate_count": 1,
            "official_candidate_count": 1,
        },
        warnings=["search_results_are_leads_not_evidence"],
    )])
    extracted = register_fake_tool(registry, "extract_search_document", [ToolResult(
        ok=True,
        evidence_facts=[_fact(url=recovered)],
    )])
    client = FakeAgentClient([
        NextAction(action="call_tool", tool_name="fetch_web_page", arguments={"url": known[0]}),
        NextAction(action="manual", reason_summary="直连失败"),
        NextAction(action="call_tool", tool_name="search_official_award",
                   arguments={"award_name": "示例奖"}),
        NextAction(action="call_tool", tool_name="extract_search_document",
                   arguments={"url": recovered}),
        NextAction(action="finish", reason_summary="域名恢复和章节提取完成"),
    ])
    outcome = EvidenceHarness(
        repository=repository,
        registry=registry,
        agent_client=client,
        allowed_roots=[tmp_path],
        verifier=EvidenceVerifier(),
    ).run(state.case_id)

    assert outcome.stopped_reason == "recommendation_ready"
    assert len(fetched.calls) == 1
    assert searched.calls == []
    assert len(extracted.calls) == 1
    assert extracted.calls[0]["url"] == known[0]
    assert "failed_candidate_extracted_without_agent_turn" in (
        outcome.state.reason_codes
    )
    store.close()


def test_s06_authoritative_partial_extract_enters_verifier_without_candidate_wander(
    tmp_path: Path,
) -> None:
    store = Store(tmp_path / "s06-partial.db")
    batch_id = store.create_batch("s06-partial")
    repository = CaseRepository(store)
    recovered = "https://www.moe.gov.cn/recovered-award-page"
    unrelated = "https://www.moe.gov.cn/application-notice"
    state, _created = repository.create_or_get(CaseSeed(
        batch_id=batch_id,
        resource_code="02040005",
        award_name="示例奖",
        year="2025",
        trigger_codes=["COVERAGE_UNKNOWN"],
        objective="核验搜索索引恢复的权威名单",
    ))
    registry = ToolRegistry()
    extracted = register_fake_tool(registry, "extract_search_document", [ToolResult(
        ok=True,
        evidence_facts=[_fact(
            url=recovered,
            level="official_secondary",
            expected=190,
            observed=112,
            complete=False,
        ).model_copy(update={
            "missing_items": ["未覆盖团队"],
            "missing_item_count": 78,
            "missing_evidence": ["权威页面正文仅恢复到第112项"],
        })],
    )])
    fetched = register_fake_tool(registry, "fetch_web_page", [
        ToolResult(ok=True, source_url=unrelated),
    ])
    state.evidence_progress.candidates = [
        EvidenceCandidate(
            url=recovered,
            source_level="official_secondary",
            provider="fake",
            rank=1,
            status="failed",
            attempts=1,
            status_reason="HTTP_ERROR",
        ),
        EvidenceCandidate(
            url=unrelated,
            source_level="official_primary",
            provider="fake",
            rank=2,
        ),
    ]
    state.evidence_progress.phase = "candidate_recovery"
    state.evidence_progress.source_failures = 1
    state.tool_trace.append(ToolObservation(
        call_id="failed-recovered",
        tool_name="fetch_web_page",
        started_at="2025-01-01T00:00:00Z",
        finished_at="2025-01-01T00:00:01Z",
        duration_ms=1000,
        input_summary={"url": recovered},
        ok=False,
        error_code="HTTP_ERROR",
    ))
    repository.save(state)
    client = FakeAgentClient([])

    outcome = EvidenceHarness(
        repository=repository,
        registry=registry,
        agent_client=client,
        allowed_roots=[tmp_path],
        verifier=EvidenceVerifier(),
    ).run(state.case_id)

    assert len(extracted.calls) == 1
    assert fetched.calls == []
    assert outcome.stopped_reason == (
        "authoritative_partial_coverage_requires_supplement"
    )
    assert outcome.state.latest_verification is not None
    assert outcome.state.latest_verification.coverage_complete == "no"
    assert "权威页面正文仅恢复到第112项" in (
        outcome.state.latest_verification.missing_evidence
    )
    assert "authoritative_partial_coverage_requires_supplement" in (
        outcome.state.reason_codes
    )
    store.close()

"""M5.1 Tool contracts, registry, executor and Fake Tool tests."""

from __future__ import annotations

import time

import openpyxl
from pydantic import BaseModel

from award_audit.agent.toolkit import (
    SafeToolExecutor,
    ToolBudgetLimits,
    ToolExecutionContext,
    ToolRegistry,
    ToolResult,
    ToolSpec,
    build_default_registry,
    web,
)
from award_audit.agent.toolkit.testing import register_fake_tool


class EchoInput(BaseModel):
    value: str
    token: str = ""
    url: str = ""


def _context(tmp_path, **limits):  # noqa: ANN001, ANN003
    return ToolExecutionContext.create([tmp_path], ToolBudgetLimits(**limits))


def test_tool_spec_renders_openai_schema() -> None:
    spec = ToolSpec(name="echo_tool", description="Echo a value.", input_model=EchoInput)
    rendered = spec.openai_schema()
    function = rendered["function"]
    assert function["name"] == "echo_tool"
    assert function["parameters"]["properties"]["value"]["type"] == "string"
    assert "input_model" not in spec.model_dump()


def test_default_registry_exposes_m51_through_m53_tools_only() -> None:
    registry = build_default_registry()
    assert [spec.name for spec in registry.specs()] == [
        "fetch_web_page",
        "download_evidence",
        "verify_page_image_roster",
        "collect_spreadsheet_attachments",
        "parse_spreadsheet",
        "inspect_pdf",
        "extract_pdf_text",
        "render_pdf_pages",
        "ocr_image",
        "vision_extract_roster",
        "compare_roster",
        "search_official_award",
        "extract_search_document",
    ]
    assert len(registry.openai_tools()) == 13


def test_registry_rejects_duplicates() -> None:
    registry = ToolRegistry()
    spec = ToolSpec(name="echo_tool", description="Echo.", input_model=EchoInput)
    registry.register(spec, lambda _args, _ctx: ToolResult(ok=True))
    try:
        registry.register(spec, lambda _args, _ctx: ToolResult(ok=True))
    except ValueError as exc:
        assert "already registered" in str(exc)
    else:
        raise AssertionError("duplicate registration must fail")


def test_executor_whitelist_input_and_output_errors_are_structured(tmp_path) -> None:
    registry = ToolRegistry()
    registry.register(
        ToolSpec(name="bad_output", description="Bad output.", input_model=EchoInput),
        lambda _args, _ctx: {"ok": False},
    )
    executor = SafeToolExecutor(registry)
    context = _context(tmp_path)

    unknown = executor.execute("not_registered", {}, context)
    invalid = executor.execute("bad_output", {}, context)
    bad_output = executor.execute("bad_output", {"value": "x"}, context)

    assert unknown.error_code == "TOOL_NOT_REGISTERED"
    assert invalid.error_code == "TOOL_INPUT_INVALID"
    assert bad_output.error_code == "TOOL_OUTPUT_INVALID"
    assert len(context.trace) == 3 and all(not item.ok for item in context.trace)


def test_executor_contains_exception_and_redacts_trace(tmp_path) -> None:
    registry = ToolRegistry()

    def explode(args, _context):  # noqa: ANN001
        raise RuntimeError(f"failed {args.url}")

    registry.register(ToolSpec(name="explode_tool", description="Explode.", input_model=EchoInput),
                      explode)
    context = _context(tmp_path)
    result = SafeToolExecutor(registry).execute(
        "explode_tool",
        {"value": "x", "token": "top-secret", "url": "https://x.cn/a?token=query-secret"},
        context,
    )
    trace_text = context.trace[0].model_dump_json()
    assert result.error_code == "TOOL_EXECUTION_ERROR"
    assert "top-secret" not in trace_text and "query-secret" not in trace_text
    assert "query-secret" not in result.error_message
    assert trace_text.count("[REDACTED]") >= 2


def test_executor_timeout_and_budget(tmp_path) -> None:
    registry = ToolRegistry()

    def slow(_args, _context):  # noqa: ANN001
        time.sleep(0.1)
        return ToolResult(ok=True)

    registry.register(ToolSpec(name="slow_tool", description="Slow.", input_model=EchoInput,
                               timeout_seconds=0.01), slow)
    context = _context(tmp_path, max_calls=1)
    executor = SafeToolExecutor(registry)
    timed_out = executor.execute("slow_tool", {"value": "x"}, context)
    over_budget = executor.execute("slow_tool", {"value": "x"}, context)
    assert timed_out.error_code == "TOOL_TIMEOUT"
    assert over_budget.error_code == "TOOL_BUDGET_EXCEEDED"
    assert context.budget.calls == 1


def test_search_budget_is_independent(tmp_path) -> None:
    registry = ToolRegistry()
    registry.register(ToolSpec(name="search_fake", description="Search.", input_model=EchoInput,
                               kind="search"), lambda _a, _c: ToolResult(ok=True))
    context = _context(tmp_path, max_calls=3, max_searches=1)
    executor = SafeToolExecutor(registry)
    assert executor.execute("search_fake", {"value": "x"}, context).ok
    assert executor.execute("search_fake", {"value": "x"}, context).error_code == (
        "TOOL_BUDGET_EXCEEDED"
    )
    assert context.budget.calls == 1 and context.budget.searches == 1


def test_fake_tool_records_calls_and_exhausts(tmp_path) -> None:
    registry = ToolRegistry()
    fake = register_fake_tool(registry, "fake_tool", [ToolResult(ok=True, data={"n": 1})],
                              input_model=EchoInput)
    executor = SafeToolExecutor(registry)
    context = _context(tmp_path)
    assert executor.execute("fake_tool", {"value": "hello"}, context).data == {"n": 1}
    exhausted = executor.execute("fake_tool", {"value": "again"}, context)
    assert exhausted.error_code == "FAKE_TOOL_EXHAUSTED"
    assert [call["value"] for call in fake.calls] == ["hello", "again"]


def test_default_parse_spreadsheet_contract(tmp_path) -> None:
    path = tmp_path / "award.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["作品", "学校"])
    sheet.append(["甲", "甲大学"])
    workbook.save(path)

    context = _context(tmp_path)
    result = SafeToolExecutor(build_default_registry()).execute(
        "parse_spreadsheet", {"path": str(path)}, context
    )
    assert result.ok and result.data["rows"][1] == ["甲", "甲大学"]
    assert len(result.sha256) == 64 and result.content_type.endswith("spreadsheetml.sheet")


def test_fetch_supplied_media_page_returns_structured_roster_evidence(
    tmp_path, monkeypatch
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XRYXM"])
    sheet.append(["入选人姓名"])
    sheet.append(["张甲"])
    sheet.append(["李乙"])
    workbook.save(submitted)
    monkeypatch.setattr(
        web,
        "fetch_page",
        lambda url: web.PageContent(
            url=url,
            status=200,
            title="2位教师入选！2025年全国最美教师公示",
            text="公示名单：张甲、李乙",
        ),
    )

    context = _context(tmp_path)
    result = SafeToolExecutor(build_default_registry()).execute(
        "fetch_web_page",
        {
            "url": "https://news.eol.cn/example",
            "expected_award_name": "最美教师",
            "expected_year": "2025",
            "submitted_path": str(submitted),
            "match_fields": ["XRYXM"],
        },
        context,
    )

    assert result.ok
    assert result.data["source_level"] == "publisher_secondary"
    assert result.data["observed_award_name"] == "最美教师"
    assert result.data["observed_year"] == "2025"
    assert result.data["observed_count"] == result.data["expected_count"] == 2
    assert result.data["coverage_complete"] is True
    facts = context.trace[0].output_summary["verification_facts"]
    assert facts["source_level"] == "publisher_secondary"
    assert facts["coverage_complete"] is True


def test_fetch_page_lists_submitted_names_missing_from_the_online_source(
    tmp_path, monkeypatch
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XRYXM"])
    sheet.append(["入选人姓名"])
    sheet.append(["张甲"])
    sheet.append(["李乙"])
    sheet.append(["王丙"])
    workbook.save(submitted)
    monkeypatch.setattr(
        web,
        "fetch_page",
        lambda url: web.PageContent(
            url=url,
            status=200,
            title="2025年全国最美教师公示",
            text="公示名单：张甲、王丙",
        ),
    )

    context = _context(tmp_path)
    result = SafeToolExecutor(build_default_registry()).execute(
        "fetch_web_page",
        {
            "url": "https://www.moe.gov.cn/example",
            "expected_award_name": "最美教师",
            "expected_year": "2025",
            "submitted_path": str(submitted),
            "match_fields": ["XRYXM"],
            "expected_scope_count": 3,
        },
        context,
    )

    assert result.ok
    assert result.data["matched_items"] == ["张甲", "王丙"]
    assert result.data["missing_items"] == ["李乙"]
    assert result.data["extra_items"] == []
    assert result.data["coverage_complete"] is False
    assert "提交名单有、该来源未找到：李乙" in result.evidence_facts[0].missing_evidence
    facts = context.trace[0].output_summary["verification_facts"]
    assert facts["missing_items"] == ["李乙"]


def test_fetch_page_defers_name_difference_until_roster_images_are_processed(
    tmp_path, monkeypatch
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XRYXM"])
    sheet.append(["入选人姓名"])
    sheet.append(["张甲"])
    sheet.append(["李乙"])
    workbook.save(submitted)
    monkeypatch.setattr(
        web,
        "fetch_page",
        lambda url: web.PageContent(
            url=url,
            status=200,
            title="2023年最美高校辅导员候选人公示",
            text="名单详见图片",
            images=["https://example.cn/roster-1.jpg"],
        ),
    )

    context = _context(tmp_path)
    result = SafeToolExecutor(build_default_registry()).execute(
        "fetch_web_page",
        {
            "url": "https://example.cn/award",
            "expected_award_name": "全国高校辅导员年度人物",
            "award_aliases": ["最美高校辅导员"],
            "section_keywords": ["最美高校辅导员"],
            "expected_year": "2023",
            "submitted_path": str(submitted),
            "match_fields": ["XRYXM"],
            "expected_scope_count": 2,
            "page_total_count": 5,
        },
        context,
    )

    assert result.ok and result.data["next_evidence_stage"] == "image_processing"
    assert result.data["missing_items"] == []
    assert result.data["unresolved_items"] == ["张甲", "李乙"]
    assert not any("该来源未找到" in item for item in result.evidence_facts[0].missing_evidence)


def test_fetch_page_matches_multi_person_cell_but_queues_unparsed_image(
    tmp_path, monkeypatch
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XRYXM"])
    sheet.append(["入选人姓名"])
    sheet.append(["李桂枝;王伟江"])
    workbook.save(submitted)
    monkeypatch.setattr(
        web,
        "fetch_page",
        lambda url: web.PageContent(
            url=url,
            status=200,
            title="2025年全国最美教师名单",
            text="李桂枝来自甲校。王伟江来自乙校。",
            images=["https://example.cn/decorative.jpg"],
        ),
    )

    result = SafeToolExecutor(build_default_registry()).execute(
        "fetch_web_page",
        {
            "url": "https://www.moe.gov.cn/award",
            "expected_award_name": "最美教师",
            "expected_year": "2025",
            "submitted_path": str(submitted),
            "match_fields": ["XRYXM"],
            "expected_scope_count": 1,
        },
        _context(tmp_path),
    )

    assert result.ok and result.data["coverage_complete"] is True
    assert result.data["matched_items"] == ["李桂枝;王伟江"]
    assert result.data["split_matched_items"] == ["李桂枝;王伟江"]
    assert result.data["missing_items"] == []
    assert "next_evidence_stage" not in result.data
    assert "candidate_image_urls" not in result.data
    assert result.data["unresolved_items"] == []


def test_fetch_page_keeps_s02_name_difference_unresolved_while_images_remain(
    tmp_path, monkeypatch
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XRYXM"])
    sheet.append(["入选人姓名"])
    sheet.append(["张甲"])
    sheet.append(["李桂枝;王伟江"])
    workbook.save(submitted)
    monkeypatch.setattr(
        web,
        "fetch_page",
        lambda url: web.PageContent(
            url=url,
            status=200,
            title="2025年全国最美教师名单",
            text="张甲",
            images=["https://www.moe.gov.cn/roster.jpg"],
        ),
    )

    result = SafeToolExecutor(build_default_registry()).execute(
        "fetch_web_page",
        {
            "url": "https://www.moe.gov.cn/award",
            "expected_award_name": "全国最美教师",
            "expected_year": "2025",
            "submitted_path": str(submitted),
            "match_fields": ["XRYXM"],
            "expected_scope_count": 2,
        },
        _context(tmp_path),
    )

    assert result.ok and result.data["coverage_complete"] is False
    assert result.data["missing_items"] == []
    assert result.data["unresolved_items"] == ["李桂枝", "王伟江"]
    assert result.data["candidate_image_urls"] == [
        "https://www.moe.gov.cn/roster.jpg"
    ]


def test_fetch_page_prioritizes_spreadsheet_attachments_over_page_images(
    tmp_path, monkeypatch
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XRYXM"])
    sheet.append(["入选人姓名"])
    sheet.append(["王奇"])
    workbook.save(submitted)
    attachment_url = "https://www.chinazy.org/system/_content/download.jsp?id=1"
    monkeypatch.setattr(
        web,
        "fetch_page",
        lambda url: web.PageContent(
            url=url,
            status=200,
            title="2023年最美高校辅导员暨高校辅导员年度人物候选人公示",
            text="名单见附件",
            attachments=[web.Attachment(
                text="2023年最美高校辅导员候选人.xlsx",
                url=attachment_url,
                is_excel=True,
            )],
            images=["https://www.chinazy.org/images/roster-preview.png"],
        ),
    )

    result = SafeToolExecutor(build_default_registry()).execute(
        "fetch_web_page",
        {
            "url": "https://www.chinazy.org/info/1014/15997.htm",
            "expected_award_name": "全国高校辅导员年度人物",
            "award_aliases": ["最美高校辅导员", "高校辅导员年度人物"],
            "expected_year": "2023",
            "submitted_path": str(submitted),
            "match_fields": ["XRYXM"],
            "expected_scope_count": 1,
        },
        _context(tmp_path),
    )

    assert result.ok
    assert result.data["next_evidence_stage"] == "spreadsheet_processing"
    assert result.data["candidate_attachment_urls"] == [attachment_url]
    assert result.data["attachment_names"] == [
        "2023年最美高校辅导员候选人.xlsx"
    ]
    assert "candidate_image_urls" not in result.data
    assert result.data["missing_items"] == []
    assert result.data["unresolved_items"] == ["王奇"]


def test_fetch_page_lists_person_vs_group_identity_difference(
    tmp_path, monkeypatch
) -> None:  # noqa: ANN001
    submitted = tmp_path / "submitted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["XRYXM"])
    sheet.append(["入选人姓名"])
    sheet.append(["丁美珍"])
    sheet.append(["马琼"])
    sheet.append(["李桂枝;王伟江"])
    workbook.save(submitted)
    monkeypatch.setattr(
        web,
        "fetch_page",
        lambda url: web.PageContent(
            url=url,
            status=200,
            title="中央宣传部、教育部联合发布2025年最美教师",
            text=(
                "丁美珍、马琼等2名同志和保定学院毕业生赴疆任教群体代表"
                "光荣入选。"
            ),
        ),
    )

    result = SafeToolExecutor(build_default_registry()).execute(
        "fetch_web_page",
        {
            "url": "https://www.moe.gov.cn/award",
            "expected_award_name": "全国最美教师",
            "expected_year": "2025",
            "submitted_path": str(submitted),
            "match_fields": ["XRYXM"],
            "expected_scope_count": 3,
        },
        _context(tmp_path),
    )

    assert result.ok and result.data["coverage_complete"] is False
    assert result.data["matched_items"] == ["丁美珍", "马琼"]
    assert result.data["missing_items"] == ["李桂枝", "王伟江"]
    assert result.data["extra_items"] == ["保定学院毕业生赴疆任教群体代表"]
    assert result.data["comparison_note"] == "来源使用群体名额，提交材料使用个人姓名"


def test_parse_html_ignores_known_government_header_images() -> None:
    _text, _attachments, images = web.parse_html(
        """
        <img src="/images/scy_jyb_lgo_03.png">
        <img src="/images/red.png">
        <img src="/files/award-roster.jpg">
        """,
        "https://www.moe.gov.cn/page.html",
    )

    assert images == ["https://www.moe.gov.cn/files/award-roster.jpg"]


def test_html_discovery_separates_detail_pages_from_download_attachments() -> None:
    html = """
    <a href="/cw/contestNews/detail/award-1">附件：获奖名单通知</a>
    <a href="/sysFile/downFile.do?fileId=pdf-1">附件：获奖名单.pdf</a>
    """
    base = "https://cpipc.acge.org.cn/cw/contestNews/list/contest/1"

    _text, attachments, _images = web.parse_html(html, base)
    related = web.discover_related_pages(html, base)

    assert [item.url for item in attachments] == [
        "https://cpipc.acge.org.cn/sysFile/downFile.do?fileId=pdf-1"
    ]
    assert [item.url for item in related] == [
        "https://cpipc.acge.org.cn/cw/contestNews/detail/award-1"
    ]


def test_default_download_records_complete_artifact(tmp_path, monkeypatch) -> None:  # noqa: ANN001
    def fake_download(_url, destination, **_kwargs):  # noqa: ANN001
        path = destination / "evidence.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["作品"])
        sheet.append(["甲"])
        workbook.save(path)
        return path

    monkeypatch.setattr(web, "download_file", fake_download)
    context = _context(tmp_path)
    result = SafeToolExecutor(build_default_registry()).execute(
        "download_evidence",
        {"url": "https://example.com/list.xlsx", "destination_dir": str(tmp_path)},
        context,
    )
    artifact = result.artifacts[0]
    assert result.ok and artifact.source_url == "https://example.com/list.xlsx"
    assert artifact.fetched_at and len(artifact.sha256) == 64 and artifact.size_bytes > 0
    assert context.budget.downloads == 1
    assert context.budget.download_bytes == artifact.size_bytes


def test_default_tools_reject_unsafe_url_and_path_without_io(tmp_path) -> None:
    executor = SafeToolExecutor(build_default_registry())
    context = _context(tmp_path)
    unsafe_url = executor.execute("fetch_web_page", {"url": "http://127.0.0.1/admin"}, context)
    outside = executor.execute(
        "parse_spreadsheet", {"path": str(tmp_path.parent / "outside.xlsx")}, context
    )
    assert unsafe_url.error_code == "UNSAFE_URL"
    assert outside.error_code == "UNSAFE_PATH"

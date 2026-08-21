"""M5.1 URL, path, type and size hard-boundary tests."""

from __future__ import annotations

import shutil
import socket
import zipfile

import openpyxl
import pytest

from award_audit.agent.toolkit import (
    SafeToolExecutor,
    ToolExecutionContext,
    build_default_registry,
    web,
)
from award_audit.agent.toolkit.safety import (
    FileInspection,
    UnsafeFileError,
    UnsafePathError,
    UnsafeUrlError,
    inspect_evidence_file,
    validate_local_path,
    validate_public_url,
)


@pytest.mark.parametrize("url", [
    "file:///etc/passwd",
    "http://localhost/admin",
    "http://127.0.0.1/admin",
    "http://10.0.0.2/",
    "http://169.254.169.254/latest/meta-data",
    "http://[::1]/",
    "https://user:password@example.com/",
])
def test_url_rejects_non_public_targets_without_dns(url: str) -> None:
    with pytest.raises(UnsafeUrlError):
        validate_public_url(url, resolve_dns=False)


def test_url_accepts_public_literal_and_rejects_private_dns() -> None:
    assert validate_public_url("https://1.1.1.1/a", resolve_dns=False).startswith("https://")

    def private_dns(_host, _port, **_kwargs):  # noqa: ANN001
        return [(socket.AF_INET, socket.SOCK_STREAM, 6, "", ("10.1.2.3", 443))]

    with pytest.raises(UnsafeUrlError, match="non-public"):
        validate_public_url("https://example.com/a", resolver=private_dns)


def test_fetch_checks_redirect_before_second_request(monkeypatch) -> None:  # noqa: ANN001
    class Response:
        status_code = 302
        headers = {"location": "http://127.0.0.1/private"}
        text = ""
        url = "https://example.com/start"

    class Client:
        def __init__(self) -> None:
            self.urls: list[str] = []

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return None

        def get(self, url):  # noqa: ANN001
            self.urls.append(url)
            return Response()

    client = Client()
    import httpx

    monkeypatch.setattr(httpx, "Client", lambda **_kwargs: client)
    monkeypatch.setattr(web, "validate_public_url",
                        lambda url: (_ for _ in ()).throw(UnsafeUrlError("private redirect"))
                        if "127.0.0.1" in url else url)
    with pytest.raises(UnsafeUrlError, match="private redirect"):
        web.fetch_page("https://example.com/start")
    assert client.urls == ["https://example.com/start"]


def test_fetch_decodes_html_meta_charset(monkeypatch) -> None:  # noqa: ANN001
    import httpx

    html = (
        '<html><head><meta http-equiv="Content-Type" '
        'content="text/html; charset=gb2312"><title>全国高校辅导员年度人物</title>'
        "</head><body>2023年候选人公示</body></html>"
    )
    response = httpx.Response(
        200,
        content=html.encode("gb18030"),
        request=httpx.Request("GET", "https://example.com/list"),
    )

    class Client:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return None

        def get(self, _url):  # noqa: ANN001
            return response

    monkeypatch.setattr(httpx, "Client", lambda **_kwargs: Client())
    monkeypatch.setattr(web, "validate_public_url", lambda url: url)
    page = web.fetch_page("https://example.com/list")
    assert page.title == "全国高校辅导员年度人物"
    assert "2023年候选人公示" in page.text


def test_html_discovery_drops_unsafe_assets() -> None:
    html = (
        '<a href="http://127.0.0.1/list.xlsx">附件名单</a>'
        '<a href="file:///etc/list.pdf">附件PDF</a>'
        '<img src="http://10.0.0.2/list.png">'
    )
    _text, attachments, images = web.parse_html(html, "https://example.com/page")
    assert attachments == [] and images == []


def test_html_discovery_filters_decorative_images() -> None:
    html = (
        '<img src="/images/share.png">'
        '<img src="/images/arrow.png">'
        '<img src="/images/gzh_qr.png">'
        '<img src="/img/bt01.jpg">'
        '<img src="/upload/official-roster-1.jpg">'
        '<img src="/upload/official-roster-1.jpg">'
    )
    _text, _attachments, images = web.parse_html(html, "https://example.com/page")
    assert images == ["https://example.com/upload/official-roster-1.jpg"]


def test_path_must_stay_under_allowed_root(tmp_path) -> None:
    inside = tmp_path / "inside.txt"
    inside.write_text("ok", encoding="utf-8")
    assert validate_local_path(inside, [tmp_path], file_only=True) == inside.resolve()
    with pytest.raises(UnsafePathError, match="outside"):
        validate_local_path(tmp_path.parent / "outside.txt", [tmp_path], must_exist=False)
    with pytest.raises(UnsafePathError, match="environment"):
        validate_local_path(tmp_path / ".env", [tmp_path], must_exist=False)


def _make_xlsx(path) -> None:  # noqa: ANN001
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["作品"])
    sheet.append(["甲"])
    workbook.save(path)


def test_file_magic_hash_and_extension_checks(tmp_path) -> None:
    source = tmp_path / "safe.xlsx"
    _make_xlsx(source)
    inspected = inspect_evidence_file(source, max_bytes=1024 * 1024,
                                      allowed_kinds={"xlsx"})
    assert inspected.kind == "xlsx" and len(inspected.sha256) == 64

    wrong_extension = tmp_path / "wrong.pdf"
    shutil.copyfile(source, wrong_extension)
    with pytest.raises(UnsafeFileError, match="extension"):
        inspect_evidence_file(wrong_extension, max_bytes=1024 * 1024)
    unknown_extension = tmp_path / "unknown.bin"
    shutil.copyfile(source, unknown_extension)
    with pytest.raises(UnsafeFileError, match="extension"):
        inspect_evidence_file(unknown_extension, max_bytes=1024 * 1024)
    with pytest.raises(UnsafeFileError, match="exceeds"):
        inspect_evidence_file(source, max_bytes=1)


def test_file_rejects_executable_and_generic_zip(tmp_path) -> None:
    executable = tmp_path / "payload.bin"
    executable.write_bytes(b"MZ" + b"\x00" * 32)
    with pytest.raises(UnsafeFileError, match="executable"):
        inspect_evidence_file(executable, max_bytes=1024)

    archive = tmp_path / "archive.zip"
    with zipfile.ZipFile(archive, "w") as handle:
        handle.writestr("payload.txt", "hello")
    with pytest.raises(UnsafeFileError, match="generic ZIP"):
        inspect_evidence_file(archive, max_bytes=1024 * 1024)


def test_download_temp_file_keeps_evidence_suffix(tmp_path, monkeypatch) -> None:  # noqa: ANN001
    import httpx

    class Response:
        status_code = 200
        headers = {
            "content-disposition": 'attachment; filename="official.xlsx"',
            "content-length": "8",
        }

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return None

        def iter_bytes(self):
            yield b"PK\x03\x04test"

    class Client:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return None

        def stream(self, _method, _url):  # noqa: ANN001
            return Response()

    inspected_suffixes: list[str] = []

    def fake_inspect(path, *, max_bytes, allowed_kinds=None):  # noqa: ANN001, ARG001
        suffix = path.suffix.lower()
        inspected_suffixes.append(suffix)
        if suffix != ".xlsx":
            raise UnsafeFileError("file extension is not an allowed evidence type")
        return FileInspection(
            kind="xlsx",
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            extension=".xlsx",
            size_bytes=8,
            sha256="a" * 64,
        )

    monkeypatch.setattr(httpx, "Client", lambda **_kwargs: Client())
    monkeypatch.setattr(web, "validate_public_url", lambda url: url)
    monkeypatch.setattr(web, "inspect_evidence_file", fake_inspect)
    downloaded = web.download_file("https://example.com/download", tmp_path)
    assert downloaded.suffix == ".xlsx" and downloaded.is_file()
    assert inspected_suffixes == [".xlsx", ".xlsx"]


def test_download_failure_does_not_delete_preexisting_evidence(tmp_path, monkeypatch) -> None:  # noqa: ANN001
    import httpx

    url = "https://example.com/evidence.pdf"
    destination = tmp_path / (web.hashlib.sha1(url.encode("utf-8")).hexdigest()[:16] + ".pdf")
    destination.write_bytes(b"preserved evidence")

    class Response:
        status_code = 200
        headers = {"content-length": "4"}

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return None

        def iter_bytes(self):
            yield b"%PDF"

    class Client:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return None

        def stream(self, _method, _url):  # noqa: ANN001
            return Response()

    monkeypatch.setattr(httpx, "Client", lambda **_kwargs: Client())
    monkeypatch.setattr(web, "validate_public_url", lambda value: value)
    monkeypatch.setattr(
        web,
        "inspect_evidence_file",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(UnsafeFileError("invalid")),
    )

    with pytest.raises(UnsafeFileError, match="invalid"):
        web.download_file(url, tmp_path, excel_only=True)

    assert destination.read_bytes() == b"preserved evidence"
    assert not list(tmp_path.glob("*.part.*"))


def test_executor_owns_optional_evidence_output_directories(
    tmp_path, monkeypatch
) -> None:  # noqa: ANN001
    import award_audit.agent.toolkit.registry as registry_module

    def fake_download(_url, destination, **_kwargs):  # noqa: ANN001
        path = destination / "official.xlsx"
        workbook = openpyxl.Workbook()
        workbook.active.append(["name"])
        workbook.save(path)
        return path

    registry = build_default_registry()
    schemas = {
        item["function"]["name"]: item["function"]["parameters"]
        for item in registry.openai_tools()
    }
    assert "destination_dir" not in schemas["download_evidence"].get("required", [])
    assert "output_dir" not in schemas["render_pdf_pages"].get("required", [])

    monkeypatch.setattr(registry_module.web, "download_file", fake_download)
    result = SafeToolExecutor(registry).execute(
        "download_evidence",
        {"url": "https://example.com/official.xlsx"},
        ToolExecutionContext.create([tmp_path]),
    )
    assert result.ok
    assert result.local_path == str((tmp_path / "official.xlsx").resolve())


def test_composite_spreadsheet_tool_merges_selected_fragments_and_counts_downloads(
    tmp_path, monkeypatch
) -> None:  # noqa: ANN001
    import award_audit.agent.toolkit.registry as registry_module

    submitted = tmp_path / "submitted.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["CSDWMC", "XCSDW"])
    sheet.append(["参赛队伍名称", "参赛单位名单"])
    sheet.append(["团队甲", "学校甲"])
    sheet.append(["团队乙", "学校乙"])
    sheet.append(["", "单位丙"])
    workbook.save(submitted)

    page_url = "https://example.gov.cn/notice"
    attachments = [
        web.Attachment(
            text="一等奖获奖团队", url="https://example.gov.cn/a", is_excel=False
        ),
        web.Attachment(
            text="二等奖获奖团队", url="https://example.gov.cn/b", is_excel=False
        ),
        web.Attachment(
            text="优秀组织单位", url="https://example.gov.cn/org", is_excel=False
        ),
        web.Attachment(
            text="获奖团队提名名单", url="https://example.gov.cn/nomination", is_excel=False
        ),
    ]
    monkeypatch.setattr(
        registry_module.web,
        "fetch_page",
        lambda _url: web.PageContent(
            url=page_url, status=200, text="", attachments=attachments
        ),
    )

    def fake_download(url, destination, **_kwargs):  # noqa: ANN001
        path = destination / (url.rsplit("/", 1)[-1] + ".xlsx")
        part = openpyxl.Workbook()
        part_sheet = part.active
        part_sheet.append(["名单"])
        part_sheet.append(["团队名称"])
        value = (
            "团队甲"
            if url.endswith("/a")
            else "团队乙"
            if url.endswith("/b")
            else "单位丙"
        )
        part_sheet.append([value])
        part.save(path)
        return path

    monkeypatch.setattr(registry_module.web, "download_file", fake_download)
    context = ToolExecutionContext.create([tmp_path])
    result = SafeToolExecutor(build_default_registry()).execute(
        "collect_spreadsheet_attachments",
        {
            "page_urls": [page_url],
            "submitted_path": str(submitted),
            "match_fields": ["CSDWMC", "XCSDW"],
            "include_attachment_keywords": ["获奖团队", "优秀组织单位"],
            "exclude_attachment_keywords": ["提名"],
        },
        context,
    )
    assert result.ok and result.data["coverage_complete"] is True
    assert result.data["attachment_count"] == 3 and len(result.artifacts) == 3
    assert result.data["submitted_match_count"] == 3
    assert context.budget.calls == 1 and context.budget.downloads == 3

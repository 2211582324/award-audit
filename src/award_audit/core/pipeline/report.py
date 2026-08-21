"""反馈报告产出：反馈意见.xlsx（一行一 Issue）+ 反馈摘要.md（批次概览）。

这是系统给乙方的主产出。Excel 按 文件 → 严重度 → 行号 排序，乙方逐条修；Markdown 给质检组一眼看清批次健康度。
"""

from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font

from award_audit.core import config
from award_audit.core.models.issue import RULES, Severity
from award_audit.core.models.triage import (
    decide_triage,
    reason_label,
    triage_label,
    triage_sort_key,
)
from award_audit.core.pipeline.engine import BatchResult

# 严重度中文与排序权重
_SEV_CN = {Severity.BLOCKER: "严重", Severity.FORMAT: "格式", Severity.REVIEW: "待复核"}
_SEV_ORDER = {Severity.BLOCKER: 0, Severity.REVIEW: 1, Severity.FORMAT: 2}
_COLUMNS = ["批次", "文件", "sheet", "数据行", "字段代码", "字段名", "规则ID", "严重度", "问题描述", "当前值", "建议"]


# 取一条 audit dict（EvidenceReport.model_dump() 形态，不含 triage）的分诊桶——派生，保 report 自足
def _triage_of(r: dict[str, Any]) -> str:
    return decide_triage(str(r.get("verdict", "")), str(r.get("confidence", "low")))


# 一条 audit dict 的降级原因中文标签（无码返回空串，展示层按需省略）
def _reasons_of(r: dict[str, Any]) -> str:
    return "；".join(reason_label(str(c)) for c in r.get("reason_codes", []))


# audit dict 的分诊排序键（转人工/低把握浮顶），供 sorted() 用
def _triage_key(r: dict[str, Any]) -> tuple[int, int]:
    return triage_sort_key(_triage_of(r), str(r.get("confidence", "low")))


# 写反馈意见.xlsx：一行一个 Issue；audit_reports 非空时增「联网核对」sheet
def write_excel_report(
    result: BatchResult,
    out_path: Path,
    audit_reports: list[dict[str, Any]] | None = None,
    audit_cases: list[dict[str, Any]] | None = None,
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None  # 新建工作簿必有活动 sheet；断言以满足类型检查
    ws.title = "反馈意见"
    ws.append(_COLUMNS)
    for c in ws[1]:
        c.font = Font(bold=True)

    for fr in result.files:
        ordered = sorted(fr.issues, key=lambda i: (_SEV_ORDER[i.severity], i.row or 0))
        for i in ordered:
            ws.append([
                i.batch, i.file, i.sheet, i.row or "",
                i.field_code or "", i.field_name or "",
                i.rule_id, _SEV_CN[i.severity], i.message,
                i.current_value or "", i.suggestion or "",
            ])
    # 列宽（粗调）
    for col, width in zip("ABCDEFGHIJK", (10, 40, 24, 8, 12, 14, 8, 8, 50, 30, 24), strict=True):
        ws.column_dimensions[col].width = width

    # 联网核对(L5)：review 统一入口传入时增一个 sheet；check/import 不传则只有「反馈意见」sheet（向后兼容）
    if audit_reports:
        audit_rows = sorted(audit_reports, key=_triage_key)
        ws2 = wb.create_sheet("联网核对")
        ws2.append(["资源项码", "奖项", "年份", "结论", "置信", "分诊", "官网条数", "提交行数",
                    "官网年份", "疑漏采", "官网未匹配", "来源", "人工入口URL", "降级原因", "备注"])
        for c in ws2[1]:
            c.font = Font(bold=True)
        for r in audit_rows:
            ws2.append([
                r.get("resource_code", ""), r.get("award_name", ""), r.get("year", ""),
                r.get("verdict", ""), r.get("confidence", ""), triage_label(_triage_of(r)),
                r.get("extracted_count", 0), r.get("submitted_count", 0), r.get("page_year", ""),
                "；".join(r.get("missing", [])), "；".join(r.get("extra", [])),
                r.get("source_kind", ""), "；".join(r.get("source_urls", [])),
                _reasons_of(r), r.get("notes", ""),
            ])
        widths = (10, 20, 6, 16, 6, 10, 8, 8, 8, 30, 30, 8, 40, 30, 30)
        for col, width in zip("ABCDEFGHIJKLMNO", widths, strict=True):
            ws2.column_dimensions[col].width = width
    if audit_cases:
        ws3 = wb.create_sheet("疑难案件")
        ws3.append([
            "案件ID", "资源项码", "奖项", "年份", "状态", "触发", "置信", "建议",
            "证据来源", "证据哈希", "获取时间", "Verifier", "人工决定", "人工摘要",
            "复核人", "版本",
        ])
        for cell in ws3[1]:
            cell.font = Font(bold=True)
        for case in audit_cases:
            ws3.append([
                case.get("case_id", ""),
                case.get("resource_code", ""),
                case.get("award_name", ""),
                case.get("year", ""),
                case.get("status", ""),
                "；".join(case.get("trigger_codes", [])),
                case.get("confidence", ""),
                case.get("recommendation", ""),
                "；".join(case.get("evidence_sources", [])),
                "；".join(case.get("evidence_hashes", [])),
                "；".join(case.get("evidence_times", [])),
                case.get("verification_action", ""),
                case.get("human_decision", ""),
                case.get("human_decision_summary", ""),
                case.get("reviewed_by", ""),
                case.get("state_version", ""),
            ])
        case_widths = (8, 10, 20, 8, 12, 24, 8, 36, 42, 36, 24, 14, 12, 36, 16, 8)
        for col, width in zip("ABCDEFGHIJKLMNOP", case_widths, strict=True):
            ws3.column_dimensions[col].width = width
    wb.save(out_path)


# 写反馈摘要.md：批次概览 + Top 问题类型；audit_reports 非空时追加「联网核对(L5)」段
def write_markdown_summary(
    result: BatchResult,
    out_path: Path,
    audit_reports: list[dict[str, Any]] | None = None,
    audit_cases: list[dict[str, Any]] | None = None,
) -> None:
    lines: list[str] = []
    lines.append(f"# 反馈摘要 —— 批次 {result.batch}\n")
    lines.append(f"- 文件数：{len(result.files)}")
    lines.append(f"- 问题总数：{result.total_issues}")
    lines.append(f"- 严重(blocker)：{result.count(Severity.BLOCKER)} ｜ "
                 f"格式(format)：{result.count(Severity.FORMAT)} ｜ 待复核(review)：{result.count(Severity.REVIEW)}\n")

    lines.append("## 各文件概览\n")
    lines.append("| 文件 | 模板 | 数据行 | 严重 | 格式 | 待复核 | 判定 |")
    lines.append("|---|---|---:|---:|---:|---:|---|")
    for fr in result.files:
        lines.append(f"| {fr.file} | {fr.claimed_table_code} | {fr.n_rows} | "
                     f"{fr.count(Severity.BLOCKER)} | {fr.count(Severity.FORMAT)} | "
                     f"{fr.count(Severity.REVIEW)} | {fr.verdict} |")

    # Top 问题类型（按规则）
    rule_counter: Counter[str] = Counter(i.rule_id for fr in result.files for i in fr.issues)
    if rule_counter:
        lines.append("\n## Top 问题类型\n")
        lines.append("| 规则ID | 查什么 | 命中次数 |")
        lines.append("|---|---|---:|")
        for rule_id, cnt in rule_counter.most_common():
            title = RULES[rule_id].title if rule_id in RULES else ""
            lines.append(f"| {rule_id} | {title} | {cnt} |")

    # 联网核对(L5)段：review 统一入口传入 audit_reports 时追加；check/import 不传则与旧产出逐字节一致
    if audit_reports:
        ordered = sorted(audit_reports, key=_triage_key)
        tri_counts = Counter(_triage_of(r) for r in audit_reports)  # 分诊统计各桶
        lines.append("\n## 联网核对(L5)\n")
        lines.append(f"- 已核对资源项：{len(audit_reports)} 个")
        lines.append(f"- 分诊：{triage_label('auto_pass')} {tri_counts.get('auto_pass', 0)} ｜ "
                     f"{triage_label('review')} {tri_counts.get('review', 0)} ｜ "
                     f"{triage_label('manual')} {tri_counts.get('manual', 0)}\n")
        lines.append("| 资源项码 | 奖项 | 年份 | 结论 | 置信 | 分诊 | 官网条数 | 提交行数 |")
        lines.append("|---|---|---|---|---|---|---:|---:|")
        for r in ordered:
            tri = triage_label(_triage_of(r))
            lines.append(f"| {r.get('resource_code', '')} | {r.get('award_name', '')} | {r.get('year', '')} | "
                         f"{r.get('verdict', '')} | {r.get('confidence', '')} | {tri} | "
                         f"{r.get('extracted_count', 0)} | {r.get('submitted_count', 0)} |")
        for r in ordered:  # 逐条明细：仅在有降级原因/疑漏/疑多/人工入口时列
            detail: list[str] = []
            reasons = _reasons_of(r)
            if reasons:
                detail.append(f"  - 降级原因：{reasons}")
            if r.get("missing"):
                detail.append(f"  - 疑漏采：{r['missing']}")
            if r.get("extra"):
                detail.append(f"  - 官网未匹配：{r['extra']}")
            for u in r.get("source_urls", []):
                detail.append(f"  - 人工核对官网：{u}")
            for a in (r.get("found_assets") or [])[:5]:
                detail.append(f"  - 名单文件/图片：{a}")
            if detail:
                lines.append(f"\n**{r.get('award_name', '')}（{r.get('resource_code', '')}）** "
                             f"[{triage_label(_triage_of(r))}]")
                lines.extend(detail)

    if audit_cases:
        status_counts = Counter(str(case.get("status", "")) for case in audit_cases)
        lines.append("\n## M5 疑难案件\n")
        lines.append(
            f"- 案件数：{len(audit_cases)} ｜ queued {status_counts.get('queued', 0)} ｜ "
            f"running {status_counts.get('running', 0)} ｜ "
            f"waiting_human {status_counts.get('waiting_human', 0)} ｜ "
            f"completed {status_counts.get('completed', 0)} ｜ "
            f"failed {status_counts.get('failed', 0)}\n"
        )
        lines.append("| 案件 | 资源项码 | 奖项 | 状态 | 触发 | Verifier | 人工决定 | 复核人 |")
        lines.append("|---:|---|---|---|---|---|---|---|")
        for case in audit_cases:
            lines.append(
                f"| {case.get('case_id', '')} | {case.get('resource_code', '')} | "
                f"{case.get('award_name', '')} | {case.get('status', '')} | "
                f"{'/'.join(case.get('trigger_codes', []))} | "
                f"{case.get('verification_action', '') or '未执行'} | "
                f"{case.get('human_decision', '') or '未终审'} | "
                f"{case.get('reviewed_by', '') or '-'} |"
            )
        for case in audit_cases:
            sources = case.get("evidence_sources", [])
            hashes = case.get("evidence_hashes", [])
            if sources or hashes or case.get("recommendation"):
                lines.append(
                    f"\n**案件 #{case.get('case_id', '')} · {case.get('award_name', '')}**"
                )
                if case.get("recommendation"):
                    lines.append(f"  - 自动建议：{case['recommendation']}")
                for source in sources:
                    lines.append(f"  - 证据来源：{source}")
                for digest in hashes:
                    lines.append(f"  - SHA-256：{digest}")
                if not case.get("human_decision"):
                    lines.append("  - 人工归宿：待复核台终审")

    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


# 一次产出两份报告，返回 (excel路径, md路径)；audit_reports 非空时（review 统一入口）两份里都追加 L5 联网核对
def write_reports(
    result: BatchResult,
    out_dir: Path | None = None,
    audit_reports: list[dict[str, Any]] | None = None,
    audit_cases: list[dict[str, Any]] | None = None,
) -> tuple[Path, Path]:
    d = out_dir or config.out_dir()
    xlsx = d / f"反馈意见-{result.batch}.xlsx"
    md = d / f"反馈摘要-{result.batch}.md"
    write_excel_report(result, xlsx, audit_reports, audit_cases)
    write_markdown_summary(result, md, audit_reports, audit_cases)
    return xlsx, md
